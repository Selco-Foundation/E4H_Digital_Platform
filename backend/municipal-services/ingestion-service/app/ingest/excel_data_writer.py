import pandas as pd

from app.ingest.service.data_writer import DataWriter


class ExcelDataWriter(DataWriter):
    def __init__(self, file_path: str, output_sheet: str = "Vendor Output"):
        self.file_path = file_path
        self.output_sheet = output_sheet

    def write_data(self, data: pd.DataFrame) -> bool:
        try:
            # Load existing workbook to preserve other sheets
            from openpyxl import load_workbook
            book = load_workbook(self.file_path)

            # Remove the existing output sheet if it exists
            if self.output_sheet in book.sheetnames:
                idx = book.sheetnames.index(self.output_sheet)
                book.remove(book.worksheets[idx])

            # Save the workbook to preserve changes
            book.save(self.file_path)

            # Now write the results to the output sheet
            with pd.ExcelWriter(self.file_path, engine='openpyxl', mode='a') as writer:
                data.to_excel(writer, sheet_name=self.output_sheet, index=False)

            print(f"Updated '{self.output_sheet}' sheet in {self.file_path}")
            return True
        except Exception as e:
            print(f"Error writing data: {str(e)}")
            return False