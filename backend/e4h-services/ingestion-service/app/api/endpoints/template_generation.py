from datetime import datetime
from typing import Optional, List
import psycopg2

import pandas as pd
from PIL import ImageDraw, Image, ImageFont
from fastapi import APIRouter, Form, HTTPException, Depends, Body
from fastapi.responses import FileResponse
from fastapi import BackgroundTasks
from openpyxl.utils import get_column_letter

from app.core.logging import AppLogger
from app.decorators.rbac_validator import get_authorized_request_info
from app.ingest.facility_template_service import FacilityTemplateService
from app.ingest.project_service import ProjectService
from app.schemas.boundary import Boundary, flatten_boundaries
from app.utils.convertor import request_info_from_json
from app.utils.excel_utils import add_dropdowns_to_excel
from app.utils.facility_service_client import FacilityServiceClient
from app.utils.file_utils import create_temp_file, cleanup_temp_file
from app.utils.mdms_client import MDMSClient
from app.utils.project_service_client import ProjectServiceClient
import os, tempfile, zipfile, qrcode, shutil

router = APIRouter()
logger = AppLogger().get_logger()

from dotenv import load_dotenv

load_dotenv()
mdms_url = os.getenv("MDMS_URL")
project_service_url = os.getenv("PROJECT_SERVICE_URL")
facility_service_url = os.getenv("FACILITY_SERVICE_URL")
DB_CONFIG = {
    "host": os.getenv("DB_HOST"),
    "port": int(os.getenv("DB_PORT", 5432)),
    "database": os.getenv("DB_NAME"),
    "user": os.getenv("DB_USER"),
    "password": os.getenv("DB_PASSWORD")
}

@router.post('/facilityIngestionTemplateWithData',
            summary='Generate facility ingestion template Excel file with schema, already present data and boundary codes',
            response_description="Returns Excel template with facility schema, facility data and boundary codes")
async def get_facility_ingestion_template_with_data(
        background_tasks: BackgroundTasks,
        facility_service: FacilityTemplateService = Depends(),
    payload: dict = Body(..., description="Payload object")
):
    request_info = request_info_from_json(payload.get("request_info", {}))
    boundary_data = payload.get("boundary_data", {})
    project_id = payload.get("project_id")
    mdms_client = MDMSClient(mdms_url)
    try:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"facility_ingestion_template_{timestamp}.xlsx"
        output_file_path = create_temp_file(suffix=".xlsx")
        try:
            facility_schema = mdms_client.get_column_definitions_with_metadata(request_info, 'data-ingestion.FacilityIngestionSchema')
            boundary_list: List[Boundary] = flatten_boundaries(boundary_data)
        except Exception as e:
            logger.error(f"Error fetching data from external services: {e}")
            cleanup_temp_file(output_file_path)
            raise HTTPException(status_code=502, detail=f"External service error: {str(e)}")

        all_facilities = []
        if facility_service_url:
            facility_client = FacilityServiceClient(facility_service_url)
            for boundary in boundary_list:
                try:
                    results = facility_client.search_facility(tenant_id='in', boundary_code=boundary.code)
                    facilities = results.get('facilities', [])
                    all_facilities.extend(facilities)
                except Exception as e:
                    print(f"Error fetching boundary facilities: {e}")

        # Fetch project-linked facilities if project_id is provided
        project_linked_facility_ids = set()
        if project_id and project_service_url:
            try:
                project_client = ProjectServiceClient(project_service_url)
                project_facilities_response = project_client.search_project_facility(request_info, project_id)
                project_facilities = project_facilities_response.get("ProjectFacilities", [])
                project_linked_facility_ids = {pf.get("facilityId") for pf in project_facilities if pf.get("facilityId")}
                logger.info(f"Found {len(project_linked_facility_ids)} facilities linked to project {project_id}")
            except Exception as e:
                logger.error(f"Error fetching project facilities: {e}")
                # Continue without project facility data if there's an error

        # Mark facilities as included in project if they are already linked
        if project_id:
            for facility in all_facilities:
                facility_id = facility.get("facility_id")
                if facility_id in project_linked_facility_ids:
                    facility["include_in_project"] = "Yes"
                    logger.info(f"Facility {facility_id} is linked to project - marking as Yes")
                else:

                    facility["include_in_project"] = "No"
                    logger.info(f"Facility {facility_id} is NOT linked to project - marking as No")
        else:
            # If no project_id provided, set all facilities to "No"
            for facility in all_facilities:
                facility["include_in_project"] = "No"
                logger.info(f"No project_id provided - marking facility {facility.get('facility_id')} as No")

        try:
            facility_service.generate_template_file_with_data(
                output_path=output_file_path,
                facility_schema=facility_schema,
                boundary_list=boundary_list,
                facility_data=all_facilities,
                project_id=project_id
            )
            logger.info(f"Successfully created facility ingestion template at {output_file_path}")
        except Exception as e:
            logger.error(f"Error generating template file: {e}")
            cleanup_temp_file(output_file_path)
            raise HTTPException(status_code=500, detail=f"Template generation error: {str(e)}")
        background_tasks.add_task(cleanup_temp_file, output_file_path)
        return FileResponse(
            path=output_file_path,
            filename=output_filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        logger.error(f"Unhandled error in get_facility_ingestion_template: {e}")
        raise HTTPException(status_code=500, detail=f"An unexpected error occurred: {str(e)}")


@router.get('/facilityIngestion',
            summary='Generate facility ingestion template Excel file with schema and boundary codes',
            response_description="Returns Excel template with facility schema and boundary codes")
async def get_facility_ingestion_template(
        facility_service: FacilityTemplateService = Depends(),
        request_info: str = Form(default="")
):
    request_info = request_info_from_json(request_info)
    get_authorized_request_info(request_info)
    mdms_client = MDMSClient(mdms_url)
    try:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"facility_ingestion_template_{timestamp}.xlsx"
        output_file_path = create_temp_file(suffix=".xlsx")
        try:
            facility_schema = mdms_client.get_column_definitions_with_metadata(request_info, 'data-ingestion.FacilityIngestionSchema')
            boundary_data = facility_service.get_all_boundaries(request_info)
            vendor_data = facility_service.get_all_vendor_codes(request_info)
        except Exception as e:
            logger.error(f"Error fetching data from external services: {e}")
            cleanup_temp_file(output_file_path)
            raise HTTPException(status_code=502, detail=f"External service error: {str(e)}")

        try:
            facility_service.generate_template_file(
                output_path=output_file_path,
                facility_schema=facility_schema,
                boundary_data=boundary_data,
                vendor_data=vendor_data
            )
            logger.info(f"Successfully created facility ingestion template at {output_file_path}")
        except Exception as e:
            logger.error(f"Error generating template file: {e}")
            cleanup_temp_file(output_file_path)
            raise HTTPException(status_code=500, detail=f"Template generation error: {str(e)}")

        return FileResponse(
            path=output_file_path,
            filename=output_filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        logger.error(f"Unhandled error in get_facility_ingestion_template: {e}")
        raise HTTPException(status_code=500, detail=f"An unexpected error occurred: {str(e)}")

@router.post('/facilityWithStaff',
            summary='Generate facility ingestion template with staff Excel file',
            response_description="Returns Excel template with facility schema")
async def get_facility_ingestion_template_with_staff(
        parent_id: str = Form(..., description="Parent project ID"),
        request_info: str = Form(..., description="Serialized RequestInfo JSON")
):
    temp_dir = tempfile.gettempdir()
    ts = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    output_filename = f"facility_staff_template_{parent_id}_{ts}.xlsx"
    output_file_path = os.path.join(temp_dir, output_filename)

    try:
        request_info = request_info_from_json(request_info)
        get_authorized_request_info(request_info)

        project_service = ProjectService()
        facilities = project_service.get_facilities(request_info, parent_id, "Staff")
        facility_template_service = FacilityTemplateService()

        try:
            original_df = pd.DataFrame(facilities)
            df = facility_template_service.add_supervisor_columns_to_dataframe(original_df)

            with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Facilities_Staff')
                worksheet = writer.sheets['Facilities_Staff']
                for i, col in enumerate(df.columns):
                    column_width = max(df[col].astype(str).map(len).max(), len(col)) + 2
                    worksheet.column_dimensions[get_column_letter(i + 1)].width = column_width

            dropdowns_map = {'Role (Mandatory) ?': ['Staff', 'Field Planner']}
            add_dropdowns_to_excel(
                file_path=output_file_path,
                sheet_name="Facilities_Staff",
                dropdowns=dropdowns_map
            )

        except Exception as e:
            logger.error(f"Error generating template file: {e}")
            cleanup_temp_file(output_file_path)
            raise HTTPException(status_code=500, detail=f"Template generation error: {str(e)}")

        return FileResponse(
            path=output_file_path,
            filename=output_filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        logger.error(f"Unhandled error in get_facility_ingestion_template_with_staff: {e}")
        cleanup_temp_file(output_file_path)
        raise HTTPException(status_code=500, detail=f"An unexpected error occurred: {str(e)}")

@router.post('/facilityWithSupervisors',
            summary='Generate facility ingestion template with supervisors Excel file',
            response_description="Returns Excel template with facility schema")
async def get_facility_ingestion_template_with_supervisors(
        parent_id: str = Form(..., description="Parent project ID"),
        request_info: str = Form(..., description="Serialized RequestInfo JSON")
):
    temp_dir = tempfile.gettempdir()
    ts = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    output_filename = f"facility_supervisors_template_{parent_id}_{ts}.xlsx"
    output_file_path = os.path.join(temp_dir, output_filename)

    try:
        request_info = request_info_from_json(request_info)
        get_authorized_request_info(request_info)

        project_service = ProjectService()
        facilities = project_service.get_facilities(request_info, parent_id, "Supervisor")
        facility_template_service = FacilityTemplateService()

        try:
            original_df = pd.DataFrame(facilities)
            df = facility_template_service.add_supervisor_columns_to_dataframe(original_df)

            with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Facilities_Supervisors')
                worksheet = writer.sheets['Facilities_Supervisors']
                for i, col in enumerate(df.columns):
                    column_width = max(df[col].astype(str).map(len).max(), len(col)) + 2
                    worksheet.column_dimensions[get_column_letter(i + 1)].width = column_width

            dropdowns_map = {'Role (Mandatory) ?': ['Supervisor', 'Field Planner']}
            add_dropdowns_to_excel(
                file_path=output_file_path,
                sheet_name="Facilities_Supervisors",
                dropdowns=dropdowns_map
            )

        except Exception as e:
            logger.error(f"Error generating template file: {e}")
            cleanup_temp_file(output_file_path)
            raise HTTPException(status_code=500, detail=f"Template generation error: {str(e)}")

        return FileResponse(
            path=output_file_path,
            filename=output_filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        logger.error(f"Unhandled error in get_facility_ingestion_template_with_supervisors: {e}")
        cleanup_temp_file(output_file_path)
        raise HTTPException(status_code=500, detail=f"An unexpected error occurred: {str(e)}")

@router.get('/facilitySelection',
            summary='Generate facility selection template Excel file',
            response_description="Returns Excel template with facility data")
async def get_facility_selection_template(
        facility_service: FacilityTemplateService = Depends(),
        parent_project_id: Optional[str] = Form(default=None),
        boundary_codes: str = Form(...),
        request_info: str = Form(default="")
):
    request_info = request_info_from_json(request_info)
    get_authorized_request_info(request_info)
    mdms_client = MDMSClient(mdms_url)

    boundary_code_list: List[str] = [code.strip() for code in boundary_codes.split(",") if code.strip()]

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_filename = f"facility_selection_template_{timestamp}.xlsx"
    output_file_path = create_temp_file(suffix=".xlsx")

    boundary_facilities = []
    project_facilities = []

    facility_client = None

    try:
        facility_selection_schema = mdms_client.fetch_facility_selection_schema(request_info=request_info)
    except Exception as e:
        logger.error(f"Error fetching data from external services: {e}")
        cleanup_temp_file(output_file_path)
        raise HTTPException(status_code=502, detail=f"External service error: {str(e)}")

    if facility_service_url:
        facility_client = FacilityServiceClient(facility_service_url)
        for boundary_code in boundary_code_list:
            try:
                results = facility_client.search_facility(tenant_id='in', boundary_code=boundary_code)
                boundary_facilities.extend(results.get('facilities', []))
            except Exception as e:
                print(f"Error fetching boundary facilities: {e}")

    if project_service_url and parent_project_id:
        project_client = ProjectServiceClient(project_service_url)
        try:
            pf_response = project_client.search_project_facility(
                request_info=request_info,
                project_id=parent_project_id
            )
            raw_project_facilities = pf_response.get("ProjectFacilities", [])
            if raw_project_facilities and facility_client:
                for pf in raw_project_facilities:
                    facility_id = pf.get("facilityId")
                    if facility_id and any(f.get('facility_id') == facility_id for f in boundary_facilities):
                        try:
                            facility_data = facility_client.search_facility(tenant_id='in', facility_id=facility_id)
                            if facility_data:
                                project_facilities.extend(facility_data.get('facilities', []))
                        except Exception as e:
                            print(f"Error fetching facility {facility_id}: {e}")
        except Exception as e:
            print(f"Error fetching project facilities: {e}")

    # Intersect by facility_id
    if parent_project_id:
        intersected_facilities = project_facilities
    else:
        intersected_facilities = boundary_facilities

    try:
        facility_service.generate_selection_template_file(
            output_path=output_file_path,
            facility_selection_schema=facility_selection_schema,
            facility_data=intersected_facilities
        )
        logger.info(f"Successfully created facility ingestion template at {output_file_path}")
    except Exception as e:
        logger.error(f"Error generating template file: {e}")
        cleanup_temp_file(output_file_path)
        raise HTTPException(status_code=500, detail=f"Template generation error: {str(e)}")

    return FileResponse(
        path=output_file_path,
        filename=output_filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@router.post('/facilityQRGeneration', summary='Generate QR for facility',
             response_description="Returns zip with QR codes")
async def get_facility_QR_for_autologin(
        request_info: str = Form(default="")
):
    request_info_obj = request_info_from_json(request_info)

    try:
        mdms_client = MDMSClient(mdms_url=mdms_url)
        mdms_content = mdms_client.get_tenant_mapping(request_info_obj, ["as", "gj", "ml", "mn", "mz", "nl", "or", "pg", "sk"])


        base_url = "https://saura-emitra-uat.selcofoundation.org"
        password = "Health@2026"

        temp_dir = tempfile.mkdtemp()

        for tenant_id, health_facility_data in mdms_content.items():

            state = health_facility_data.get("address")
            district = health_facility_data.get("city", {}).get("districtName")
            block = health_facility_data.get("city", {}).get("blockCode")
            facility_name = health_facility_data.get("name")

            if not all([state, district, block, facility_name]):
                print(f"Skipping tenant {tenant_id}: Missing state/district/block/facility_name.")
                continue

            qr_folder = os.path.join(temp_dir, state, district, block, facility_name)
            os.makedirs(qr_folder, exist_ok=True)

            conn = psycopg2.connect(**DB_CONFIG)
            with conn.cursor() as cursor:
                sql = "SELECT code FROM eg_hrms_employee WHERE tenantid = %s"
                cursor.execute(sql, (health_facility_data["code"],))
                rows = cursor.fetchall()

            if not rows:
                print(f"Skipping tenant {tenant_id} ({facility_name}): No HRMS employee code found for tenantid {health_facility_data['code']}")
                continue

            username = rows[0][0]


            if state=='Karnataka':
                url_state_name = 'digit-ui'
            else:
                url_state_name = state.lower()

            login_url = f"{base_url}/{url_state_name}/employee/user/login?tenantid={tenant_id}&username={username}&passwd={password}"

            qr = qrcode.make(login_url).convert("RGB")

            # Create a new image (taller) to hold QR and text
            width, height = qr.size
            font_size = 30
            padding = 10

            try:
                font = ImageFont.truetype("DejaVuSans-Bold.ttf", font_size)
            except:
                font = ImageFont.load_default()

            # Create a new image with extra space for text
            new_height = height + font_size + 2 * padding
            combined = Image.new("RGB", (width, new_height), "white")
            combined.paste(qr, (0, 0))

            # Draw the facility name
            draw = ImageDraw.Draw(combined)
            text = facility_name
            bbox = draw.textbbox((0, 0), text, font=font)
            text_width = bbox[2] - bbox[0]
            text_position = ((width - text_width) // 2, height + padding)
            draw.text(text_position, text, font=font, fill="black")

            # Save the combined image
            qr_filename = f"{username}.png"
            img_path = os.path.join(qr_folder, qr_filename)
            combined.save(img_path)

        # Create ZIP
        zip_path = os.path.join(tempfile.gettempdir(), "facility_qr_codes.zip")
        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zipf:
            for root, _, files in os.walk(temp_dir):
                for file in files:
                    abs_file = os.path.join(root, file)
                    rel_path = os.path.relpath(abs_file, temp_dir)
                    zipf.write(abs_file, arcname=rel_path)

        shutil.rmtree(temp_dir)

        return FileResponse(
            path=zip_path,
            filename="facility_qr_codes.zip",
            media_type="application/zip"
        )
    except Exception as e:
        return {"status": "error", "message": str(e)}


