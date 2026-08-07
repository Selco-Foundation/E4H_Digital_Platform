import io
import json
import math
import os
import shutil
import tempfile
import zipfile
from datetime import datetime
import uuid
from typing import Optional, Dict, List, Set

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Protection, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

from app.utils.amc_scheduler_service_client import AMCSchedulerServiceClient
from app.utils.excel_utils import (
    FACILITY_IDENTIFIER_COLUMNS,
    autofit_columns,
    normalize_excel_integer_columns,
    prepare_dataframe_for_excel_export,
)
from app.utils.facility_validator import (
    project_facility_validation,
    facility_validation,
    field_plan_facility_validation,
    collect_hfr_nin_errors_for_row,
    collect_anganwadi_poc_username_errors_for_row,
)
from fastapi import APIRouter, File, Form, UploadFile, HTTPException, BackgroundTasks, Depends
from fastapi.responses import FileResponse
import psycopg2
from starlette.responses import JSONResponse, StreamingResponse
import requests

from app.core.logging import AppLogger
from app.decorators.rbac_validator import get_authorized_request_info
from app.ingest.excel_data_writer import ExcelDataWriter
from app.processor.factory.boundary_data_processor_factory import BoundaryDataProcessorFactory
from app.processor.factory.vendor_data_processor_factory import VendorDataProcessorFactory
from app.schemas.request_info import RequestInfo
from app.producer.producer import Producer
from app.utils.convertor import request_info_from_json, create_vendor_request, create_facility_payload, \
    resolve_mapped_vendor_for_facility_row, build_field_plan_facility_bulk_entry, \
    build_field_plan_facility_additional_details, build_field_plan_facility_additional_fields, \
    get_project_creation_payload, check_role_mismatch_for_user_type, get_user_creation_payload_staff, \
    get_user_creation_payload_supervisors, \
    get_staff_creation_payload, create_project_payload, get_installation_spoc_creation_payload, \
    get_staff_search_payload, create_update_payload, get_incident_request_info, \
    resolve_boundary_codes_for_dataframe
from app.utils.boundary_service_client import BoundaryServiceClient
from app.utils.facility_service_client import FacilityServiceClient
from app.utils.fieldplan_activity_service_client import FieldPlanActivityServiceClient
from app.utils.fieldplan_service_client import FieldPlanServiceClient
from app.utils.assessment_service_client import AssessmentServiceClient
from app.utils.assessment_fieldplan_handoff import (
    extract_assessment_link_meta,
    load_eligible_facility_map,
    merge_assessment_validation_errors,
    parse_assessment_plan_ids,
    validate_assessment_handoff_rows,
)
from app.utils.icc_report_converter import validate_and_convert, ICCValidationError, SYSTEM_TYPE_TO_INTERNAL
from app.utils.file_utils import cleanup_temp_file
from app.utils.im_service_client import IMServiceClient
from app.utils.mdms_client import MDMSClient
from app.utils.organization_service_client import OrganizationServiceClient
from app.utils.project_service_client import ProjectServiceClient
from app.utils.hrms_service_client import HRMSServiceClient

router = APIRouter()
logger = AppLogger().get_logger()

from dotenv import load_dotenv
from collections import defaultdict


ALLOWED_EXCEL_EXTENSIONS = {"xlsx", "xls"}


def _has_allowed_excel_extension(filename: Optional[str]) -> bool:
    if not filename or "." not in filename:
        return False
    return filename.rsplit(".", 1)[-1].strip().lower() in ALLOWED_EXCEL_EXTENSIONS


async def _save_upload_to_temp_file(upload_file: UploadFile, suffix: str = ".xlsx", chunk_size: int = 1024 * 1024):
    """
    Persist an UploadFile to disk in chunks to avoid loading
    large uploads entirely into memory.
    """
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
    total_bytes = 0
    try:
        while True:
            chunk = await upload_file.read(chunk_size)
            if not chunk:
                break
            temp_file.write(chunk)
            total_bytes += len(chunk)
    finally:
        temp_file.close()
    return temp_file, total_bytes


VALIDATION_ERRORS_SHEET = "Validation Errors"


def _write_validation_errors_sheet(xlsx_path: str, error_message: str) -> int:
    """Replaces the `VALIDATION_ERRORS_SHEET` sheet in the workbook at xlsx_path with one row per
    line of error_message, and saves in place. A pre-existing sheet from an earlier failed attempt
    is dropped first, so re-uploading a previously-annotated file always reflects only the latest
    validation run, never a stale/appended mix of old and new errors. All other sheets/data are
    untouched. Splitting on newline (rather than the "\n- " bullet marker Validations 3/4 use)
    handles both the single-sentence messages from Validations 1/2 and the multi-line bulleted
    messages from Validations 3/4 uniformly, with no special-casing.

    Returns the number of distinct errors represented by error_message: the number of "- "-prefixed
    bullet lines for a multi-line Validation 3/4 message, or 1 for a single-sentence Validation 1/2
    message that has no bullets.
    """
    wb = load_workbook(xlsx_path)
    if VALIDATION_ERRORS_SHEET in wb.sheetnames:
        del wb[VALIDATION_ERRORS_SHEET]
    ws = wb.create_sheet(VALIDATION_ERRORS_SHEET)
    ws.append(["Validation Error"])
    ws["A1"].font = Font(bold=True)
    lines = error_message.strip().split("\n")
    for line in lines:
        ws.append([line])
    wb.save(xlsx_path)
    autofit_columns(xlsx_path, VALIDATION_ERRORS_SHEET)
    bullet_count = sum(1 for line in lines if line.strip().startswith("- "))
    return bullet_count if bullet_count else 1


load_dotenv()
mdms_url = os.getenv("MDMS_URL")
org_service_url = os.getenv("VENDOR_SERVICE_URL")
project_service_url = os.getenv("PROJECT_SERVICE_URL")
fieldPlan_service_url = os.getenv("FIELDPLAN_SERVICE_URL")
fieldPlan_activity_service_url = os.getenv("FIELDPLAN_ACTIVITY_SERVICE_URL")
facility_service_url = os.getenv("FACILITY_SERVICE_URL")
hrms_service_url = os.getenv("HRMS_SERVICE_URL")
im_services_url = os.getenv("IM_SERVICES_URL")
amc_scheduler_service_url = os.getenv("AMC_SCHEDULER_SERVICE_URL")
localization_service_url = os.getenv("LOCALIZATION_SERVICE_URL")
boundary_service_url = os.getenv("BOUNDARY_SERVICE_URL")
DEFAULT_AMC_ASSET_TYPES = ["INVERTER", "PANEL", "BATTERY"]
BULK_INGEST_CHUNK_SIZE = 200
AMC_CONFIGURATION_BULK_CHUNK_SIZE = 400
ENVIRONMENT = os.getenv("ENVIRONMENT", "uat").lower()
base_path = os.path.dirname(os.path.abspath(__file__))
config_path = os.path.abspath(os.path.join(base_path, "..", "..", "config"))

with open(os.path.join(config_path, "tenant_creator_mapping.json"), 'r') as f:
    TENANT_CREATOR_MAPPING = json.load(f).get(ENVIRONMENT, {})

with open(os.path.join(config_path, "user_profiles.json"), 'r') as f:
    USER_PROFILE = json.load(f).get(ENVIRONMENT, {})

DB_CONFIG = {
    "host": os.getenv("DB_HOST"),
    "port": int(os.getenv("DB_PORT", 5432)),
    "database": os.getenv("DB_NAME"),
    "user": os.getenv("DB_USER"),
    "password": os.getenv("DB_PASSWORD")
}

FACILITY_VENDOR_CODE_COLUMN = "Vendor Code (Mandatory)"


def _bulk_create_facilities_for_ingestion(
        df: pd.DataFrame,
        facility_client: FacilityServiceClient,
        org_client: Optional[OrganizationServiceClient],
        request_info: RequestInfo,
        facility_schema: List[Dict],
        are_facilities_onm_ready: bool,
        vendor_mapping_cache: Dict[str, Dict[str, Optional[str]]],
        hfr_nin_db_cache: Dict[str, bool],
) -> None:
    """Validate rows and bulk-create facilities in chunks (vendor jurisdictions handled by facility-service)."""
    pending_creates: List[tuple] = []

    for index, row in df[df['status'] != 'success'].iterrows():
        hfr_nin_errs = collect_hfr_nin_errors_for_row(
            row, index, df, facility_client, hfr_nin_db_cache,
        )
        anganwadi_poc_errs = collect_anganwadi_poc_username_errors_for_row(
            row, index, df, facility_schema,
        )
        pre_errs = list(dict.fromkeys([*hfr_nin_errs, *anganwadi_poc_errs]))
        if pre_errs:
            df.at[index, 'status'] = 'failed'
            df.at[index, 'error'] = '; '.join(pre_errs)
            continue

        pending_creates.append((index, row.copy()))

    if not pending_creates:
        return

    logger.info("Processing %s facilities using bulk create API", len(pending_creates))
    request_info_payload = request_info.model_dump(by_alias=True, exclude_none=True)

    for chunk_start in range(0, len(pending_creates), BULK_INGEST_CHUNK_SIZE):
        chunk = pending_creates[chunk_start:chunk_start + BULK_INGEST_CHUNK_SIZE]
        bulk_payload = {
            "RequestInfo": request_info_payload,
            "facilities": [],
        }
        creation_meta: List[int] = []

        for index, row_data in chunk:
            vendor_mapping = resolve_mapped_vendor_for_facility_row(
                org_client,
                request_info,
                row_data,
                FACILITY_VENDOR_CODE_COLUMN,
                vendor_mapping_cache,
            )
            single_payload = create_facility_payload(
                request_info,
                row_data,
                are_facilities_onm_ready,
                facility_schema,
                mapped_vendor_name=vendor_mapping.get("mappedVendorName"),
                mapped_vendor_user_name=vendor_mapping.get("mappedVendorUserName"),
            )
            facilities = single_payload.get("facilities", [])
            if facilities:
                bulk_payload["facilities"].append(facilities[0])
                creation_meta.append(index)
            else:
                df.at[index, 'status'] = 'failed'
                df.at[index, 'error'] = 'Invalid facility payload'

        if not bulk_payload["facilities"]:
            continue

        create_resp = None
        try:
            create_resp = facility_client.create_facility(bulk_payload)
        except Exception as exc:
            for index in creation_meta:
                df.at[index, 'status'] = 'failed'
                df.at[index, 'error'] = f'Exception during bulk create: {str(exc)}'
            continue

        if create_resp.status_code in (200, 201):
            for row_idx in creation_meta:
                df.at[row_idx, 'status'] = 'success'
                df.at[row_idx, 'error'] = ''
        elif create_resp.status_code == 400:
            try:
                error_data = create_resp.json()
                error_message = error_data.get('Errors', [{}])[0].get('message', 'Unknown error')
            except Exception:
                error_message = create_resp.text
            for index in creation_meta:
                df.at[index, 'status'] = 'failed'
                df.at[index, 'error'] = error_message
        else:
            for index in creation_meta:
                df.at[index, 'status'] = 'failed'
                df.at[index, 'error'] = f'{create_resp.status_code}: {create_resp.text}'


@router.post('/vendors',
             summary='Upload and process vendor Excel file with multiple sheets',
             response_description="Returns processed Excel file with validation results")
async def upload_vendors_excel_sheet(
        vendor_file: UploadFile = File(description="Excel file containing vendor data and boundary codes"),
        vendor_sheet_name: str = Form(default="Vendor Input", description="Name of the sheet containing vendor data"),
        boundary_sheet_name: str = Form(default="Boundary Code",
                                        description="Name of the sheet containing boundary codes"),
        request_info: str = Form(default="")
):
    logger.trace("Starting vendor Excel file upload and processing")
    input_temp_file = None
    output_temp_file = None
    request_info = request_info_from_json(request_info)
    #get_authorized_request_info(request_info)
    logger.info(f"Processing vendor file: vendor_sheet={vendor_sheet_name}, boundary_sheet={boundary_sheet_name}")

    try:
        logger.debug("Creating temporary files for vendor processing")
        input_temp_file, uploaded_size = await _save_upload_to_temp_file(vendor_file, suffix=".xlsx")
        vendor_file_path = input_temp_file.name
        logger.debug(f"Saved uploaded file to: {vendor_file_path}, size: {uploaded_size} bytes")

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"vendor_validation_results_{timestamp}.xlsx"
        output_temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        output_temp_file.close()
        output_file_path = output_temp_file.name

        logger.info("Creating vendor data processor")
        processor = VendorDataProcessorFactory.create_processor(
            file_path=vendor_file_path,
            vendor_sheet=vendor_sheet_name,
            boundary_sheet=boundary_sheet_name,
            mdms_url=mdms_url,
            request_info=request_info
        )
        logger.info("Processing vendor data")
        tuple_vendors = processor.process_data()
        vendors = tuple_vendors[0]
        vendor_df = tuple_vendors[1]
        logger.info(f"Vendor processing completed: {len(vendors)} valid vendors, {len(vendor_df)} total rows")

        if org_service_url and vendors:
            logger.info(f"Creating {len(vendors)} vendors in organization service")
            org_client = OrganizationServiceClient(org_service_url)

            success_count = 0
            for index, vendor in enumerate(vendors):
                logger.trace(f"Creating vendor {index + 1}/{len(vendors)}: {vendor.vendor_name}")
                vendor_payload = create_vendor_request(request_info, vendor)

                try:
                    org_data = org_client.create_vendor(vendor_payload)
                    if org_data and org_data.get("organisations"):
                        vendor_df.at[index, "status"] = "success"
                        vendor_df.at[index, "error"] = None
                        vendor_id = org_data["organisations"][0].get("id")
                        vendor_df.at[index, "vendor_id"] = vendor_id
                        success_count += 1
                        logger.debug(f"Vendor created successfully: {vendor.vendor_name}, id={vendor_id}")
                    else:
                        logger.warning(f"Failed to create vendor: {vendor.vendor_name} - no organization data returned")
                except Exception as e:
                    logger.error(f"Error creating vendor {vendor.vendor_name} in org service: {e}", exc_info=True)
            logger.info(f"Vendor creation completed: {success_count}/{len(vendors)} successful")

        logger.info("Writing processed vendor data to Excel file")
        with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
            vendor_df.to_excel(writer, sheet_name="Vendor Output", index=False)
            boundary_df = pd.read_excel(vendor_file_path, sheet_name=boundary_sheet_name)
            boundary_df.to_excel(writer, sheet_name=boundary_sheet_name, index=False)
        logger.info(f"Vendor processing completed successfully: {output_filename}")

        return FileResponse(
            path=output_file_path,
            filename=output_filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        logger.error(f"Error processing vendor data: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to process vendor data: {str(e)}")

    finally:
        if input_temp_file and os.path.exists(input_temp_file.name):
            os.unlink(input_temp_file.name)


@router.post('/boundaries',
             summary='Upload and process boundary Excel file',
             response_description="Returns processed Excel file with validation results")
async def upload_boundaries_excel_sheet(
        boundary_file: UploadFile = File(description="Excel file containing boundary data"),
        boundary_sheet_name: str = Form(default="Boundary Data",
                                        description="Name of the sheet containing boundary data"),
        request_info: str = Form(default="")
):
    logger.trace("Starting boundary Excel file upload and processing")
    input_temp_file = None
    output_temp_file = None
    request_info = request_info_from_json(request_info)
    logger.info(f"Processing boundary file: boundary_sheet={boundary_sheet_name}")

    try:
        logger.debug("Creating temporary files for boundary processing")
        input_temp_file, uploaded_size = await _save_upload_to_temp_file(boundary_file, suffix=".xlsx")
        boundary_file_path = input_temp_file.name
        logger.debug(f"Saved uploaded file to: {boundary_file_path}, size: {uploaded_size} bytes")

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"boundary_validation_results_{timestamp}.xlsx"
        output_temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        output_temp_file.close()
        output_file_path = output_temp_file.name

        with open(boundary_file_path, 'rb') as src, open(output_file_path, 'wb') as dst:
            dst.write(src.read())

        logger.info("Creating boundary data processor")
        processor = BoundaryDataProcessorFactory.create_processor(
            file_path=output_file_path,
            boundary_sheet=boundary_sheet_name,
            mdms_url=mdms_url,
            request_info=request_info
        )
        logger.info("Processing boundary data")
        boundary_df = processor.process_data()
        logger.info(f"Boundary processing completed: {len(boundary_df)} boundaries processed")

        writer = ExcelDataWriter(output_file_path, output_sheet="Boundary Data")
        writer.write_data(boundary_df)

        error_count = int(
            boundary_df["status"].astype(str).str.strip().str.lower().eq("fail").sum()
        )

        response = FileResponse(
            path=output_file_path,
            filename=output_filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response.headers["X-Error-Count"] = str(error_count)
        return response

    except Exception as e:
        logger.error(f"Error processing boundary data: {e}")
        raise HTTPException(
            status_code=500,
            detail="Failed to process boundary data"
        ) from e


    finally:
        if input_temp_file and os.path.exists(input_temp_file.name):
            os.unlink(input_temp_file.name)


@router.post('/addFacilitiesValidateData',
             summary='Validate add bulk facility Excel file before processing',
             response_description='Returns validation report Excel with PASSED/FAILED rows')
async def validate_facilities_excel_sheet(
        background_tasks: BackgroundTasks,
        facility_file: UploadFile = File(..., description="Excel file containing facility data"),
        facility_sheet_name: str = Form(default="FacilityIngestionTemplate",
                                        description="Name of the sheet containing facility data"),
        boundary_sheet_name: str = Form(default="BlockBoundaryCodes",
                                        description="Name of the sheet containing boundary data"),
        request_info: str = Form(default="")
):
    temp_input_file = None
    request_info_obj = request_info_from_json(request_info)
    mdms_client = MDMSClient(mdms_url)
    facility_client = FacilityServiceClient(facility_service_url)

    try:
        # Save uploaded Excel to a temp file
        temp_input_file, _ = await _save_upload_to_temp_file(facility_file, suffix=".xlsx")

        # Load workbook to preserve everything
        wb = load_workbook(temp_input_file.name)

        # ----------------- Read Boundary Sheet ----------------- #
        if boundary_sheet_name not in wb.sheetnames:
            raise HTTPException(status_code=400, detail=f"Boundary sheet '{boundary_sheet_name}' not found")

        boundary_data_df = pd.read_excel(temp_input_file.name, sheet_name=boundary_sheet_name)

        # ----------------- Read Facility Sheet ----------------- #
        if facility_sheet_name not in wb.sheetnames:
            raise HTTPException(status_code=400, detail=f"Facility sheet '{facility_sheet_name}' not found")

        df = pd.read_excel(temp_input_file.name, sheet_name=facility_sheet_name)
        df.columns = [str(c).strip() for c in df.columns]
        df = df.loc[:, ~df.columns.str.startswith('Unnamed')]
        df = normalize_excel_integer_columns(df, force_columns=FACILITY_IDENTIFIER_COLUMNS)

        # ----------------- Read Facility Column ----------------- #
        if 'Facility Id' not in df.columns:
            raise HTTPException(status_code=400, detail=f"Facility Column in '{facility_sheet_name}' not found")

        if FACILITY_VENDOR_CODE_COLUMN not in df.columns:
            raise HTTPException(
                status_code=400,
                detail=f"Missing mandatory column '{FACILITY_VENDOR_CODE_COLUMN}'. "
                "Facilities cannot be validated without vendor mapping.",
            )

        # Ensure status/error columns exist
        if 'status' not in df.columns:
            df['status'] = ''
        if 'error' not in df.columns:
            df['error'] = ''

        # Resolve boundary codes from State/District/Block via localization
        df = resolve_boundary_codes_for_dataframe(
            df, localization_service_url,
            boundary_code_column='Boundary Code (Mandatory)',
            logger=logger,
        )

        # ----------------- Verify resolved codes via Boundary Service ----------------- #
        resolved_codes = df.loc[
            df['Boundary Code (Mandatory)'].astype(str).str.strip() != '',
            'Boundary Code (Mandatory)'
        ].unique().tolist()

        if resolved_codes and boundary_service_url:
            boundary_client = BoundaryServiceClient(boundary_service_url)
            existing_codes = set()
            chunk_size = 50
            for i in range(0, len(resolved_codes), chunk_size):
                chunk = resolved_codes[i:i + chunk_size]
                try:
                    response_data = boundary_client.search_boundaries(
                        request_info=request_info_obj,
                        tenant_id="in",
                        codes=chunk,
                    )
                    if response_data and "Boundary" in response_data:
                        for b in response_data["Boundary"]:
                            existing_codes.add(b["code"])
                except Exception as e:
                    logger.error(f"Error verifying boundary codes: {e}", exc_info=True)

            missing_codes = set(resolved_codes) - existing_codes
            if missing_codes:
                logger.warning(f"Boundary codes not found in boundary service: {missing_codes}")
                for index, row in df.iterrows():
                    code = str(row.get('Boundary Code (Mandatory)', '') or '').strip()
                    if code in missing_codes:
                        state_val = str(row.get('State (Mandatory)', '') or '').strip()
                        district_val = str(row.get('District (Mandatory)', '') or '').strip()
                        block_val = str(row.get('Block (Mandatory)', '') or '').strip()
                        df.at[index, 'Boundary Code (Mandatory)'] = ''
                        df.at[index, 'status'] = 'FAILED'
                        df.at[index, 'error'] = f"Boundary code for State '{state_val}' District '{district_val}' Block '{block_val}' not found"

        # ----------------- Run Validation ----------------- #
        validation_errors = facility_validation(
            df,
            mdms_client,
            request_info_obj,
            facility_client,
            boundary_data_df,
            'data-ingestion.FacilityIngestionSchema'
        )

        org_client = OrganizationServiceClient(org_service_url)
        registered_vendor_codes = org_client.fetch_registered_vendor_codes(request_info_obj)
        for i in range(len(df)):
            row = df.iloc[i]
            fid = row.get("Facility Id")
            is_new = pd.isna(fid) or str(fid).strip() == ""
            if not is_new:
                continue
            vendor_code = org_client.normalize_facility_vendor_code(row.get(FACILITY_VENDOR_CODE_COLUMN))
            if not vendor_code:
                validation_errors[i].append(
                    f"{FACILITY_VENDOR_CODE_COLUMN} is required; facilities cannot be created without a vendor mapping."
                )
            elif registered_vendor_codes is not None and vendor_code not in registered_vendor_codes:
                validation_errors[i].append(
                    f"Vendor code '{vendor_code}' is not registered in the vendor service; "
                    "register the vendor before facility ingestion."
                )

        # Mark rows based on validation results, preserving earlier boundary errors
        error_count = 0
        for i, errs in enumerate(validation_errors):
            existing_status = str(df.at[i, 'status']).strip().upper() if pd.notna(df.at[i, 'status']) else ''
            existing_error = str(df.at[i, 'error']).strip() if pd.notna(df.at[i, 'error']) else ''

            if existing_status == 'FAILED':
                if errs:
                    df.at[i, 'error'] = existing_error + "; " + "; ".join(dict.fromkeys(errs))
                error_count += 1
            elif errs:
                df.at[i, 'status'] = 'FAILED'
                df.at[i, 'error'] = "; ".join(dict.fromkeys(errs))
                error_count += 1
            else:
                df.at[i, 'status'] = 'PASSED'
                df.at[i, 'error'] = ''

        # ----------------- Update Facility Sheet In-Place ----------------- #
        ws = wb[facility_sheet_name]
        header_values = [cell.value for cell in ws[1]]

        # Add columns in same order as DataFrame: status, error, Boundary Code
        for col_name in ["status", "error", "Boundary Code (Mandatory)"]:
            if col_name not in header_values:
                new_col_idx = len(header_values) + 1
                cell = ws.cell(row=1, column=new_col_idx, value=col_name)
                cell.font = Font(bold=True)
                header_values.append(col_name)

        grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        # Write data rows back (without header row)
        export_df = prepare_dataframe_for_excel_export(df)
        for r_idx, row in enumerate(dataframe_to_rows(export_df, index=False, header=False), start=2):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)

                # force lock for status/error columns
                # if ws.cell(1, c_idx).value in ["status", "error"]:
                #     cell.protection = Protection(locked=True)
                #     cell.fill = grey_fill

        # Ensure sheet protection is ON
        # ws.protection.sheet = True
        # ws.protection.enable()

        # ----------------- Save to new temp file ----------------- #
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_temp_file_path = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx").name
        wb.save(output_temp_file_path)

        autofit_columns(output_temp_file_path, facility_sheet_name, auto_fit=True)

        background_tasks.add_task(cleanup_temp_file, output_temp_file_path)

        response = FileResponse(
            path=output_temp_file_path,
            filename=f"facility_validation_results_{timestamp}.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response.headers["X-Error-Count"] = str(error_count)

        return response

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Validation failed: {str(e)}")
    finally:
        if temp_input_file and os.path.exists(temp_input_file.name):
            os.unlink(temp_input_file.name)

@router.post('/facilities',
             summary='Upload and process facility Excel file',
             response_description='Returns processed Excel file with validations results')
async def upload_facilities_excel_sheet(
        facility_file: UploadFile = File(description="Excel file containing facility data"),
        facility_sheet_name: str = Form(default="FacilityIngestionTemplate",
                                        description="Name of the sheet containing facility data"),
        request_info: str = Form(default=""),
        are_facilities_onm_ready: bool = Form(description="FieldPlan ID")
):
    input_temp_file = None
    output_temp_file = None
    request_info = request_info_from_json(request_info)
    mdms_client = MDMSClient(mdms_url)

    try:
        input_temp_file, _ = await _save_upload_to_temp_file(facility_file, suffix=".xlsx")
        facility_file_path = input_temp_file.name

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"facility_ingestion_results_{timestamp}.xlsx"
        output_temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        output_temp_file.close()
        output_file_path = output_temp_file.name

        with open(facility_file_path, 'rb') as src, open(output_file_path, 'wb') as dst:
            dst.write(src.read())

        df = pd.read_excel(facility_file_path, sheet_name=facility_sheet_name)
        df.columns = [str(c).strip() for c in df.columns]
        df = df.loc[:, ~df.columns.str.startswith('Unnamed')]
        df = normalize_excel_integer_columns(df, force_columns=FACILITY_IDENTIFIER_COLUMNS)

        if 'status' not in df.columns:
            df['status'] = ''
        if 'error' not in df.columns:
            df['error'] = ''
        df['status'] = df['status'].fillna('').astype(str)
        df['error'] = df['error'].fillna('').astype(str)

        # Resolve boundary codes from State/District/Block via localization
        df = resolve_boundary_codes_for_dataframe(
            df, localization_service_url,
            boundary_code_column='Boundary Code (Mandatory)',
            logger=logger,
        )

        if facility_service_url and not df.empty:
            facility_client = FacilityServiceClient(facility_service_url)
            facility_schema = mdms_client.get_column_definitions_with_metadata(request_info,'data-ingestion.FacilityIngestionSchema')
            org_client = OrganizationServiceClient(org_service_url) if org_service_url else None
            vendor_mapping_cache: Dict[str, Dict[str, Optional[str]]] = {}
            hfr_nin_db_cache: Dict[str, bool] = {}

            _bulk_create_facilities_for_ingestion(
                df=df,
                facility_client=facility_client,
                org_client=org_client,
                request_info=request_info,
                facility_schema=facility_schema,
                are_facilities_onm_ready=are_facilities_onm_ready,
                vendor_mapping_cache=vendor_mapping_cache,
                hfr_nin_db_cache=hfr_nin_db_cache,
            )

        writer = ExcelDataWriter(output_file_path, output_sheet=facility_sheet_name)
        writer.write_data(df)

        return FileResponse(
            path=output_file_path,
            filename=output_filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        logger.error(f"Error processing facility data: {e}")
        raise HTTPException(
            status_code=500,
            detail=f"Failed to process facility data: {str(e)}"
        )
    finally:
        if input_temp_file and os.path.exists(input_temp_file.name):
            os.unlink(input_temp_file.name)

@router.post('/workStreamWithFacilities',
             summary='Upload and process workstream with facilities excel file.',
             response_description='Returns processed Excel file with validation results')
async def upload_facilities_with_workstream(
        project_id_with_type_field_plan: str = Form(default="Project id of the project with type field plan"),
        request_info: str = Form(default=""),
        installation_spoc_user_name:str = Form(default=""),
        installation_spoc_user_mobile_number:str = Form(default=""),
        installation_spoc_user_email:str = Form(default="")
)->JSONResponse:
    request_info = request_info_from_json(request_info)
    #get_authorized_request_info(request_info)

    try:
        # Fetch project of type Field Plan using project_id
        if project_service_url and hrms_service_url:
            project_client = ProjectServiceClient(project_service_url)
            hrms_client = HRMSServiceClient(hrms_service_url)
            field_plan_project = project_client.search_project(request_info, project_id_with_type_field_plan)
            project = field_plan_project["Project"][0]
            if not project:
                raise Exception("Field plan id is not correct.")
            field_plan_project_facilities = project_client.search_project_facility(request_info,
                                                                                   project_id_with_type_field_plan)
            work_stream_creation_payload = get_project_creation_payload(
                request_info,
                project["project"]['name'] + "_work_stream",
                "Work Stream",
                project_id_with_type_field_plan,
                project["project"]["startDate"],
                project["project"]["endDate"],
                "Installation"
            )
            work_stream_creation_response = json.loads(project_client.create_project(work_stream_creation_payload).text)
            work_stream = work_stream_creation_response['Project'][0]

            # Link work stream project to facilities
            facilities = field_plan_project_facilities["ProjectFacilities"]
            for facility in facilities:
                project_client.create_project_facility(request_info,
                    work_stream["id"],
                    facility["facilityId"]
                )
            # Create a installation spoc user
            installation_spoc_creation_payload = get_installation_spoc_creation_payload(request_info, installation_spoc_user_name, installation_spoc_user_mobile_number,
                                                   installation_spoc_user_email)
            user_creation_response = json.loads(hrms_client.create_user(installation_spoc_creation_payload).text)
            user = user_creation_response['Employees'][0]
            staff_creation_payload = get_staff_creation_payload(request_info, user["uuid"], work_stream["id"])
            staff_creation_response = json.loads(project_client.create_project_staff(staff_creation_payload).text)
            staff = staff_creation_response['ProjectStaff']
            return JSONResponse(
                status_code=200,
                content={"staff": staff, "user": user, "work_stream": work_stream}
            )
        return JSONResponse(
            status_code=500,
            content="Connection failed with project service and hrms service."
        )

    except Exception as e:
        logger.error(f"Error processing facility data: {e}")
        raise HTTPException(
            status_code=500,
            detail=f"Failed to process facility data: {str(e)}"
        )


@router.post('/facilityWithStaff',
             summary='Upload and process facility with staff Excel file',
             response_description="Returns processed Excel file with validation results")
async def upload_facility_with_staff_excel_sheet(
        facility_with_staff: UploadFile = File(
            description="Excel file containing facility with staff data"),
        facility_sheet: str = Form(default="Facilities_Staff",
                                    description="Name of the sheet containing facility data"),
        request_info: str = Form(default=""),
        work_stream_project_id:str = Form(default="")
):
    input_temp_file = None
    output_temp_file = None
    request_info = request_info_from_json(request_info)
    #get_authorized_request_info(request_info)

    try:
        # Create input temporary file
        input_temp_file, _ = await _save_upload_to_temp_file(facility_with_staff, suffix=".xlsx")
        facility_with_staff_file_path = input_temp_file.name

        # Create output file with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"facility_with_staff_ingestion_results_{timestamp}.xlsx"
        output_temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        output_temp_file.close()
        output_file_path = output_temp_file.name

        # Copy input to output file first
        with open(facility_with_staff_file_path, 'rb') as src, open(output_file_path, 'wb') as dst:
            dst.write(src.read())

        # Read the Excel file
        df = pd.read_excel(facility_with_staff_file_path, sheet_name=facility_sheet)

        # Add status and error columns if they don't exist
        if 'status' not in df.columns:
            df['status'] = ''
        if 'error' not in df.columns:
            df['error'] = ''

        # Process each row if services are available
        if project_service_url and not df.empty:
            project_client = ProjectServiceClient(project_service_url)
            hrms_client = HRMSServiceClient(hrms_service_url)
            work_stream_project = project_client.search_project(request_info, work_stream_project_id)
            work_stream = work_stream_project["Project"][0]
            for index, row in df.iterrows():
                if row.get("status", "") != "success":
                    try:
                        # Create project of type facility
                        facility_creation_payload = get_project_creation_payload(request_info, row.get('Health Centre Name (Mandatory)', ''), "Facility",
                                                                                 work_stream_project_id, work_stream["startDate"],work_stream["endDate"],"")
                        facility_creation_response = project_client.create_project(facility_creation_payload)
                        facility = json.loads(facility_creation_response.text)
                        if facility_creation_response.status_code in [200, 201, 202]:
                            df.at[index, 'status'] = 'success'
                            # Check if user already exists and validate roles
                            user_search_payload = get_user_creation_payload_staff(request_info, row)
                            existing_user_response = hrms_client.search_user(user_search_payload)
                            existing_user = None
                            if existing_user_response.status_code == 200:
                                response_data = existing_user_response.json()
                                employees = response_data.get("Employees", [])
                                if employees:
                                    existing_user = employees[0]

                            if existing_user:
                                # Check for role mismatch
                                role_check = check_role_mismatch_for_user_type(existing_user, "staff")
                                if role_check["has_mismatch"]:
                                    df.at[index, 'status'] = 'error'
                                    df.at[index, 'error'] = f"Role mismatch detected: {role_check['mismatch_details']}. Current roles: {', '.join(role_check['current_roles'])}. Expected roles: {', '.join(role_check['expected_roles'])}"
                                    continue
                                else:
                                    # Use existing user
                                    user_uuid = existing_user.get("uuid")
                                    df.at[index, 'status'] = 'success'
                            else:
                                # Create new user
                                user_creation_payload = get_user_creation_payload_staff(request_info, row)
                                user_creation_response = hrms_client.create_user(user_creation_payload)
                                user = json.loads(user_creation_response.text)
                                if user_creation_response.status_code in [200, 201, 202]:
                                    user_uuid = user["Employees"][0]["uuid"]
                                    df.at[index, 'status'] = 'success'
                                else:
                                    df.at[index, 'status'] = 'failed'
                                    df.at[index, 'error'] = f"User Creation Error: {user_creation_response.status_code} - {user.get('Errors', [{}])[0].get('message', 'Unknown error')}"
                                    continue

                            # Validate user_uuid before staff creation
                            if not user_uuid:
                                df.at[index, 'status'] = 'failed'
                                df.at[index, 'error'] = "User UUID is required for staff creation but was not obtained"
                                continue

                            # Create staff
                            staff_creation_payload = get_staff_creation_payload(request_info, user_uuid, facility["Project"][0]["id"])
                            staff_creation_response = project_client.create_project_staff(staff_creation_payload)
                            if staff_creation_response.status_code in [200, 201, 202]:
                                df.at[index,'status'] = 'success'
                                df.at[index, 'error'] = ''
                            else:
                                df.at[index, 'status'] = 'failed'
                                df.at[index, 'error'] = f"Staff Creation Error: {staff_creation_response.status_code} - {staff_creation_response.text}"
                        else:
                            df.at[index, 'status'] = 'failed'
                            df.at[index, 'error'] = f"Facility Creation Error: {facility_creation_response.status_code} - {facility_creation_response.text}"
                    except Exception as e:
                        df.at[index, 'status'] = 'failed'
                        df.at[index, 'error'] = f"Processing Error: {str(e)}"

        # Write data to the same sheet name that was read
        with pd.ExcelWriter(output_file_path, engine='openpyxl', mode='a') as writer:
            # Remove the existing sheet if it exists
            if facility_sheet in writer.book.sheetnames:
                idx = writer.book.sheetnames.index(facility_sheet)
                writer.book.remove(writer.book.worksheets[idx])
            # Write data to the sheet
            df.to_excel(writer, sheet_name=facility_sheet, index=False)

        return FileResponse(
            path=output_file_path,
            filename=output_filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        logger.error(f"Error processing facility data: {e}")
        raise HTTPException(
            status_code=500,
            detail=f"Failed to process facility data: {str(e)}"
        )
    finally:
        if input_temp_file and os.path.exists(input_temp_file.name):
            os.unlink(input_temp_file.name)



@router.post('/facilityWithSupervisors',
             summary='Upload and process facility with supervisors Excel file',
             response_description="Returns processed Excel file with validation results")
async def upload_facility_with_supervisors_excel_sheet(
        facility_with_supervisors: UploadFile = File(
            description="Excel file containing facility with supervisors data"),
        facility_sheet: str = Form(default="Facilities_Supervisors",
                                    description="Name of the sheet containing facility data"),
        request_info: str = Form(default=""),
        work_stream_project_id:str = Form(default="")
):
    input_temp_file = None
    output_temp_file = None
    request_info = request_info_from_json(request_info)
    #get_authorized_request_info(request_info)

    try:
        # Create input temporary file
        input_temp_file, _ = await _save_upload_to_temp_file(facility_with_supervisors, suffix=".xlsx")
        facility_with_supervisors_file_path = input_temp_file.name

        # Create output file with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"facility_with_supervisor_ingestion_results_{timestamp}.xlsx"
        output_temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        output_temp_file.close()
        output_file_path = output_temp_file.name

        # Copy input to output file first
        with open(facility_with_supervisors_file_path, 'rb') as src, open(output_file_path, 'wb') as dst:
            dst.write(src.read())

        # Read the Excel file
        df = pd.read_excel(facility_with_supervisors_file_path, sheet_name=facility_sheet)

        # Add status and error columns if they don't exist
        if 'status' not in df.columns:
            df['status'] = ''
        if 'error' not in df.columns:
            df['error'] = ''

        # Process each row if services are available
        if project_service_url and not df.empty:
            project_client = ProjectServiceClient(project_service_url)
            hrms_client = HRMSServiceClient(hrms_service_url)
            work_stream_project = project_client.search_project(request_info, work_stream_project_id)
            work_stream = work_stream_project["Project"][0]
            for index, row in df.iterrows():
                if row.get("status", "") != "success":
                    try:
                        existing_facility = project_client.search_project_facility(request_info, work_stream_project_id)
                        facility_list = existing_facility.get('ProjectFacilities', [])

                        facility_created = False
                        if facility_list:
                            facility = facility_list[0]
                            facility_created = False  # existing, not newly created
                        else:
                            # Create facility if not found
                            facility_creation_payload = get_project_creation_payload(request_info, row.get('Health Centre Name (Mandatory)', ''), "Facility",
                                                                                 work_stream_project_id, work_stream["startDate"],work_stream["endDate"],"")
                            facility_creation_response = project_client.create_project(facility_creation_payload)
                            if facility_creation_response.status_code not in [200, 201, 202]:
                                df.at[index, 'status'] = 'failed'
                                df.at[index, 'error'] = (
                                    f"Facility Creation Error: {facility_creation_response.status_code} - {facility_creation_response.text}"
                                )
                                continue

                            facility = json.loads(facility_creation_response.text)
                            facility_created = True

                        # 🧠 Correctly extract project ID based on the structure
                        if facility_created:
                            project_id = facility["Project"][0]["id"]
                        else:
                            project_id = facility["projectId"]
                        # Check if user already exists and validate roles
                        user_search_payload = get_user_creation_payload_supervisors(request_info, row)
                        existing_user_response = hrms_client.search_user(user_search_payload)
                        existing_user = None
                        if existing_user_response.status_code == 200:
                            response_data = existing_user_response.json()
                            employees = response_data.get("Employees", [])
                            if employees:
                                existing_user = employees[0]

                        if existing_user:
                            # Check for role mismatch
                            role_check = check_role_mismatch_for_user_type(existing_user, "supervisor")
                            if role_check["has_mismatch"]:
                                df.at[index, 'status'] = 'error'
                                df.at[index, 'error'] = f"Role mismatch detected: {role_check['mismatch_details']}. Current roles: {', '.join(role_check['current_roles'])}. Expected roles: {', '.join(role_check['expected_roles'])}"
                                continue
                            else:
                                # Use existing user
                                user_uuid = existing_user.get("uuid")
                                df.at[index, 'status'] = 'success'
                        else:
                            # Create new user
                            user_creation_payload = get_user_creation_payload_supervisors(request_info, row)
                            user_creation_response = hrms_client.create_user(user_creation_payload)
                            user = json.loads(user_creation_response.text)
                            if user_creation_response.status_code in [200, 201, 202]:
                                user_uuid = user["Employees"][0]["uuid"]
                                df.at[index, 'status'] = 'success'
                            else:
                                df.at[index, 'status'] = 'failed'
                                df.at[index, 'error'] = f"User Creation Error: {user_creation_response.status_code} - {user.get('Errors', [{}])[0].get('message', 'Unknown error')}"
                                continue

                        # Create staff
                        staff_creation_payload = get_staff_creation_payload(request_info, user_uuid, project_id)
                        staff_creation_response = project_client.create_project_staff(staff_creation_payload)
                        if staff_creation_response.status_code in [200, 201, 202]:
                            df.at[index,'status'] = 'success'
                            df.at[index, 'error'] = ''
                        else:
                            df.at[index, 'status'] = 'failed'
                            df.at[index, 'error'] = f"Staff Creation Error: {staff_creation_response.status_code} - {staff_creation_response.text}"
                    except Exception as e:
                        df.at[index, 'status'] = 'failed'
                        df.at[index, 'error'] = f"Processing Error: {str(e)}"

        # Write data to the same sheet name that was read
        with pd.ExcelWriter(output_file_path, engine='openpyxl', mode='a') as writer:
            # Remove the existing sheet if it exists
            if facility_sheet in writer.book.sheetnames:
                idx = writer.book.sheetnames.index(facility_sheet)
                writer.book.remove(writer.book.worksheets[idx])
            # Write data to the sheet
            df.to_excel(writer, sheet_name=facility_sheet, index=False)

        return FileResponse(
            path=output_file_path,
            filename=output_filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        logger.error(f"Error processing facility data: {e}")
        raise HTTPException(
            status_code=500,
            detail=f"Failed to process facility data: {str(e)}"
        )
    finally:
        if input_temp_file and os.path.exists(input_temp_file.name):
            os.unlink(input_temp_file.name)

@router.post('/facilityWithSupervisorUpdateWorkflowState',
             summary='Upload and process facility with supervisors Excel file',
             response_description="Returns processed Excel file with validation results")
async def upload_facility_with_supervisors_workflow_state_excel_sheet(
        facility_with_supervisors: UploadFile = File(
            description="Excel file containing facility with supervisors data"),
        facility_sheet: str = Form(default="Facilities_Supervisors",
                                    description="Name of the sheet containing facility data"),
        request_info: str = Form(default=""),
        work_stream_project_id:str = Form(default="")
):
    input_temp_file = None
    output_temp_file = None
    request_info = request_info_from_json(request_info)
    #get_authorized_request_info(request_info)

    try:
        # Create input temporary file
        input_temp_file, _ = await _save_upload_to_temp_file(facility_with_supervisors, suffix=".xlsx")
        facility_with_supervisors_file_path = input_temp_file.name

        # Create output file with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"facility_with_supervisor_ingestion_results_{timestamp}.xlsx"
        output_temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        output_temp_file.close()
        output_file_path = output_temp_file.name

        # Copy input to output file first
        with open(facility_with_supervisors_file_path, 'rb') as src, open(output_file_path, 'wb') as dst:
            dst.write(src.read())

        # Read the Excel file
        df = pd.read_excel(facility_with_supervisors_file_path, sheet_name=facility_sheet)

        # Add status and error columns if they don't exist
        if 'status' not in df.columns:
            df['status'] = ''
        if 'error' not in df.columns:
            df['error'] = ''

        # Process each row if services are available
        if project_service_url and not df.empty:
            project_client = ProjectServiceClient(project_service_url)
            hrms_client = HRMSServiceClient(hrms_service_url)
            work_stream_project = project_client.search_project(request_info, work_stream_project_id)
            work_stream = work_stream_project["Project"][0]
            for index, row in df.iterrows():
                if row.get("status", "") != "success":
                    try:
                        existing_facility = project_client.search_project_facility(request_info, work_stream_project_id)
                        facility_list = existing_facility.get('ProjectFacilities', [])

                        facility_created = False
                        if facility_list:
                            facility = facility_list[0]
                            facility_created = False  # existing, not newly created
                        else:
                            # Create facility if not found
                            facility_creation_payload = get_project_creation_payload(request_info, row.get('Health Centre Name (Mandatory)', ''), "Facility",
                                                                                 work_stream_project_id, work_stream["startDate"],work_stream["endDate"],"")
                            facility_creation_response = project_client.create_project(facility_creation_payload)
                            if facility_creation_response.status_code not in [200, 201, 202]:
                                df.at[index, 'status'] = 'failed'
                                df.at[index, 'error'] = (
                                    f"Facility Creation Error: {facility_creation_response.status_code} - {facility_creation_response.text}"
                                )
                                continue

                            facility = json.loads(facility_creation_response.text)
                            facility_created = True

                        # 🧠 Correctly extract project ID based on the structure
                        if facility_created:
                            project_id = facility["Project"][0]["id"]
                        else:
                            project_id = facility["projectId"]
                        # Check if user already exists and validate roles
                        # Determine user type based on Role column
                        user_type = "supervisor"  # default
                        if 'Role' in df.columns:
                            role_value = df.at[index, 'Role']
                            if role_value and str(role_value).strip().lower() == 'supervisor':
                                user_type = "supervisor"
                            else:
                                user_type = "staff"

                        # Create search payload based on user type
                        if user_type == "supervisor":
                            user_search_payload = get_user_creation_payload_supervisors(request_info, row)
                        else:
                            user_search_payload = get_user_creation_payload_staff(request_info, row)

                        existing_user_response = hrms_client.search_user(user_search_payload)
                        existing_user = None
                        if existing_user_response.status_code == 200:
                            response_data = existing_user_response.json()
                            employees = response_data.get("Employees", [])
                            if employees:
                                existing_user = employees[0]

                        if existing_user:
                            # Check for role mismatch
                            role_check = check_role_mismatch_for_user_type(existing_user, user_type)
                            if role_check["has_mismatch"]:
                                df.at[index, 'status'] = 'error'
                                df.at[index, 'error'] = f"Role mismatch detected: {role_check['mismatch_details']}. Current roles: {', '.join(role_check['current_roles'])}. Expected roles: {', '.join(role_check['expected_roles'])}"
                                continue
                            else:
                                # Use existing user
                                user_uuid = existing_user.get("uuid")
                        else:
                            # Create new user based on role type
                            if user_type == "supervisor":
                                user_creation_payload = get_user_creation_payload_supervisors(request_info, row)
                            else:
                                user_creation_payload = get_user_creation_payload_staff(request_info, row)

                            user_creation_response = hrms_client.create_user(user_creation_payload)
                            user = json.loads(user_creation_response.text)
                            if user_creation_response.status_code in [200, 201, 202]:
                                user_uuid = user["Employees"][0]["uuid"]
                            else:
                                df.at[index, 'status'] = 'failed'
                                df.at[index, 'error'] = f"User Creation Error: {user_creation_response.status_code} - {user.get('Errors', [{}])[0].get('message', 'Unknown error')}"
                                continue

                        # Validate user_uuid before staff creation
                        if not user_uuid:
                            df.at[index, 'status'] = 'failed'
                            df.at[index, 'error'] = "User UUID is required for staff creation but was not obtained"
                            continue

                        # Create staff
                        staff_creation_payload = get_staff_creation_payload(request_info, user_uuid, project_id)
                        staff_creation_response = project_client.create_project_staff(staff_creation_payload)
                        if staff_creation_response.status_code in [200, 201, 202]:
                            # Validate Role column exists
                            if 'Role' not in df.columns:
                                df.at[index, 'status'] = 'failed'
                                df.at[index, 'error'] = "Role column is required for workflow state updates"
                                continue
                            # update workflow state
                            role_value = df.at[index,'Role']
                            if role_value and str(role_value).strip().lower() == 'supervisor':
                                update_workflow_state_response = project_client.update_workflow(request_info, work_stream_project_id, 'ASSIGN_FIELD_SUPERVISOR')
                            else:
                                update_workflow_state_response = project_client.update_workflow(request_info, work_stream_project_id,
                                                                                                'ASSIGN_FIELD_STAFF')
                            if update_workflow_state_response.status_code in [200, 201, 202]:
                                df.at[index,'status'] = 'success'
                                df.at[index, 'error'] = ''
                            else:
                                df.at[index, 'status'] = 'failed'
                                df.at[
                                    index, 'error'] = f"Update Workflow state Error: {update_workflow_state_response.status_code} - {update_workflow_state_response.text}"
                        else:
                            df.at[index, 'status'] = 'failed'
                            df.at[index, 'error'] = f"Staff Creation Error: {staff_creation_response.status_code} - {staff_creation_response.text}"
                    except Exception as e:
                        df.at[index, 'status'] = 'failed'
                        df.at[index, 'error'] = f"Processing Error: {str(e)}"

        # Write data to the same sheet name that was read
        with pd.ExcelWriter(output_file_path, engine='openpyxl', mode='a') as writer:
            # Remove the existing sheet if it exists
            if facility_sheet in writer.book.sheetnames:
                idx = writer.book.sheetnames.index(facility_sheet)
                writer.book.remove(writer.book.worksheets[idx])
            # Write data to the sheet
            df.to_excel(writer, sheet_name=facility_sheet, index=False)

        return FileResponse(
            path=output_file_path,
            filename=output_filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        logger.error(f"Error processing facility data: {e}")
        raise HTTPException(
            status_code=500,
            detail=f"Failed to process facility data: {str(e)}"
        )
    finally:
        if input_temp_file and os.path.exists(input_temp_file.name):
            os.unlink(input_temp_file.name)

@router.post('/projects',
             summary='Upload and process project Excel file',
             response_description='Returns processed Excel file with validations results')
async def upload_projects_excel_sheet(
        project_file: UploadFile = File(description="Excel file containing project data"),
        project_sheet_name: str = Form(default="Project Data",
                                        description="Name of the sheet containing project data"),
        request_info: str = Form(default="")
):
    input_temp_file = None
    output_temp_file = None
    request_info = request_info_from_json(request_info)
    #get_authorized_request_info(request_info)

    try:
        input_temp_file, _ = await _save_upload_to_temp_file(project_file, suffix=".xlsx")
        project_file_path = input_temp_file.name

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"project_ingestion_results_{timestamp}.xlsx"
        output_temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        output_temp_file.close()
        output_file_path = output_temp_file.name

        with open(project_file_path, 'rb') as src, open(output_file_path, 'wb') as dst:
            dst.write(src.read())

        df = pd.read_excel(project_file_path, sheet_name=project_sheet_name)

        if 'status' not in df.columns:
            df['status'] = ''
        if 'error' not in df.columns:
            df['error'] = ''
        if 'Project ID' not in df.columns:
            df['Project ID'] = ''

        if project_service_url and hrms_service_url and not df.empty:
            hrms_client = HRMSServiceClient(hrms_service_url)
            project_client = ProjectServiceClient(project_service_url)
            for index, row in df[df['status'] != 'success'].iterrows():
                try:

                    project_data_payload = create_project_payload(request_info, row)
                    response = project_client.create_project(project_data_payload)
                    response_data = response.json()

                    if df.at[index, 'Project Type'] == 'Field Plan':
                        name = df.at[index, 'Name']
                        mobile_number_raw = df.at[index, 'Mobile Number']
                        email_value = df.at[index, 'Email']
                        if pd.isna(email_value) or not email_value:
                            df.at[index, 'status'] = 'failed'
                            df.at[index, 'error'] = 'Email is required for Field Plan projects'
                            continue
                        if pd.isna(mobile_number_raw) or not mobile_number_raw:
                            df.at[index, 'status'] = 'failed'
                            df.at[index, 'error'] = 'Mobile Number is required for Field Plan projects'
                            continue


                        mobile_number = str(int(mobile_number_raw))
                        email = str(email_value).strip()

                        spoc_payload = get_installation_spoc_creation_payload(request_info, name, mobile_number, email)

                        user_response = hrms_client.search_user(spoc_payload)

                        if user_response.status_code not in [200, 201, 202]:
                            df.at[index, 'status'] = 'failed'
                            df.at[
                                index, 'error'] = f"User search failed with status: {user_response.status_code} - {user_response.text}"
                            continue

                        response_body = json.loads(user_response.text)
                        employee_list = response_body.get("Employees", [])

                        # Filter for matching email
                        matched_user = None
                        for emp in employee_list:
                            if emp["user"]["emailId"].strip() == email:
                                matched_user = emp["user"]
                                break

                        if not matched_user:
                            df.at[index, 'status'] = 'failed'
                            df.at[index, 'error'] = f"No matching user found for email: {email}"
                            continue

                    if response.status_code in [200, 201, 202] and isinstance(response_data.get('Project'), list) and response_data[
                        'Project']:
                        if df.at[index, 'Project Type'] == 'Field Plan':

                            user_uuid = matched_user.get("uuid")
                            project_id = response_data['Project'][0].get('id')

                            staff_payload = get_staff_creation_payload(request_info, user_uuid, project_id)
                            staff_response = project_client.create_project_staff(staff_payload)

                            if staff_response.status_code in [200, 201, 202]:

                                staff_search_payload = get_staff_search_payload(request_info, user_uuid)
                                staff_search_response = project_client.search_project_staff_by_id(staff_search_payload)
                                if staff_search_response.status_code in [200, 201]:
                                    logger.debug(f"Staff search response for user {user_uuid}: {staff_search_response.text}")

                                    staff_list = staff_search_response.json().get("ProjectStaff", [])
                                    logger.debug(f"Found {len(staff_list)} staff members for user {user_uuid}")

                                    if len(staff_list) == 1:
                                        sms_request = {
                                            "mobileNumber": mobile_number,
                                            "message": "Yor are assigned to the field plan",
                                            "expiryTime": None
                                        }
                                        producer = Producer()
                                        producer.send("egov.core.notification.sms", sms_request)
                                        producer.close()


                                df.at[index, 'status'] = 'success'
                                df.at[index, 'error'] = ''
                                df.at[index, 'Project ID'] = project_id
                            else:
                                df.at[index, 'status'] = 'failed'
                                df.at[index, 'error'] = (
                                    f"Staff Creation Error: {staff_response.status_code} - {staff_response.text}")
                        else:
                            project_id = response_data['Project'][0].get('id')
                            df.at[index, 'status'] = 'success'
                            df.at[index, 'error'] = ''
                            df.at[index, 'Project ID'] = project_id
                    elif response.status_code == 400:
                        error_data = response.json()
                        error_message = error_data.get('Errors', [{}])[0].get('message', 'Unknown error')
                        df.at[index, 'status'] = 'failed'
                        df.at[index, 'error'] = error_message
                    else:
                        df.at[index, 'status'] = 'failed'
                        df.at[index, 'error'] = f'{response.status_code}: {response.text}'
                except Exception as e:
                    df.at[index, 'status'] = 'failed'
                    df.at[index, 'error'] = f'Exception: {str(e)}'

        writer = ExcelDataWriter(output_file_path, output_sheet=project_sheet_name)
        writer.write_data(df)

        return FileResponse(
            path=output_file_path,
            filename=output_filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        logger.error(f"Error processing project data: {e}")
        raise HTTPException(
            status_code=500,
            detail=f"Failed to process project data: {str(e)}"
        )
    finally:
        if input_temp_file and os.path.exists(input_temp_file.name):
            os.unlink(input_temp_file.name)

@router.post('/facilitySelection',
             summary='Upload and process facility selection Excel file',
             response_description='Returns processed Excel file with validations results')
async def upload_facility_selection_excel_sheet(
        facility_selection_file: UploadFile = File(description="Excel file containing facility selection data"),
        project_id: str = Form(...),
        facility_selection_sheet_name: str = Form(default="Facility Selection Template",
                                        description="Name of the sheet containing facility selection data"),
        request_info: str = Form(default="")
):
    input_temp_file = None
    output_temp_file = None
    request_info = request_info_from_json(request_info)
    #get_authorized_request_info(request_info)

    try:
        input_temp_file, _ = await _save_upload_to_temp_file(facility_selection_file, suffix=".xlsx")
        project_file_path = input_temp_file.name

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"project_facility_ingestion_results_{timestamp}.xlsx"
        output_temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        output_temp_file.close()
        output_file_path = output_temp_file.name

        with open(project_file_path, 'rb') as src, open(output_file_path, 'wb') as dst:
            dst.write(src.read())

        df = pd.read_excel(project_file_path, sheet_name=facility_selection_sheet_name)

        if 'status' not in df.columns:
            df['status'] = ''
        if 'error' not in df.columns:
            df['error'] = ''

        if project_service_url and not df.empty:
            project_client = ProjectServiceClient(project_service_url)
            for index, row in df[(df['status'] != 'success') & (df['Selection?'] == 'Yes')].iterrows():
                try:
                    facility_id = row.get("HC ID")
                    if pd.isna(facility_id):
                        df.at[index, 'status'] = 'failed'
                        df.at[index, 'error'] = 'HC ID must not be null.'
                        continue
                    response = project_client.create_project_facility(request_info, project_id, facility_id)
                    if response.status_code in (200, 201, 202):
                        df.at[index, 'status'] = 'success'
                        df.at[index, 'error'] = ''
                    elif response.status_code == 400:
                        error_data = response.json()
                        error_message = error_data.get('Errors', [{}])[0].get('message', 'Unknown error')
                        df.at[index, 'status'] = 'failed'
                        df.at[index, 'error'] = error_message
                    else:
                        df.at[index, 'status'] = 'failed'
                        df.at[index, 'error'] = f'{response.status_code}: {response.text}'
                except Exception as e:
                    df.at[index, 'status'] = 'failed'
                    df.at[index, 'error'] = f'Exception: {str(e)}'

        writer = ExcelDataWriter(output_file_path, output_sheet=facility_selection_sheet_name)
        writer.write_data(df)

        return FileResponse(
            path=output_file_path,
            filename=output_filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        logger.error(f"Error processing facility selection data: {e}")
        raise HTTPException(
            status_code=500,
            detail=f"Failed to process facility selection data: {str(e)}"
        )
    finally:
        if input_temp_file and os.path.exists(input_temp_file.name):
            os.unlink(input_temp_file.name)


@router.post('/icc-reports',
             summary='Bulk-upload ICC report Excel files, validate each against its System Type, '
                     'convert to JSON, and store them via field-planner in one call',
             response_description='Returns the field-planner bulk template creation response')
async def upload_icc_reports(
        background_tasks: BackgroundTasks,
        items: str = Form(
            ...,
            description='JSON array of metadata objects, one per file, paired positionally with '
                        'icc_files: [{"systemType": "...", "totalSystemCapacity": "...", '
                        '"fieldPlanId": "...", "tenantId": "in"}]'
        ),
        icc_files: List[UploadFile] = File(
            ..., description="ICC Report Excel files (.xlsx), positionally paired with items"
        ),
        request_info: str = Form(default="")
):
    """
    Accepts N metadata items and N Excel files, paired positionally (items[i] <-> icc_files[i]).
    Validation is all-or-nothing: every item is validated first, and if ANY item fails, nothing is
    forwarded to field-planner. Request-level problems (malformed `items` JSON, missing required
    field(s), wrong file extension, item/file count mismatch) are still reported as a plain JSON
    400 with `{"message", "errors": [{"index", "error"}, ...]}`. But a per-file ICC validation
    failure (structural mismatch, wrong System Type, BOM/brand problems) is instead returned the
    same way `/fieldPlanfacilitiesValidateData` reports its row failures: a 200 response whose body
    is a downloadable Excel attachment (the offending file, with a new "Validation Errors" sheet
    appended listing every problem found in it) and an `X-Error-Count` header holding the total
    number of individual errors across every failing file (not the file count), instead of an HTTP
    error status. If more than one file in the batch fails, all of them (each annotated) are bundled
    into a `.zip` attachment instead.
    """
    request_info_obj = request_info_from_json(request_info)
    mdms_client = MDMSClient(mdms_url)

    try:
        parsed_items = json.loads(items)
    except json.JSONDecodeError as e:
        raise HTTPException(status_code=400, detail=f"`items` is not valid JSON: {e}")

    if not isinstance(parsed_items, list) or len(parsed_items) == 0:
        raise HTTPException(status_code=400, detail="`items` must be a non-empty JSON array")

    if len(parsed_items) != len(icc_files):
        raise HTTPException(
            status_code=400,
            detail=(
                f"Item count ({len(parsed_items)}) does not match file count ({len(icc_files)}); "
                "items[i] and icc_files[i] must be paired positionally."
            ),
        )

    required_keys = ("systemType", "totalSystemCapacity", "fieldPlanId")
    field_errors = []
    for idx, item in enumerate(parsed_items):
        if not isinstance(item, dict):
            field_errors.append({"index": idx, "error": "item must be a JSON object"})
            continue
        missing = [k for k in required_keys if not item.get(k)]
        if missing:
            field_errors.append({"index": idx, "error": f"missing required field(s): {', '.join(missing)}"})
        if not _has_allowed_excel_extension(icc_files[idx].filename):
            field_errors.append({
                "index": idx,
                "error": f"'{icc_files[idx].filename}' is not an Excel file (.xlsx/.xls required)",
            })

    if field_errors:
        raise HTTPException(status_code=400, detail={"message": "Invalid items in batch", "errors": field_errors})

    temp_files = []
    converted = []
    failed_files = []  # [(annotated_temp_path, original_filename), ...]
    total_error_count = 0

    try:
        for idx, (item, upload) in enumerate(zip(parsed_items, icc_files)):
            temp_file, _ = await _save_upload_to_temp_file(upload, suffix=".xlsx")
            temp_files.append((temp_file, upload))
            try:
                detected_type, icc_json, fallback_fields, unmatched_fields = validate_and_convert(
                    temp_file.name, item["systemType"],
                    mdms_client=mdms_client, request_info=request_info_obj,
                )
                converted.append({
                    "tenant_id": item.get("tenantId", "in"),
                    "field_plan_id": item["fieldPlanId"],
                    "system_type": item["systemType"],
                    "total_capacity": item["totalSystemCapacity"],
                    "template_data": icc_json,
                })
                logger.info(
                    f"ICC report[{idx}] converted: systemType={item['systemType']} "
                    f"(detected={detected_type}), keys={len(icc_json)}, "
                    f"fallback_keys={len(fallback_fields)}, unmatched={len(unmatched_fields)}"
                )
            except ICCValidationError as e:
                logger.warning(f"ICC report[{idx}] upload rejected: {e}")
                total_error_count += _write_validation_errors_sheet(temp_file.name, str(e))
                failed_files.append((temp_file.name, upload.filename or f"icc_report_{idx}.xlsx"))

        if failed_files:
            logger.warning(
                f"ICC bulk upload rejected: {len(failed_files)} of {len(parsed_items)} item(s) failed validation"
            )
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            if len(failed_files) == 1:
                src_path, _ = failed_files[0]
                output_path = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx").name
                shutil.copyfile(src_path, output_path)
                response_filename = f"icc_report_validation_errors_{timestamp}.xlsx"
                media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            else:
                output_path = tempfile.NamedTemporaryFile(delete=False, suffix=".zip").name
                with zipfile.ZipFile(output_path, "w") as zf:
                    for src_path, filename in failed_files:
                        zf.write(src_path, arcname=filename)
                response_filename = f"icc_report_validation_errors_{timestamp}.zip"
                media_type = "application/zip"

            background_tasks.add_task(cleanup_temp_file, output_path)
            response = FileResponse(path=output_path, filename=response_filename, media_type=media_type)
            response.headers["X-Error-Count"] = str(total_error_count)
            return response

        if not fieldPlan_service_url:
            raise HTTPException(status_code=500, detail="FIELDPLAN_SERVICE_URL is not configured")

        files_payload = []
        for temp_file, upload in temp_files:
            with open(temp_file.name, "rb") as f:
                files_payload.append((upload.filename or "icc_report.xlsx", f.read()))

        field_plan_client = FieldPlanServiceClient(fieldPlan_service_url)
        response = field_plan_client.create_field_plan_templates(request_info_obj, converted, files_payload)

        if response.status_code in (200, 201, 202):
            return JSONResponse(status_code=response.status_code, content=response.json())

        logger.error(f"field-planner rejected the bulk template request: {response.status_code} - {response.text}")
        raise HTTPException(
            status_code=response.status_code if response.status_code >= 400 else 502,
            detail=f"field-planner rejected the template batch: {response.text}",
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error processing ICC report batch: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to process ICC report batch: {str(e)}")
    finally:
        for temp_file, _ in temp_files:
            if temp_file and os.path.exists(temp_file.name):
                os.unlink(temp_file.name)


@router.post('/icc-reports/_update',
             summary='Bulk-update existing ICC report field plan templates, optionally replacing '
                     'their Excel files, via field-planner in one call',
             response_description='Returns the field-planner bulk template update response')
async def update_icc_reports(
        background_tasks: BackgroundTasks,
        items: str = Form(
            ...,
            description='JSON array of metadata objects, one per template to update: '
                        '[{"id": "...", "systemType": "...", "totalSystemCapacity": "...", '
                        '"fieldPlanId": "...", "tenantId": "in"}]'
        ),
        icc_files: Optional[List[UploadFile]] = File(
            default=None,
            description="Optional ICC Report Excel files (.xlsx), positionally paired 1:1 with "
                        "items. Omit entirely to update metadata only and keep every template's "
                        "existing file - a partial list (fewer files than items) is rejected."
        ),
        request_info: str = Form(default="")
):
    """
    Accepts N metadata items, each identifying an existing FieldPlanTemplate by "id", and either
    zero or exactly N Excel files paired positionally (items[i] <-> icc_files[i]). Validation is
    all-or-nothing: every item (and every file that was provided) is validated first, and if ANY
    item fails, nothing is forwarded to field-planner. Request-level problems (malformed `items`
    JSON, missing required field(s), wrong file extension, item/file count mismatch) are still
    reported as a plain JSON 400 with `{"message", "errors": [{"index", "error"}, ...]}`. But a
    per-file ICC validation failure (structural mismatch, wrong System Type, BOM/brand problems) is
    instead returned the same way `/fieldPlanfacilitiesValidateData` reports its row failures: a
    200 response whose body is a downloadable Excel attachment (the offending file, with a new
    "Validation Errors" sheet appended listing every problem found in it) and an `X-Error-Count`
    header holding the total number of individual errors across every failing file (not the file
    count), instead of an HTTP error status. If more than one file in the batch fails, all of them
    (each annotated) are bundled into a `.zip` attachment
    instead.
    """
    request_info_obj = request_info_from_json(request_info)
    mdms_client = MDMSClient(mdms_url)

    try:
        parsed_items = json.loads(items)
    except json.JSONDecodeError as e:
        raise HTTPException(status_code=400, detail=f"`items` is not valid JSON: {e}")

    if not isinstance(parsed_items, list) or len(parsed_items) == 0:
        raise HTTPException(status_code=400, detail="`items` must be a non-empty JSON array")

    has_files = icc_files is not None and len(icc_files) > 0
    if has_files and len(icc_files) != len(parsed_items):
        raise HTTPException(
            status_code=400,
            detail=(
                f"Item count ({len(parsed_items)}) does not match file count ({len(icc_files)}); "
                "provide exactly one file per item, or omit icc_files entirely to update "
                "metadata only."
            ),
        )

    required_keys = ("id", "systemType", "totalSystemCapacity", "fieldPlanId")
    field_errors = []
    for idx, item in enumerate(parsed_items):
        if not isinstance(item, dict):
            field_errors.append({"index": idx, "error": "item must be a JSON object"})
            continue
        missing = [k for k in required_keys if not item.get(k)]
        if missing:
            field_errors.append({"index": idx, "error": f"missing required field(s): {', '.join(missing)}"})
        if has_files and not _has_allowed_excel_extension(icc_files[idx].filename):
            field_errors.append({
                "index": idx,
                "error": f"'{icc_files[idx].filename}' is not an Excel file (.xlsx/.xls required)",
            })

    if field_errors:
        raise HTTPException(status_code=400, detail={"message": "Invalid items in batch", "errors": field_errors})

    temp_files = []
    converted = []
    failed_files = []  # [(annotated_temp_path, original_filename), ...]
    total_error_count = 0

    try:
        for idx, item in enumerate(parsed_items):
            upload = icc_files[idx] if has_files else None
            template_data = None
            if upload is not None:
                temp_file, _ = await _save_upload_to_temp_file(upload, suffix=".xlsx")
                temp_files.append((temp_file, upload))
                try:
                    detected_type, icc_json, fallback_fields, unmatched_fields = validate_and_convert(
                        temp_file.name, item["systemType"],
                        mdms_client=mdms_client, request_info=request_info_obj,
                    )
                    template_data = icc_json
                    logger.info(
                        f"ICC report[{idx}] (update, id={item['id']}) converted: "
                        f"systemType={item['systemType']} (detected={detected_type}), "
                        f"keys={len(icc_json)}, fallback_keys={len(fallback_fields)}, "
                        f"unmatched={len(unmatched_fields)}"
                    )
                except ICCValidationError as e:
                    logger.warning(f"ICC report[{idx}] update rejected: {e}")
                    total_error_count += _write_validation_errors_sheet(temp_file.name, str(e))
                    failed_files.append((temp_file.name, upload.filename or f"icc_report_{idx}.xlsx"))
                    continue

            converted.append({
                "id": item["id"],
                "tenant_id": item.get("tenantId", "in"),
                "field_plan_id": item["fieldPlanId"],
                "system_type": item["systemType"],
                "total_capacity": item["totalSystemCapacity"],
                "template_data": template_data,
            })

        if failed_files:
            logger.warning(
                f"ICC bulk update rejected: {len(failed_files)} of {len(parsed_items)} item(s) failed validation"
            )
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            if len(failed_files) == 1:
                src_path, _ = failed_files[0]
                output_path = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx").name
                shutil.copyfile(src_path, output_path)
                response_filename = f"icc_report_validation_errors_{timestamp}.xlsx"
                media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            else:
                output_path = tempfile.NamedTemporaryFile(delete=False, suffix=".zip").name
                with zipfile.ZipFile(output_path, "w") as zf:
                    for src_path, filename in failed_files:
                        zf.write(src_path, arcname=filename)
                response_filename = f"icc_report_validation_errors_{timestamp}.zip"
                media_type = "application/zip"

            background_tasks.add_task(cleanup_temp_file, output_path)
            response = FileResponse(path=output_path, filename=response_filename, media_type=media_type)
            response.headers["X-Error-Count"] = str(total_error_count)
            return response

        if not fieldPlan_service_url:
            raise HTTPException(status_code=500, detail="FIELDPLAN_SERVICE_URL is not configured")

        files_payload = []
        if has_files:
            for temp_file, upload in temp_files:
                with open(temp_file.name, "rb") as f:
                    files_payload.append((upload.filename or "icc_report.xlsx", f.read()))

        field_plan_client = FieldPlanServiceClient(fieldPlan_service_url)
        response = field_plan_client.update_field_plan_templates(request_info_obj, converted, files_payload)

        if response.status_code in (200, 201, 202):
            return JSONResponse(status_code=response.status_code, content=response.json())

        logger.error(f"field-planner rejected the bulk template update: {response.status_code} - {response.text}")
        raise HTTPException(
            status_code=response.status_code if response.status_code >= 400 else 502,
            detail=f"field-planner rejected the template update batch: {response.text}",
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error processing ICC report update batch: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to process ICC report update batch: {str(e)}")
    finally:
        for temp_file, _ in temp_files:
            if temp_file and os.path.exists(temp_file.name):
                os.unlink(temp_file.name)


def get_hrms_employee_info(codes: List[str], db_conn) -> Dict[str, str]:
    try:
        with db_conn.cursor() as cursor:
            sql = "SELECT code, tenantid  FROM eg_hrms_employee WHERE code = ANY (%s)"
            cursor.execute(sql, (codes,))
            rows = cursor.fetchall()
            return {row[0]: row[1] for row in rows}
    except Exception as e:
        logger.error(f"Error fetching HRMS employee info: {e}")
        return {}

def get_tenant_mapping(request_info: RequestInfo, tenant_ids: List[str]) -> Dict:
    """
    Fetch tenant mapping from MDMS for PHC subtypes
    """
    all_tenant_data = {}

    for tenant_id in tenant_ids:
        try:
            search_url = f"{mdms_url}/egov-mdms-service/v1/_search"
            search_payload = {
                "RequestInfo": request_info.model_dump(by_alias=True, exclude_none=True),
                "MdmsCriteria": {
                    "tenantId": tenant_id,
                    "moduleDetails": [
                        {
                            "moduleName": "tenant",
                            "masterDetails": [
                                {
                                    "name": "tenants"
                                }
                            ]
                        }
                    ]
                }
            }
            response = requests.post(search_url, json=search_payload)
            if response.status_code == 200:
                data = response.json()
                tenants = data.get("MdmsRes", {}).get("tenant", {}).get("tenants", [])
                all_tenant_data.update({t["code"]: t for t in tenants if t.get("code") and t["code"] not in all_tenant_data})
        except Exception as e:
            logger.error(f"Error fetching tenant mapping from MDMS: {e}")

    return all_tenant_data


def get_block_mapping_from_mdms(request_info: RequestInfo, tenant_ids: List[str]) -> Dict[str, dict]:
    """
    Fetch block mapping from MDMS where moduleName is 'Incident' and masterDetails name is 'Block'.
    Returns a dictionary with 'code' from each 'data' object as the key.
    """
    block_mapping = {}

    for tenant_id in tenant_ids:
        try:
            search_url = f"{mdms_url}/egov-mdms-service/v1/_search"
            search_payload = {
                "RequestInfo": request_info.model_dump(by_alias=True, exclude_none=True),
                "MdmsCriteria": {
                    "tenantId": tenant_id,
                    "moduleDetails": [
                        {
                            "moduleName": "Incident",
                            "masterDetails": [
                                {
                                    "name": "Block"
                                }
                            ]
                        }
                    ]
                }
            }

            response = requests.post(search_url, json=search_payload)
            if response.status_code == 200:
                data = response.json()
                mdms_blocks = data.get("MdmsRes", {}).get("Incident", {}).get("Block", [])

                for block in mdms_blocks:
                    code = block.get("code")
                    if code and code not in block_mapping:
                        block_mapping[code] = block

        except Exception as e:
            logger.error(f"Error fetching block mapping from MDMS for tenant {tenant_id}: {e}")

    return block_mapping

def create_mapping_dicts(mapping_file: UploadFile, sheet_name: str):
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_file:
        temp_file.write(mapping_file.file.read())
        temp_file_path = temp_file.name

    mapping_df = pd.read_excel(temp_file_path, sheet_name=sheet_name)
    mapping_df.columns = mapping_df.columns.str.strip()
    mapping_df = mapping_df.astype(str).apply(lambda x: x.str.strip())

    os.unlink(temp_file_path)

    subtype_mapping = {
        (row['Existing Issue Type'], row['Existing Ticket Sub Type ( Saure eMitra)']):
        (row['New Issue Type'], row['New Ticket Sub type'])
        for _, row in mapping_df.iterrows()
    }

    return subtype_mapping

def get_user_info_for_mizoram(usernames: List[str], db_conn) -> Dict[str, str]:
    try:
        with db_conn.cursor() as cursor:
            sql = "SELECT username, tenantid FROM eg_user WHERE username = ANY (%s)"
            cursor.execute(sql, (usernames,))
            rows = cursor.fetchall()
            return {row[0]: row[1] for row in rows}
    except Exception as e:
        logger.error(f"Error fetching user info for Mizoram: {e}")
        return {}


@router.post("/legacy_ticket_ingestion", summary="Upload and ingest legacy tickets Excel file")
async def upload_legacy_ticket_excel_sheet(
    legacy_ticket_file: UploadFile = File(...),
    legacy_ticket_sheet_name: str = Form(default="Legacy Tickets"),
    mapping_type_subtype_file: UploadFile = File(...),
    mapping_type_subtype_sheet_name: str = Form(default="Mapping Old_New_v1.0"),
    request_info: str = Form(default="")
):
    migration_id = str(uuid.uuid4())
    request_info_obj = request_info_from_json(request_info)
    get_authorized_request_info(request_info_obj)

    subtype_mapping = create_mapping_dicts(mapping_type_subtype_file, mapping_type_subtype_sheet_name)
    tenant_creator_mapping = TENANT_CREATOR_MAPPING

    input_temp_file, _ = await _save_upload_to_temp_file(legacy_ticket_file, suffix=".xlsx")
    excel_file_path = input_temp_file.name

    df = pd.read_excel(excel_file_path, sheet_name=legacy_ticket_sheet_name)
    df.columns = df.columns.str.strip()
    df = df.reindex(columns=df.columns.tolist() + ['ticket_id', 'employee_info'], fill_value='')

    unique_states = df["State"].dropna().str.strip().unique()
    tenant_ids = [tenant_creator_mapping.get(state, {}).get("tenantId") for state in unique_states]

    tenant_mapping = get_tenant_mapping(request_info_obj, tenant_ids)
    block_mapping = get_block_mapping_from_mdms(request_info_obj, tenant_ids)

    conn = psycopg2.connect(**DB_CONFIG)
    codes = [str(row.get("NIN_HFR ID", "")).strip() for i, row in df.iterrows()
             if str(df.at[i, 'status']).strip().lower() not in ['duplicate', 'error']]
    employee_info = get_hrms_employee_info(codes, conn)

    usernames = [str(row.get("Actual User Name", "")).strip() for i, row in df.iterrows()
                 if str(row.get("State", "")).strip() == "Mizoram" and str(df.at[i, 'status']).strip().lower() not in ['duplicate', 'error']]
    user_info = get_user_info_for_mizoram(usernames, conn)

    for idx, row in df.iterrows():
        try:
            status = str(df.at[idx, 'status']).strip().lower()
            if status in ['duplicate', 'error']:
                continue

            state = str(row.get("State", "")).strip()
            if state == "Mizoram":
                identifier = str(row.get("Actual User Name", "")).strip()
                tenant_id = user_info.get(identifier)
            else:
                identifier = str(row.get("NIN_HFR ID", "")).strip()
                tenant_id = employee_info.get(identifier)

            if not tenant_id:
                df.at[idx, 'status'] = 'failed'
                df.at[idx, 'error'] = f'Employee not found for code: {identifier}'
                df.at[idx, 'employee_info'] = 'Not found'
                continue

            df.at[idx, 'employee_info'] = 'Found'

            tenant_details = tenant_mapping.get(tenant_id, {})
            if not tenant_details:
                df.at[idx, 'status'] = 'failed'
                df.at[idx, 'error'] = f'Tenant mapping not found for tenant ID: {tenant_id}'
                continue

            incident_payload = build_incident_payload(row, identifier, tenant_details, block_mapping, migration_id,
                                                      tenant_creator_mapping.get(state, {}), subtype_mapping)
            response = submit_incident_payload(incident_payload, tenant_creator_mapping.get(state, {}))
            process_response(response, df, idx, identifier)

        except Exception as e:
            df.at[idx, 'status'] = 'failed'
            df.at[idx, 'error'] = str(e)

    return write_and_return_excel(df, legacy_ticket_sheet_name)

def build_incident_payload(row, identifier, tenant_details, block_mapping, migration_id, creator_info, subtype_mapping):
    ticket_type = str(row.get("Ticket Type", "")).strip()
    ticket_subtype = str(row.get("Ticket Sub Type", "")).strip()
    system_functional = {"Yes": "FUNCTIONAL", "No": "NON_FUNCTIONAL"}.get(
        str(row.get("Is the solar system working?", "")).strip(), "")
    comments = str(row.get("Comments", "")).strip()[:256]
    mapped_pair = subtype_mapping.get((ticket_type, ticket_subtype))

    if not ticket_type or not ticket_subtype or not mapped_pair:
        raise ValueError("Missing or invalid Ticket Type/Sub Type")

    block_code = tenant_details.get("city", {}).get("blockCode", "")
    block = block_mapping.get(block_code, {}).get("name", "")

    incident_payload = {
        "incidentType": mapped_pair[0],
        "incidentSubtype": mapped_pair[1],
        "comments": comments,
        "systemFunctional": system_functional,
        "tenantId": tenant_details.get("code", ""),
        "migrationId": migration_id,
        "district": tenant_details.get("city", {}).get("districtCode", ""),
        "block": block,
        "phcType": tenant_details.get("code", ""),
        "phcSubType": tenant_details.get("centreType", ""),
        "additionalDetail": {"fileStoreId": [], "reopenreason": [], "rejectReason": [],
                              "sendBackReason": [], "sendBackSubReason": []},
        "source": "web",
        "reporter": {
            "uuid": creator_info.get("uuid"),
            "tenantId": creator_info.get("tenantId")
        }
    }

    if pd.notnull(row.get("Unique_ID")):
        incident_payload["legacyId"] = str(row.get("Unique_ID")).strip()

    reported_date = row.get("Actual_Reported_Date (mm/dd/yyyy)", None)
    if pd.notnull(reported_date):
        dt = pd.to_datetime(reported_date, format="%d/%m/%Y", errors='coerce') if isinstance(reported_date, str) else pd.to_datetime(reported_date, errors='coerce')
        if pd.notnull(dt):
            incident_payload["filedDate"] = int(dt.timestamp() * 1000)

    return incident_payload

def submit_incident_payload(payload, creator):
    profile = USER_PROFILE

    return requests.post(
        f"{im_services_url}/im-services/v2/request/_create",
        json={
            "RequestInfo": {
                "apiId": "Rainmaker",
                "authToken": "79967889-fbf5-42c6-9bd3-4adc0dbe7692",
                "userInfo": {
                    "id": creator.get("id"),
                    "uuid": creator.get("uuid"),
                    "userName": profile["userName"],
                    "name": profile["name"],
                    "mobileNumber": creator.get("mobileNumber"),
                    "emailId": None,
                    "locale": None,
                    "type": "EMPLOYEE",
                    "roles": [
                        {
                            "name": "Complainant",
                            "code": "COMPLAINANT",
                            "tenantId": creator.get("tenantId")
                        },
                        {
                            "name": "Employee",
                            "code": "EMPLOYEE",
                            "tenantId": creator.get("tenantId")
                        }
                    ],
                    "active": True,
                    "tenantId": creator.get("tenantId"),
                    "permanentCity": None
                },
                "msgId": "1744021633700|en_IN",
                "plainAccessRequest": {}
            },
            "incident": payload,
            "workflow": {"action": "APPLY", "verificationDocuments": []}
        },
        headers={"Content-Type": "application/json"}
    )

def process_response(response, df, idx, identifier):
    if response.status_code in [200, 201]:
        incident = response.json().get("IncidentWrappers", [{}])[0].get("incident")
        df.at[idx, 'status'] = 'success'
        df.at[idx, 'error'] = ''
        df.at[idx, 'ticket_id'] = incident.get("incidentId", '')
    else:
        error_msg = response.json().get('Errors', [{}])[0].get('message', response.text)
        df.at[idx, 'status'] = 'failed'
        df.at[idx, 'error'] = error_msg

def write_and_return_excel(df, sheet_name):
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = f"/tmp/legacy_ticket_ingestion_results_{timestamp}.xlsx"
    df.to_excel(output_path, sheet_name=sheet_name, index=False)
    return FileResponse(output_path, filename=os.path.basename(output_path), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@router.post("/check_duplicates")
async def check_duplicate_tickets(
        legacy_ticket_file: UploadFile = File(...),
        legacy_ticket_sheet_name: str = Form(default="Duplication Template"),
):
    input_temp_file = None
    try:
        # Save uploaded file temporarily
        input_temp_file, _ = await _save_upload_to_temp_file(legacy_ticket_file, suffix=".xlsx")
        excel_path = input_temp_file.name

        # Read Excel file
        df = pd.read_excel(excel_path, sheet_name=legacy_ticket_sheet_name)
        df.columns = df.columns.str.strip()
        df = df.reindex(columns=df.columns.tolist() + ['status', 'error'], fill_value='')

        conn = psycopg2.connect(**DB_CONFIG)
        cursor = conn.cursor()

        # Step 1: Fetch NIN_HFR ID -> tenantId mapping
        nin_hfr_ids = df["NIN_HFR ID"].dropna().astype(str).str.strip().unique().tolist()
        cursor.execute("SELECT code, tenantid FROM eg_hrms_employee WHERE code IN %s", (tuple(nin_hfr_ids),))
        code_tenant_map = dict(cursor.fetchall())

        # Step 2: Check each row for incident duplication
        for idx, row in df.iterrows():
            code = str(row.get("NIN_HFR ID", "")).strip()
            ticket_type = str(row.get("Ticket Type", "")).strip()
            ticket_subtype = str(row.get("Ticket Sub Type", "")).strip()

            tenant_id = code_tenant_map.get(code)
            if not tenant_id:
                df.at[idx, 'status'] = 'error'
                df.at[idx, 'error'] = 'Invalid NIN_HFR ID (not in eg_hrms_employee)'
                continue

            # Step 3: Check for matching incidents
            cursor.execute("""
                SELECT 1 FROM eg_incident_v2
                WHERE tenantid = %s
                AND incidenttype = %s
                AND incidentsubtype = %s
                AND applicationstatus NOT IN ('CLOSEDAFTERRESOLUTION', 'RESOLVED', 'REJECTED')
                LIMIT 1
            """, (tenant_id, ticket_type, ticket_subtype))
            exists = cursor.fetchone()

            if exists:
                df.at[idx, 'status'] = 'duplicate'

        conn.close()

        # Save updated Excel
        df.to_excel(excel_path, index=False, sheet_name=legacy_ticket_sheet_name)

        return FileResponse(
            path=excel_path,
            filename=legacy_ticket_file.filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error during duplicate check: {str(e)}")

    finally:
        if input_temp_file and os.path.exists(input_temp_file.name):
            pass


def get_request_info_to_send_back_workflow():
    return {
        "apiId": "project-api",
        "ver": "1.0",
        "ts": "",
        "action": "update",
        "did": "",
        "key": "",
        "msgId": "20240617",
        "authToken": "f6a27ba4-bead-483d-b4d8-23d46c74d153",
        "userInfo": {
            "id": 178,
            "uuid": "72743f47-9f1a-47de-ac43-b12cde70afc1",
            "userName": "dummy_manager",
            "name": "dummy_manager",
            "mobileNumber": "9911223345",
            "emailId": None,
            "locale": None,
            "type": "EMPLOYEE",
            "roles": [
                {"name": "Installation Report Part A editor", "code": "INSTALLATION_REPORT_PART_A_EDITOR",
                 "tenantId": "in"},
                {"name": "Installation Report Part B editor", "code": "INSTALLATION_REPORT_PART_B_EDITOR",
                 "tenantId": "in"},
                {"name": "Installation Report Part A reviewer", "code": "INSTALLATION_REPORT_PART_A_REVIEWER",
                 "tenantId": "in"},
                {"name": "Project manager", "code": "PROJECT_MANAGER", "tenantId": "in"},
                {"name": "Installation Report Approver QC team", "code": "INSTALLATION_REPORT_APPROVER_QC_TEAM",
                 "tenantId": "in"}
            ],
            "active": True,
            "tenantId": "in",
            "permanentCity": None
        }
    }

@router.post("/flag_for_qc")
async def flag_for_qc(
        facility_file: UploadFile = File(...),
        facility_sheet_name: str = Form(default="Facilities"),
):
    input_temp_file = None
    try:
        # Save uploaded file temporarily
        input_temp_file, _ = await _save_upload_to_temp_file(facility_file, suffix=".xlsx")
        excel_path = input_temp_file.name

        # Read Excel file
        df = pd.read_excel(excel_path, sheet_name=facility_sheet_name)
        df.columns = df.columns.str.strip()

        # Add system columns for audit/error tracking
        if 'status' not in df.columns:
            df['status'] = ''
        if 'error' not in df.columns:
            df['error'] = ''
        if 'auditTrail' not in df.columns:
            df['auditTrail'] = ''


        for idx, row in df.iterrows():
            business_id = row.get("BusinessId")

            # Prepare request body for workflow API
            payload = {
                "RequestInfo" : get_request_info_to_send_back_workflow(),
                "projectId": business_id,
                "workflow": {
                    "action": "REMOVE_FLAG",
                }
            }

            # Call workflow API
            workflow_update = f"{project_service_url}/project/v1/project/workflow/update"
            try:
                resp = requests.post(workflow_update, json=payload, headers={"Content-Type": "application/json"})
                if resp.status_code != 200:
                    df.at[idx, "error"] = f"WF API failed: {resp.status_code} {resp.text}"
                else:
                    df.at[idx, "auditTrail"] = f"Sent back to pending approval"
            except Exception as e:
                df.at[idx, "error"] = f"WF API call error: {str(e)}"

        df.to_excel(excel_path, index=False, sheet_name=facility_sheet_name)

        return FileResponse(
            path=excel_path,
            filename=facility_file.filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error during facility status update: {str(e)}")

    finally:
        if input_temp_file and os.path.exists(input_temp_file.name):
            pass

@router.post('/incidents/dataUpdate',
             summary='Update incidents data from Excel file',
             response_description='Returns result status for each incident')
async def update_incidents_data_from_excel(
        incidents_file: UploadFile = File(..., description="Excel file containing incidents to update data"),
        incidents_sheet_name: str = Form(default="Incidents",
                                         description="Name of the sheet containing incident data"),
        request_info: str = Form(default="", description="Request info in JSON format")
):
    temp_file = None
    request_info = request_info_from_json(request_info)
    #get_authorized_request_info(request_info)

    try:
        temp_file, _ = await _save_upload_to_temp_file(incidents_file, suffix=".xlsx")

        df = pd.read_excel(temp_file.name, sheet_name=incidents_sheet_name)
        df.columns = df.columns.str.strip()


        for col in ['status', 'error']:
            if col not in df.columns:
                df[col] = ''

        incident_client = IMServiceClient(im_services_url)

        for index, row in df.iterrows():
            if pd.isna(row.get('Ticket No.')):
                df.at[index, 'status'] = 'skipped'
                df.at[index, 'error'] = 'Missing ticket_no'
                continue

            if pd.isna(row.get('Tenant ID')):
                df.at[index, 'status'] = 'skipped'
                df.at[index, 'error'] = 'Missing Tenant ID'
                continue

            incident_request_info = get_incident_request_info()

            try:
                search_response = incident_client.search_incident(
                    incident_id=row['Ticket No.'].strip(),
                    tenant_id=row['Tenant ID'].strip(),
                    request_info=incident_request_info
                )

                incident_wrappers = search_response.get("IncidentWrappers", [])
                if not incident_wrappers:
                    df.at[index, 'status'] = 'failed'
                    df.at[index, 'error'] = f"No incident found for Ticket No. {row['Ticket No.']} and Tenant ID {row['Tenant ID']}"
                    continue

                update_data = {
                    "systemFunctional": (
                        {"yes": "FUNCTIONAL", "no": "NON_FUNCTIONAL"}.get(str(row.get("Is the solar system working?", "")).strip().lower(), "")
                    )
                }

                update_payload = create_update_payload(search_response, update_data)
                update_response = incident_client.update_incident_data(update_payload)

                process_update_incident_data_response(update_response, df, index)

            except Exception as e:
                df.at[index, 'status'] = 'failed'
                df.at[index, 'error'] = str(e)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"incident_data_update_results_{timestamp}.xlsx"

        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as output_temp_file:
            df.to_excel(output_temp_file.name, sheet_name=incidents_sheet_name, index=False)

        return FileResponse(
            path=output_temp_file.name,
            filename=output_filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Failed to process incident updates: {str(e)}"
        ) from e
    finally:
        if temp_file and os.path.exists(temp_file.name):
            os.unlink(temp_file.name)

def process_update_incident_data_response(response, df, idx):
    try:
        if 'Errors' in response and response['Errors']:
            error_msg = response['Errors'][0].get('message', str(response['Errors'][0]))
            df.at[idx, 'status'] = 'failed'
            df.at[idx, 'error'] = error_msg
        else:
            df.at[idx, 'status'] = 'success'
            df.at[idx, 'error'] = ''
    except Exception as e:
        df.at[idx, 'status'] = 'failed'
        df.at[idx, 'error'] = str(e)


@router.post('/facilitiesValidateData',
             summary='Validate facility Excel file before processing',
             response_description='Returns validation report Excel with PASSED/FAILED rows')
async def validate_facilities_excel_sheet(
        background_tasks: BackgroundTasks,
        facility_file: UploadFile = File(..., description="Excel file containing facility data"),
        project_id: str = Form(description="Project ID"),
        facility_sheet_name: str = Form(default="FacilityMapping",
                                        description="Name of the sheet containing facility data"),
        boundary_sheet_name: str = Form(default="BoundaryCodes",
                                        description="Name of the sheet containing boundary data"),
        request_info: str = Form(default="")
):
    temp_input_file = None
    request_info_obj = request_info_from_json(request_info)
    mdms_client = MDMSClient(mdms_url)
    facility_client = FacilityServiceClient(facility_service_url)
    project_client = ProjectServiceClient(project_service_url)

    try:
        # Save uploaded Excel to a temp file
        temp_input_file, _ = await _save_upload_to_temp_file(facility_file, suffix=".xlsx")

        # Load workbook to preserve everything
        wb = load_workbook(temp_input_file.name)

        # ----------------- Read Boundary Sheet ----------------- #
        if boundary_sheet_name not in wb.sheetnames:
            raise HTTPException(status_code=400, detail=f"Boundary sheet '{boundary_sheet_name}' not found")

        boundary_data_df = pd.read_excel(temp_input_file.name, sheet_name=boundary_sheet_name)

        # ----------------- Validate Boundary Sheet Against Project ----------------- #
        projects = project_client.search_project(request_info_obj, project_id)
        if not projects or "Project" not in projects or len(projects["Project"]) == 0:
            raise HTTPException(status_code=400, detail=f"No project found for id {project_id}")

        project = projects["Project"][0]["project"]
        geography = project.get("additionalDetails", {}).get("geographyDetails", {})

        # Valid codes directly as a set (no loop needed)
        valid_boundary_codes = {str(block["code"]).strip() for block in geography.get("blocks", []) if
                                block.get("code")}

        # Uploaded codes directly as a set
        uploaded_codes = set(boundary_data_df["BoundaryCode"].dropna().astype(str).str.strip())

        # Equality check
        if uploaded_codes != valid_boundary_codes:
            raise HTTPException(
                status_code=400,
                detail={
                    "error": "BoundaryCode mismatch",
                    "missing": list(valid_boundary_codes - uploaded_codes),
                    "extra": list(uploaded_codes - valid_boundary_codes)
                }
            )

        # ----------------- Read Facility Sheet ----------------- #
        if facility_sheet_name not in wb.sheetnames:
            raise HTTPException(status_code=400, detail=f"Facility sheet '{facility_sheet_name}' not found")

        df = pd.read_excel(temp_input_file.name, sheet_name=facility_sheet_name)
        df.columns = [str(c).strip() for c in df.columns]
        df = df.loc[:, ~df.columns.str.startswith('Unnamed')]

        # ----------------- Read Facility Column ----------------- #
        if 'Facility Id' not in df.columns:
            raise HTTPException(status_code=400, detail=f"Facility Column in '{facility_sheet_name}' not found")

        # Ensure status/error columns exist
        if 'status' not in df.columns:
            df['status'] = ''
        if 'error' not in df.columns:
            df['error'] = ''

        # ----------------- Run Validation ----------------- #
        validation_errors = project_facility_validation(
            df,
            mdms_client,
            request_info_obj,
            facility_client,
            boundary_data_df,
            'data-ingestion.FacilityIngestionSchema'
        )

        # Mark rows based on validation results
        error_count = 0
        for i, errs in enumerate(validation_errors):
            if errs:
                df.at[i, 'status'] = 'FAILED'
                df.at[i, 'error'] = "; ".join(dict.fromkeys(errs))
                error_count += 1
            else:
                df.at[i, 'status'] = 'PASSED'
                df.at[i, 'error'] = ''

        # ----------------- Update Facility Sheet In-Place ----------------- #
        ws = wb[facility_sheet_name]
        header_values = [cell.value for cell in ws[1]]

        # Add status/error columns if missing
        for col_name in ["status", "error"]:
            if col_name not in header_values:
                new_col_idx = len(header_values) + 1
                cell = ws.cell(row=1, column=new_col_idx, value=col_name)
                cell.font = Font(bold=True)
                header_values.append(col_name)

                # lock header cell
                cell.protection = Protection(locked=True)

                # lock all data cells in this new column
                for r_idx in range(2, ws.max_row + 1):
                    ws.cell(row=r_idx, column=new_col_idx).protection = Protection(locked=True)

        grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        # Write data rows back (without header row)
        export_df = prepare_dataframe_for_excel_export(df)
        for r_idx, row in enumerate(dataframe_to_rows(export_df, index=False, header=False), start=2):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)

                # force lock for status/error columns
                if ws.cell(1, c_idx).value in ["status", "error"]:
                    cell.protection = Protection(locked=True)
                    cell.fill = grey_fill

        # Ensure sheet protection is ON
        ws.protection.sheet = True
        ws.protection.enable()

        # ----------------- Save to new temp file ----------------- #
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_temp_file_path = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx").name
        wb.save(output_temp_file_path)

        autofit_columns(output_temp_file_path, facility_sheet_name, auto_fit=True)

        background_tasks.add_task(cleanup_temp_file, output_temp_file_path)

        response = FileResponse(
            path=output_temp_file_path,
            filename=f"facility_validation_results_{timestamp}.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response.headers["X-Error-Count"] = str(error_count)

        return response

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Validation failed: {str(e)}")
    finally:
        if temp_input_file and os.path.exists(temp_input_file.name):
            os.unlink(temp_input_file.name)


@router.post('/fieldPlanfacilitiesValidateData',
             summary='Validate facility Excel file before processing',
             response_description='Returns validation report Excel with PASSED/FAILED rows')
async def validate_facilities_excel_sheet(
        background_tasks: BackgroundTasks,
        facility_file: UploadFile = File(..., description="Excel file containing facility data"),
        facility_sheet_name: str = Form(default="FacilityMapping",
                                        description="Name of the sheet containing facility data"),
        boundary_sheet_name: str = Form(default="BoundaryCodes",
                                        description="Name of the sheet containing boundary data"),
        request_info: str = Form(default=""),
        project_id: str = Form(default="", description="Project ID (required for assessment handoff validation)"),
        assessment_plan_ids: str = Form(default="", description="JSON array or comma-separated assessment plan IDs"),
        tenant_id: str = Form(default="in"),
):
    temp_input_file = None
    request_info_obj = request_info_from_json(request_info)
    mdms_client = MDMSClient(mdms_url)
    facility_client = FacilityServiceClient(facility_service_url)

    try:
        # Save uploaded Excel to a temp file
        temp_input_file, _ = await _save_upload_to_temp_file(facility_file, suffix=".xlsx")

        # Load workbook to preserve everything
        wb = load_workbook(temp_input_file.name)

        # ----------------- Read Boundary Sheet ----------------- #
        if boundary_sheet_name not in wb.sheetnames:
            raise HTTPException(status_code=400, detail=f"Boundary sheet '{boundary_sheet_name}' not found")

        boundary_data_df = pd.read_excel(temp_input_file.name, sheet_name=boundary_sheet_name)

        # ----------------- Read Facility Sheet ----------------- #
        if facility_sheet_name not in wb.sheetnames:
            raise HTTPException(status_code=400, detail=f"Facility sheet '{facility_sheet_name}' not found")

        df = pd.read_excel(temp_input_file.name, sheet_name=facility_sheet_name)
        df.columns = [str(c).strip() for c in df.columns]
        df = df.loc[:, ~df.columns.str.startswith('Unnamed')]

        # ----------------- Read Facility Column ----------------- #
        if 'Facility Id' not in df.columns:
            raise HTTPException(status_code=400, detail=f"Facility Column in '{facility_sheet_name}' not found")

        # Ensure status/error columns exist
        if 'status' not in df.columns:
            df['status'] = ''
        if 'error' not in df.columns:
            df['error'] = ''

        # ----------------- Run Validation ----------------- #
        validation_errors = field_plan_facility_validation(
            df,
            mdms_client,
            request_info_obj,
            facility_client,
            boundary_data_df,
            'data-ingestion.FieldPlanFacilityIngestionSchema'
        )

        parsed_assessment_plan_ids = parse_assessment_plan_ids(assessment_plan_ids)
        if parsed_assessment_plan_ids:
            if not project_id:
                raise HTTPException(
                    status_code=400,
                    detail="project_id is required when assessment_plan_ids is provided",
                )
            assessment_client = AssessmentServiceClient(fieldPlan_service_url)
            eligible_map = load_eligible_facility_map(
                assessment_client,
                request_info_obj,
                project_id,
                tenant_id,
                parsed_assessment_plan_ids,
            )
            assessment_errors = validate_assessment_handoff_rows(df, eligible_map)
            validation_errors = merge_assessment_validation_errors(validation_errors, assessment_errors)

        # Mark rows based on validation results
        error_count = 0
        for i, errs in enumerate(validation_errors):
            if errs:
                df.at[i, 'status'] = 'FAILED'
                df.at[i, 'error'] = "; ".join(dict.fromkeys(errs))
                error_count += 1
            else:
                df.at[i, 'status'] = 'PASSED'
                df.at[i, 'error'] = ''

        # ----------------- Update Facility Sheet In-Place ----------------- #
        ws = wb[facility_sheet_name]
        header_values = [cell.value for cell in ws[1]]

        # Add status/error columns if missing
        for col_name in ["status", "error"]:
            if col_name not in header_values:
                new_col_idx = len(header_values) + 1
                cell = ws.cell(row=1, column=new_col_idx, value=col_name)
                cell.font = Font(bold=True)
                header_values.append(col_name)

                # lock header cell
                cell.protection = Protection(locked=True)

                # lock all data cells in this new column
                for r_idx in range(2, ws.max_row + 1):
                    ws.cell(row=r_idx, column=new_col_idx).protection = Protection(locked=True)

        grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        # Write data rows back (without header row)
        export_df = prepare_dataframe_for_excel_export(df)
        for r_idx, row in enumerate(dataframe_to_rows(export_df, index=False, header=False), start=2):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)

                # force lock for status/error columns
                if ws.cell(1, c_idx).value in ["status", "error"]:
                    cell.protection = Protection(locked=True)
                    cell.fill = grey_fill

        # Ensure sheet protection is ON
        ws.protection.sheet = True
        ws.protection.enable()

        # ----------------- Save to new temp file ----------------- #
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_temp_file_path = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx").name
        wb.save(output_temp_file_path)

        autofit_columns(output_temp_file_path, facility_sheet_name, auto_fit=True)

        background_tasks.add_task(cleanup_temp_file, output_temp_file_path)

        response = FileResponse(
            path=output_temp_file_path,
            filename=f"facility_validation_results_{timestamp}.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response.headers["X-Error-Count"] = str(error_count)

        return response

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Validation failed: {str(e)}")
    finally:
        if temp_input_file and os.path.exists(temp_input_file.name):
            os.unlink(temp_input_file.name)


@router.post('/createFacilityAndUpdateProject',
             summary='Create passed facility in Excel file and add them to project',
             response_description='Created facilities from PASSED rows and added to the given project if selected')
async def create_facilities_and_update_project(
        background_tasks: BackgroundTasks,
        facility_file: UploadFile = File(description="Validated Excel file with PASSED/FAILED status"),
        facility_sheet_name: str = Form(default="FacilityMapping",
                                        description="Name of the sheet containing facility data"),
        project_id: str = Form(description="Project ID"),
        request_info: str = Form(default="")
):
    input_temp_file = None

    # parse
    request_info = request_info_from_json(request_info)
    mdms_client = MDMSClient(mdms_url)

    try:
        # ---------- save uploaded file ----------
        input_temp_file, uploaded_size = await _save_upload_to_temp_file(facility_file, suffix=".xlsx")
        facility_file_path = input_temp_file.name
        logger.info(f"Received createFacilityAndUpdateProject file of size {uploaded_size} bytes")

        # ---------- prepare output path & load workbook ----------
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"facility_creation_and_project_update_results_{timestamp}.xlsx"
        output_temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        output_temp_file.close()
        output_file_path = output_temp_file.name

        wb = load_workbook(facility_file_path)
        if facility_sheet_name not in wb.sheetnames:
            raise HTTPException(status_code=400, detail=f"Sheet '{facility_sheet_name}' not found")
        ws = wb[facility_sheet_name]

        # ---------- read sheet into DataFrame ----------
        df = pd.read_excel(facility_file_path, sheet_name=facility_sheet_name)
        df.columns = [str(c).strip() for c in df.columns]

        # sanity checks
        # status/error are validation artifacts from previous step
        if 'status' not in df.columns or 'error' not in df.columns:
            raise HTTPException(status_code=400, detail="Missing 'status'/'error' columns. Please upload validated file.")

        # Ensure all rows are PASSED
        failed_rows = df[df['status'].str.upper() != 'PASSED']
        if not failed_rows.empty:
            raise HTTPException(
                status_code=400,
                detail="Validation failed: Some rows are not marked as PASSED. Please upload a fully validated file."
            )

        # helper to find a column by partial name (case insensitive)
        def find_col(partial):
            for c in df.columns:
                if partial.lower() in str(c).lower():
                    return c
            return None

        include_col = find_col("Include in Project")
        facility_id_col = find_col("Facility Id") or "Facility Id"
        status_col = find_col("status") or "status"

        # add result columns if missing
        if 'Facility Creation Status' not in df.columns:
            df['Facility Creation Status'] = ''
        if 'Project Linking Status' not in df.columns:
            df['Project Linking Status'] = ''

        facility_client = FacilityServiceClient(facility_service_url)
        project_client = ProjectServiceClient(project_service_url)
        facility_schema = mdms_client.get_column_definitions_with_metadata(
            request_info, 'data-ingestion.FacilityIngestionSchema'
        )

        # --- NEW: fetch already linked facilities once ---
        linked_facilities_resp = project_client.search_project_facility(request_info, project_id)
        linked_facilities = linked_facilities_resp.get("ProjectFacilities", []) if linked_facilities_resp else []
        linked_facility_ids = {pf.get("facilityId") for pf in linked_facilities if pf.get("facilityId")}

        creation_tasks = []
        pending_bulk_links = []
        existing_or_skipped_indexes = []
        for index, row in df.iterrows():
            include_val = ''
            if include_col:
                include_val = str(row.get(include_col, "")).strip().lower()
            else:
                include_val = str(row.get("Include in Project (Mandatory)", "")).strip().lower()
            should_link = include_val == "yes"

            facility_id_val = row.get(facility_id_col, None)
            facility_id = str(facility_id_val).strip() if pd.notna(facility_id_val) and str(facility_id_val).strip() else None
            row_status = str(row.get(status_col, "")).strip().upper()

            if facility_id:
                existing_or_skipped_indexes.append((index, row, should_link, facility_id))
            elif row_status != "PASSED":
                df.at[index, 'Facility Creation Status'] = "Skipped (Validation not PASSED)"
                df.at[index, 'Project Linking Status'] = "Not Attempted"
            else:
                creation_tasks.append((index, row.copy(), should_link))

        for index, row, should_link, facility_id in existing_or_skipped_indexes:
            try:
                df.at[index, 'Facility Creation Status'] = "Already Exists"
                if facility_id in linked_facility_ids:
                    if should_link:
                        df.at[index, 'Project Linking Status'] = "Already Linked"
                    else:
                        try:
                            project_facility_data = next((pf for pf in linked_facilities if pf.get("facilityId") == facility_id), None)
                            project_client.unlink_project_facility(
                                request_info=request_info,
                                project_id=project_id,
                                facility_id=facility_id,
                                project_facility_data=project_facility_data
                            )
                            df.at[index, 'Project Linking Status'] = "Unlinked"
                            linked_facility_ids.remove(facility_id)
                        except Exception as e:
                            df.at[index, 'Project Linking Status'] = f"Exception during unlink: {str(e)}"
                else:
                    if should_link:
                        pending_bulk_links.append((index, facility_id))
                    else:
                        df.at[index, 'Project Linking Status'] = "Skipped (Include in Project != Yes)"
            except Exception as e:
                df.at[index, 'Facility Creation Status'] = f"Exception: {str(e)}"
                df.at[index, 'Project Linking Status'] = "Not Attempted"
                continue

        if creation_tasks:
            logger.info(f"Processing {len(creation_tasks)} new facilities using bulk create API")
            bulk_payload = {
                "RequestInfo": request_info.model_dump(by_alias=True, exclude_none=True),
                "facilities": []
            }
            creation_meta = []

            for idx, row_data, link_required in creation_tasks:
                single_payload = create_facility_payload(request_info, row_data, False, facility_schema)
                facilities = single_payload.get("facilities", [])
                if facilities:
                    bulk_payload["facilities"].append(facilities[0])
                    creation_meta.append((idx, link_required))
                else:
                    df.at[idx, 'Facility Creation Status'] = "Failed: Invalid facility payload"
                    df.at[idx, 'Project Linking Status'] = "Not Attempted"

            create_resp = None
            try:
                if bulk_payload["facilities"]:
                    create_resp = facility_client.create_facility(bulk_payload)
            except Exception as exc:
                for idx, _ in creation_meta:
                    df.at[idx, 'Facility Creation Status'] = f"Exception during bulk create: {str(exc)}"
                    df.at[idx, 'Project Linking Status'] = "Not Attempted"

            if create_resp is not None:
                if create_resp.status_code in (200, 201):
                    created_facilities = []
                    try:
                        created_facilities = create_resp.json() or []
                    except Exception as exc:
                        logger.warning(f"Could not parse bulk create response JSON: {exc}")

                    for result_idx, (row_idx, link_required) in enumerate(creation_meta):
                        created_id = None
                        if result_idx < len(created_facilities):
                            created_id = created_facilities[result_idx].get("facility_id")

                        creation_status = "Created" if created_id else "Created (id missing)"
                        df.at[row_idx, 'Facility Creation Status'] = creation_status

                        if created_id:
                            df.at[row_idx, facility_id_col] = created_id
                            linked_facility_ids.add(created_id)

                        if link_required and created_id:
                            pending_bulk_links.append((row_idx, created_id))
                        elif link_required and not created_id:
                            df.at[row_idx, 'Project Linking Status'] = "Skipped (no facility id after create)"
                        else:
                            df.at[row_idx, 'Project Linking Status'] = "Skipped (Include in Project != Yes)"
                elif create_resp.status_code == 400:
                    try:
                        error_data = create_resp.json()
                        error_message = error_data.get('Errors', [{}])[0].get('message', 'Unknown error')
                    except Exception:
                        error_message = create_resp.text
                    for idx, _ in creation_meta:
                        df.at[idx, 'Facility Creation Status'] = f"Failed: {error_message}"
                        df.at[idx, 'Project Linking Status'] = "Not Attempted"
                else:
                    for idx, _ in creation_meta:
                        df.at[idx, 'Facility Creation Status'] = f"Failed: {create_resp.status_code} {create_resp.text}"
                        df.at[idx, 'Project Linking Status'] = "Not Attempted"

        # Bulk-link facilities to project (for include=yes rows not already linked)
        if pending_bulk_links:
            chunk_size = BULK_INGEST_CHUNK_SIZE
            for i in range(0, len(pending_bulk_links), chunk_size):
                chunk = pending_bulk_links[i:i + chunk_size]
                chunk_facility_ids = [facility_id for _, facility_id in chunk]
                try:
                    bulk_resp = project_client.create_project_facility_bulk(
                        request_info=request_info,
                        project_id=project_id,
                        facility_ids=chunk_facility_ids
                    )
                    if bulk_resp.status_code in (200, 201, 202):
                        for row_idx, facility_id in chunk:
                            df.at[row_idx, 'Project Linking Status'] = "Linked"
                            linked_facility_ids.add(facility_id)
                    else:
                        for row_idx, _ in chunk:
                            df.at[row_idx, 'Project Linking Status'] = f"Failed: {bulk_resp.status_code} {bulk_resp.text}"
                except Exception as exc:
                    for row_idx, _ in chunk:
                        df.at[row_idx, 'Project Linking Status'] = f"Exception: {str(exc)}"

        # ---------- write results back into workbook preserving formatting ----------
        # Ensure headers exist in sheet (without wiping template)
        header_values = [cell.value for cell in ws[1]]

        for col_name in ["Facility Creation Status", "Project Linking Status"]:
            if col_name not in header_values:
                cell = ws.cell(row=1, column=len(header_values) + 1, value=col_name)
                cell.font = Font(bold=True)
                header_values.append(col_name)
                if col_name not in df.columns:
                    df[col_name] = ""  # ensure column exists in dataframe

        # Delete data rows only (preserve header row and template formatting)
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)

        # Write data rows back (without header row)
        export_df = prepare_dataframe_for_excel_export(df)
        for r_idx, row in enumerate(dataframe_to_rows(export_df, index=False, header=False), start=2):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        wb.save(output_file_path)

        autofit_columns(output_file_path, facility_sheet_name , auto_fit=True)

        background_tasks.add_task(cleanup_temp_file, output_file_path)

        return FileResponse(
            path=output_file_path,
            filename=output_filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        logger.error(f"Error finalizing facility data: {e}")
        raise HTTPException(
            status_code=500,
            detail=f"Failed to finalize facility data: {str(e)}"
        )
    finally:
        if input_temp_file and os.path.exists(input_temp_file.name):
            os.unlink(input_temp_file.name)


@router.post('/createFieldPlanFacility',
             summary='Create passed facility in Excel file and add them to project',
             response_description='Created facilities from PASSED rows and added to the given project if selected')
async def create_fielplan_facilities(
        background_tasks: BackgroundTasks,
        facility_file: UploadFile = File(description="Validated Excel file with PASSED/FAILED status"),
        facility_sheet_name: str = Form(default="FacilityMapping",
                                        description="Name of the sheet containing facility data"),
        fieldplan_id: str = Form(description="FieldPlan ID"),
        request_info: str = Form(default=""),
        project_id: str = Form(default="", description="Project ID (required for assessment handoff apply)"),
        assessment_plan_ids: str = Form(default="", description="JSON array or comma-separated assessment plan IDs"),
        tenant_id: str = Form(default="in"),
):
    input_temp_file = None

    # parse
    request_info = request_info_from_json(request_info)
    mdms_client = MDMSClient(mdms_url)

    try:
        # ---------- save uploaded file ----------
        input_temp_file, uploaded_size = await _save_upload_to_temp_file(facility_file, suffix=".xlsx")
        facility_file_path = input_temp_file.name
        logger.info(f"Received createFieldPlanFacility file of size {uploaded_size} bytes")

        # ---------- prepare output path & load workbook ----------
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"facility_fieldplan_update_results_{timestamp}.xlsx"
        output_temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        output_temp_file.close()
        output_file_path = output_temp_file.name

        wb = load_workbook(facility_file_path)
        if facility_sheet_name not in wb.sheetnames:
            raise HTTPException(status_code=400, detail=f"Sheet '{facility_sheet_name}' not found")
        ws = wb[facility_sheet_name]

        # ---------- read sheet into DataFrame ----------
        df = pd.read_excel(facility_file_path, sheet_name=facility_sheet_name)
        df.columns = [str(c).strip() for c in df.columns]

        # sanity checks
        # status/error are validation artifacts from previous step
        if 'status' not in df.columns or 'error' not in df.columns:
            raise HTTPException(status_code=400, detail="Missing 'status'/'error' columns. Please upload validated file.")

        # Ensure all rows are PASSED
        failed_rows = df[df['status'].str.upper() != 'PASSED']
        if not failed_rows.empty:
            raise HTTPException(
                status_code=400,
                detail="Validation failed: Some rows are not marked as PASSED. Please upload a fully validated file."
            )

        # helper to find a column by partial name (case insensitive)
        def find_col(partial):
            for c in df.columns:
                if partial.lower() in str(c).lower():
                    return c
            return None

        # Editable fields on an already-linked FieldPlanFacility can only be updated while the
        # parent FieldPlan is still a Draft - matches field-planner's FieldPlannerConstants.DRAFT_STATUS.
        DRAFT_FIELD_PLAN_STATUS = "DRAFT"

        # Fields a user may update on an already-linked FieldPlanFacility (facilityType and the
        # facilityId/fieldPlanId link itself stay immutable via this endpoint).
        EDITABLE_ADDITIONAL_FIELD_KEYS = {
            "systemType", "solarSolutionDesignType", "totalSystemCapacity",
            "customSolarSolutionDesignType", "customTotalSystemCapacity",
        }

        def validate_custom_capacity_numeric(value):
            """Same numeric-only rule used for create; returns an error message or None."""
            value_str = str(value).strip() if value not in (None, "") else ""
            if not value_str:
                return None
            try:
                if not math.isfinite(float(value_str)):
                    raise ValueError(value_str)
            except ValueError:
                return f"Error: Custom Total System Capacity '{value_str}' must be numeric"
            return None

        def flatten_additional_fields(additional_fields):
            if not additional_fields:
                return {}
            return {
                f.get("key"): f.get("value")
                for f in additional_fields.get("fields") or []
                if f.get("key")
            }

        include_col = find_col("Included in Field Plan")
        facility_id_col = find_col("Facility Id") or "Facility Id"
        status_col = find_col("status") or "status"
        facility_type_col = find_col("Type of HC")
        system_type_col = find_col("System Type")
        solution_design_type_col = find_col("Solution Design Type")
        total_system_capacity_col = find_col("Total System Capacity")
        custom_solution_design_col = find_col("Custom Solution Design Type")
        custom_total_system_capacity_col = find_col("Custom Total System Capacity")

        # MDMS schema for facilityType/systemType/solarSolutionDesignType/totalSystemCapacity
        # code lookups (see build_field_plan_facility_additional_details) - falls back to raw
        # Excel labels if the schema can't be fetched, same graceful-degradation pattern used
        # elsewhere in this endpoint for other external service calls.
        field_plan_facility_schema = []
        try:
            field_plan_facility_schema = mdms_client.get_column_definitions_with_metadata(
                request_info, 'data-ingestion.FieldPlanFacilityIngestionSchema'
            )
        except Exception as e:
            logger.warning(f"Could not fetch FieldPlanFacilityIngestionSchema for code lookup: {e}")

        # add result columns if missing
        if 'Field Plan Linking Status' not in df.columns:
            df['Field Plan Linking Status'] = ''

        fieldplan_client = FieldPlanServiceClient(fieldPlan_service_url)
        fieldplan_activity_client = FieldPlanActivityServiceClient(fieldPlan_activity_service_url)

        parsed_assessment_plan_ids = parse_assessment_plan_ids(assessment_plan_ids)
        assessment_client = None
        eligible_map = {}
        if parsed_assessment_plan_ids:
            if not project_id:
                raise HTTPException(
                    status_code=400,
                    detail="project_id is required when assessment_plan_ids is provided",
                )
            assessment_client = AssessmentServiceClient(fieldPlan_service_url)
            eligible_map = load_eligible_facility_map(
                assessment_client,
                request_info,
                project_id,
                tenant_id,
                parsed_assessment_plan_ids,
            )

        # Fetch fieldplan-linked facilities if fieldplan_id is provided
        fieldplan_linked_facility_ids = set()
        if fieldplan_id:
            try:
                fieldplan_facilities_response = fieldplan_client.search_fieldplan_facility(request_info, fieldplan_id)
                fieldplan_facilities = fieldplan_facilities_response.get("FieldPlanFacilities", [])
                fieldplan_linked_facility_ids = {pf.get("facilityId") for pf in fieldplan_facilities if
                                                 pf.get("facilityId")}
                logger.info(
                    f"Found {len(fieldplan_linked_facility_ids)} facilities linked to fieldplan {fieldplan_id}")

                # Get FieldPlan status
                fieldplan_response = fieldplan_client.search_fieldPlan(request_info, fieldplan_id)
                fieldplan_data = fieldplan_response.get("FieldPlans", [])
                fieldplan_status = fieldplan_data[0].get("status") if fieldplan_data else None

                fieldplan_assignment_response = fieldplan_activity_client.search_fieldplan_activity_assignment(request_info, fieldplan_id)
                fieldplan_assignment_data = fieldplan_assignment_response.get("ActivitiesAssignments", [])
                role_to_ids = defaultdict(list)

                for item in fieldplan_assignment_data:
                    role = item.get("role")
                    if role:
                        code = role.get("code")
                        if code:
                            role_to_ids[code].append(item.get("assignedTo"))

                pending_bulk_fieldplan_links = []
                pending_assessment_handoffs = []
                pending_bulk_fieldplan_updates = []
                # iterate all rows — handle existing facility ids (linking/unlinking)
                for index, row in df.iterrows():
                    try:
                        # normalize facility id and include flag
                        facility_id_val = row.get(facility_id_col, None)
                        facility_id = None
                        if pd.notna(facility_id_val) and str(facility_id_val).strip():
                            facility_id = str(facility_id_val).strip()

                        include_val = ''
                        if include_col:
                            include_val = str(row.get(include_col, "")).strip().lower()
                        else:
                            include_val = str(row.get("Included in Field Plan (Mandatory)", "")).strip().lower()

                        should_link = include_val == "yes"

                        # ---------- CASE A: existing facility_id present -> skip creation, attempt linking if requested ----------
                        if facility_id:
                            # df.at[index, 'Facility Creation Status'] = "Already Exists"
                            # attempt linking if requested
                            if facility_id in fieldplan_linked_facility_ids:
                                if should_link:
                                    # already linked → check if the editable fields changed and
                                    # need an update, otherwise no-op
                                    fieldPlan_facility_data = next(
                                        (pf for pf in fieldplan_facilities if pf.get("facilityId") == facility_id),
                                        None)
                                    existing_values = flatten_additional_fields(
                                        fieldPlan_facility_data.get("additionalFields") if fieldPlan_facility_data else None
                                    )
                                    row_values = build_field_plan_facility_additional_details(
                                        row,
                                        column_list=field_plan_facility_schema,
                                        system_type_column=system_type_col,
                                        total_system_capacity_column=total_system_capacity_col,
                                        solution_design_type_column=solution_design_type_col,
                                        custom_solution_design_column=custom_solution_design_col,
                                        custom_total_system_capacity_column=custom_total_system_capacity_col,
                                    ) or {}

                                    # Only a key the Excel row actually fills in and that differs
                                    # from the stored value counts as "changed" - a blank cell
                                    # means "leave this field as-is", same semantics used at
                                    # create time (blank optional fields are simply omitted).
                                    changed_keys = {
                                        key for key in EDITABLE_ADDITIONAL_FIELD_KEYS
                                        if row_values.get(key) and row_values.get(key) != existing_values.get(key)
                                    }

                                    if not changed_keys:
                                        df.at[index, 'Field Plan Linking Status'] = "Already Linked"
                                    elif fieldplan_status != DRAFT_FIELD_PLAN_STATUS:
                                        df.at[index, 'Field Plan Linking Status'] = (
                                            f"Error: Cannot update - FieldPlan status must be DRAFT "
                                            f"(current status: {fieldplan_status})"
                                        )
                                    elif fieldPlan_facility_data is None or not fieldPlan_facility_data.get("id"):
                                        df.at[index, 'Field Plan Linking Status'] = (
                                            "Error: Cannot update - FieldPlanFacility record id not found"
                                        )
                                    else:
                                        error = None
                                        if "customTotalSystemCapacity" in changed_keys:
                                            error = validate_custom_capacity_numeric(
                                                row_values.get("customTotalSystemCapacity")
                                            )
                                        if error:
                                            df.at[index, 'Field Plan Linking Status'] = error
                                        else:
                                            additional_fields = build_field_plan_facility_additional_fields(
                                                {key: row_values.get(key) for key in changed_keys}
                                            )
                                            pending_bulk_fieldplan_updates.append(
                                                (
                                                    index,
                                                    {
                                                        "id": fieldPlan_facility_data["id"],
                                                        "facilityId": facility_id,
                                                        "fieldPlanId": fieldplan_id,
                                                        "additionalFields": additional_fields,
                                                    },
                                                )
                                            )
                                else:
                                    # linked but Excel says No → unlink
                                    try:
                                        fieldPlan_facility_data = next(
                                            (pf for pf in fieldplan_facilities if pf.get("facilityId") == facility_id),
                                            None)
                                        fieldplan_client.unlink_fieldplan_facility(
                                            request_info=request_info,
                                            fieldplan_id=fieldplan_id,
                                            facility_id=facility_id,
                                            fieldplan_facility_data=fieldPlan_facility_data
                                        )

                                        facilities_activity_response = fieldplan_activity_client.search_facility_activity(
                                            request_info, fieldplan_id, facility_id)
                                        facilities_activity = facilities_activity_response.get("FacilityActivities",[])
                                        facility_activity_ids = list({fa.get("activityFacility").get("id") for fa in facilities_activity if fa.get("activityFacility").get("id")})
                                        fieldplan_activity_client.delete_facility_activity(request_info=request_info, facility_activity_id=facility_activity_ids)

                                        df.at[index, 'Field Plan Linking Status'] = "Unlinked"
                                        fieldplan_linked_facility_ids.remove(facility_id)
                                    except Exception as e:
                                        df.at[index, 'Field Plan Linking Status'] = f"Exception during unlink: {str(e)}"
                            else:
                                if should_link:
                                    custom_capacity_val = row.get(custom_total_system_capacity_col, None) \
                                        if custom_total_system_capacity_col else None
                                    custom_capacity_str = str(custom_capacity_val).strip() \
                                        if pd.notna(custom_capacity_val) else ""
                                    error = validate_custom_capacity_numeric(custom_capacity_str)
                                    if error:
                                        df.at[index, 'Field Plan Linking Status'] = error
                                        continue

                                    bulk_entry = build_field_plan_facility_bulk_entry(
                                        row,
                                        facility_id,
                                        column_list=field_plan_facility_schema,
                                        facility_type_column=facility_type_col,
                                        system_type_column=system_type_col,
                                        total_system_capacity_column=total_system_capacity_col,
                                        solution_design_type_column=solution_design_type_col,
                                        custom_solution_design_column=custom_solution_design_col,
                                        custom_total_system_capacity_column=custom_total_system_capacity_col,
                                    )
                                    plan_facility_id, _ = extract_assessment_link_meta(row, df)
                                    if (
                                        plan_facility_id
                                        and parsed_assessment_plan_ids
                                        and plan_facility_id in eligible_map
                                    ):
                                        pending_assessment_handoffs.append(
                                            (index, bulk_entry, plan_facility_id)
                                        )
                                    else:
                                        pending_bulk_fieldplan_links.append((index, bulk_entry))
                                else:
                                    df.at[index, 'Field Plan Linking Status'] = "Skipped (Include in Field Plan != Yes)"

                                # continue to next row
                                continue

                    except Exception as e:
                        # any unexpected error per row
                        df.at[index, 'Field Plan Linking Status'] = "Not Attempted"
                        continue

                if pending_bulk_fieldplan_links:
                    chunk_size = BULK_INGEST_CHUNK_SIZE
                    for i in range(0, len(pending_bulk_fieldplan_links), chunk_size):
                        chunk = pending_bulk_fieldplan_links[i:i + chunk_size]
                        facilities_chunk = [entry for _, entry in chunk]
                        try:
                            fieldplan_resp = fieldplan_client.create_fieldPlan_facility_bulk(
                                request_info=request_info,
                                fieldPlan_id=fieldplan_id,
                                facilities=facilities_chunk,
                            )

                            if fieldplan_resp.status_code in (200, 201, 202):
                                for row_idx, entry in chunk:
                                    facility_id = entry["facilityId"]
                                    df.at[row_idx, 'Field Plan Linking Status'] = "Linked"
                                    fieldplan_linked_facility_ids.add(facility_id)

                                    if fieldplan_data:
                                        fieldplan = fieldplan_data[0]
                                        if fieldplan.get("status") == 'SCHEDULED':
                                            try:
                                                facility_activity_resp = fieldplan_activity_client.create_facility_activity(
                                                    request_info=request_info,
                                                    fieldPlan=fieldplan,
                                                    roleToIds=role_to_ids,
                                                    facility_id=facility_id
                                                )
                                                logger.info(f"Facility activity created successfully for facility {facility_id}")
                                                logger.debug(f"Facility activity response: {facility_activity_resp}")
                                            except Exception as activity_exc:
                                                logger.error(f"Error creating facility activity for {facility_id}: {activity_exc}", exc_info=True)
                            else:
                                for row_idx, _ in chunk:
                                    df.at[row_idx, 'Field Plan Linking Status'] = f"Failed: {fieldplan_resp.status_code} {fieldplan_resp.text}"
                        except Exception as bulk_exc:
                            for row_idx, _ in chunk:
                                df.at[row_idx, 'Field Plan Linking Status'] = f"Exception: {str(bulk_exc)}"

                if pending_assessment_handoffs and assessment_client:
                    for row_idx, bulk_entry, plan_facility_id in pending_assessment_handoffs:
                        facility_id = bulk_entry["facilityId"]
                        try:
                            create_resp = fieldplan_client.create_fieldPlan_facility(
                                request_info=request_info,
                                fieldPlan_id=fieldplan_id,
                                facility_id=facility_id,
                                source_plan_facility_id=plan_facility_id,
                                additional_fields=bulk_entry.get("additionalFields"),
                            )
                            field_plan_facility = (
                                create_resp.get("FieldPlanFacility")
                                or create_resp.get("fieldPlanFacility")
                                or {}
                            )
                            field_plan_facility_id = field_plan_facility.get("id")
                            if not field_plan_facility_id:
                                search_resp = fieldplan_client.search_fieldplan_facility(
                                    request_info, fieldplan_id
                                )
                                for pf in search_resp.get("FieldPlanFacilities", []):
                                    if pf.get("facilityId") == facility_id:
                                        field_plan_facility_id = pf.get("id")
                                        break
                            if not field_plan_facility_id:
                                raise ValueError("FieldPlanFacility id not returned after create")

                            assessment_client.apply_facility_handoff(
                                request_info=request_info,
                                plan_facility_id=plan_facility_id,
                                installation_field_plan_id=fieldplan_id,
                                field_plan_facility_id=field_plan_facility_id,
                            )
                            df.at[row_idx, 'Field Plan Linking Status'] = "Linked + Assessment Handoff"
                            fieldplan_linked_facility_ids.add(facility_id)

                            if fieldplan_data:
                                fieldplan = fieldplan_data[0]
                                if fieldplan.get("status") == 'SCHEDULED':
                                    try:
                                        fieldplan_activity_client.create_facility_activity(
                                            request_info=request_info,
                                            fieldPlan=fieldplan,
                                            roleToIds=role_to_ids,
                                            facility_id=facility_id,
                                        )
                                    except Exception as activity_exc:
                                        logger.error(
                                            f"Error creating facility activity for {facility_id}: {activity_exc}",
                                            exc_info=True,
                                        )
                        except Exception as handoff_exc:
                            logger.error(
                                f"Assessment handoff failed for planFacilityId={plan_facility_id}: {handoff_exc}",
                                exc_info=True,
                            )
                            df.at[row_idx, 'Field Plan Linking Status'] = f"Handoff failed: {handoff_exc}"

                if pending_bulk_fieldplan_updates:
                    chunk_size = BULK_INGEST_CHUNK_SIZE
                    for i in range(0, len(pending_bulk_fieldplan_updates), chunk_size):
                        chunk = pending_bulk_fieldplan_updates[i:i + chunk_size]
                        updates_chunk = [entry for _, entry in chunk]
                        try:
                            update_resp = fieldplan_client.update_fieldPlan_facility_bulk(
                                request_info=request_info,
                                updates=updates_chunk,
                            )

                            if update_resp.status_code in (200, 201, 202):
                                for row_idx, _ in chunk:
                                    df.at[row_idx, 'Field Plan Linking Status'] = "Updated"
                            else:
                                for row_idx, _ in chunk:
                                    df.at[row_idx, 'Field Plan Linking Status'] = f"Failed: {update_resp.status_code} {update_resp.text}"
                        except Exception as bulk_exc:
                            for row_idx, _ in chunk:
                                df.at[row_idx, 'Field Plan Linking Status'] = f"Exception: {str(bulk_exc)}"

            except Exception as e:
                logger.error(f"Error fetching fieldplan facilities: {e}")
                # Continue without fieldplan facility data if there's an error

        # ---------- write results back into workbook preserving formatting ----------
        # Ensure headers exist in sheet (without wiping template)
        header_values = [cell.value for cell in ws[1]]

        for col_name in ["Field Plan Linking Status"]:
            if col_name not in header_values:
                cell = ws.cell(row=1, column=len(header_values) + 1, value=col_name)
                cell.font = Font(bold=True)
                header_values.append(col_name)
                if col_name not in df.columns:
                    df[col_name] = ""  # ensure column exists in dataframe

        # Delete data rows only (preserve header row and template formatting)
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)

        # Write data rows back (without header row)
        export_df = prepare_dataframe_for_excel_export(df)
        for r_idx, row in enumerate(dataframe_to_rows(export_df, index=False, header=False), start=2):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        wb.save(output_file_path)

        autofit_columns(output_file_path, facility_sheet_name , auto_fit=True)

        background_tasks.add_task(cleanup_temp_file, output_file_path)

        return FileResponse(
            path=output_file_path,
            filename=output_filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        logger.error(f"Error finalizing facility data: {e}")
        raise HTTPException(
            status_code=500,
            detail=f"Failed to finalize facility data: {str(e)}"
        )
    finally:
        if input_temp_file and os.path.exists(input_temp_file.name):
            os.unlink(input_temp_file.name)


@router.post('/amcConfigurationValidateData',
             summary='Validate AMC configuration Excel file before processing',
             response_description='Returns validation report Excel with PASSED/FAILED rows')
async def validate_amc_configurations_excel_sheet(
        background_tasks: BackgroundTasks,
        amc_file: UploadFile = File(..., description="Excel file containing AMC configuration data"),
        amc_sheet_name: str = Form(default="amc-configurations", description="Name of the sheet containing AMC data"),
        request_info: str = Form(default="")
):
    input_temp_file = None
    output_temp_file = None
    request_info_obj = request_info_from_json(request_info)

    try:
        input_temp_file, _ = await _save_upload_to_temp_file(amc_file, suffix=".xlsx")

        # Prepare output file
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"amc_configuration_validation_results_{timestamp}.xlsx"
        output_temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        output_temp_file.close()
        output_file_path = output_temp_file.name

        # Read Excel file
        wb = load_workbook(input_temp_file.name)
        if amc_sheet_name not in wb.sheetnames:
            raise HTTPException(status_code=400, detail=f"Sheet '{amc_sheet_name}' not found")

        df = pd.read_excel(input_temp_file.name, sheet_name=amc_sheet_name)
        df.columns = [str(c).strip() for c in df.columns]

        # Status/error must be object dtype (Excel may load as float); avoids FutureWarning on df.loc writes.
        for _col in ("status", "error"):
            if _col not in df.columns:
                df[_col] = ""
            else:
                df[_col] = df[_col].map(lambda x: "" if pd.isna(x) else str(x))
            df[_col] = df[_col].astype("object")

        required_columns = [
            "Facility Id",
            "Health Facility Name",
            # "Vendor",
            "AMC-Frequency",
            "AMC-Duration"
        ]

        # Column names
        # vendor_col = "Vendor" if required_columns else "vendor"
        frequency_col = "AMC-Frequency" if required_columns else "amc-frequency"
        duration_col = "AMC-Duration" if required_columns else "amc-duration"

        # Validate each row - only check vendor, AMC frequency, and AMC duration
        error_count = 0
        for index, row in df.iterrows():
            validation_errors = []

            try:
                # Validate vendor, AMC frequency, and AMC duration
                # vendor_name = str(row.get(vendor_col, "")).strip() if not pd.isna(row.get(vendor_col)) else ""
                amc_frequency = str(row.get(frequency_col, "")).strip() if not pd.isna(row.get(frequency_col)) else ""
                amc_duration = str(row.get(duration_col, "")).strip() if not pd.isna(row.get(duration_col)) else ""

                # A row left entirely blank means the user does not want an AMC on that facility - and
                # is how an existing configuration gets removed at ingest time. Only a half-filled row
                # is an error, because it is always a mistake rather than an opt-out.
                if amc_frequency and not amc_duration:
                    validation_errors.append(
                        "Please ensure AMC duration is selected when an AMC frequency is set."
                    )
                elif amc_duration and not amc_frequency:
                    validation_errors.append(
                        "Please ensure AMC frequency is selected when an AMC duration is set."
                    )

                # Set status and error
                if validation_errors:
                    df.at[index, 'status'] = 'FAILED'
                    df.at[index, 'error'] = "; ".join(validation_errors)
                    error_count += 1
                else:
                    df.at[index, 'status'] = 'PASSED'
                    df.at[index, 'error'] = ''

            except Exception as e:
                df.at[index, 'status'] = 'FAILED'
                df.at[index, 'error'] = f'Unexpected error: {str(e)}'
                error_count += 1
                logger.error(f"Error validating row {index}: {e}")

        # Write results to Excel
        ws = wb[amc_sheet_name]
        header_values = [cell.value for cell in ws[1]]

        # Add status/error columns if missing
        for col_name in ["status", "error"]:
            if col_name not in header_values:
                new_col_idx = len(header_values) + 1
                cell = ws.cell(row=1, column=new_col_idx, value=col_name)
                cell.font = Font(bold=True)
                header_values.append(col_name)

                # lock header cell
                cell.protection = Protection(locked=True)

                # lock all data cells in this new column
                for r_idx in range(2, ws.max_row + 1):
                    ws.cell(row=r_idx, column=new_col_idx).protection = Protection(locked=True)

        grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        # Write data rows back (without header row)
        export_df = prepare_dataframe_for_excel_export(df)
        for r_idx, row in enumerate(dataframe_to_rows(export_df, index=False, header=False), start=2):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)

                # force lock for status/error columns
                if ws.cell(1, c_idx).value in ["status", "error"]:
                    cell.protection = Protection(locked=True)
                    cell.fill = grey_fill

        # Ensure sheet protection is ON
        ws.protection.sheet = True
        ws.protection.enable()

        # Save to output file
        wb.save(output_file_path)

        autofit_columns(output_file_path, amc_sheet_name, auto_fit=True)

        background_tasks.add_task(cleanup_temp_file, output_file_path)
        background_tasks.add_task(cleanup_temp_file, input_temp_file.name)

        response = FileResponse(
            path=output_file_path,
            filename=output_filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response.headers["X-Error-Count"] = str(error_count)

        return response

    except HTTPException:
        if input_temp_file and os.path.exists(input_temp_file.name):
            os.unlink(input_temp_file.name)
        if output_temp_file and os.path.exists(output_temp_file.name):
            os.unlink(output_temp_file.name)
        raise
    except Exception as e:
        logger.error(f"Unhandled error in validate_amc_configurations_excel_sheet: {e}")
        if input_temp_file and os.path.exists(input_temp_file.name):
            os.unlink(input_temp_file.name)
        if output_temp_file and os.path.exists(output_temp_file.name):
            os.unlink(output_temp_file.name)
        raise HTTPException(status_code=500, detail=f"An unexpected error occurred: {str(e)}")


def _extract_boundary_codes(geography_details: dict, key: str) -> Set[str]:
    """
    Boundary codes under geography_details[key] ("districts" or "blocks").

    The platform emits geographyDetails in two shapes - a list of plain code strings, and a list of
    {"code": ..., "name": ...} objects - so both are accepted here rather than picking one and
    silently returning nothing for the other.
    """
    codes: Set[str] = set()
    for entry in geography_details.get(key) or []:
        if isinstance(entry, dict):
            code = entry.get("code")
        else:
            code = entry
        if code:
            codes.add(str(code).strip())
    codes.discard("")
    return codes


def _build_amc_geography_scope(geography_details: dict):
    """
    Predicate telling whether a boundary code falls inside the districts/blocks the user selected,
    plus whether a scope was actually supplied.

    Boundary codes are hierarchical ("India_Karnataka_Mysuru_HD_Kote"), so a district is matched by
    prefix. When no district or block was selected the scope is unrestricted - a bulk upload must not
    silently drop every row because the caller omitted the geography filters. Callers must check the
    second return value before doing anything destructive: an unrestricted scope covers the whole
    project, which is exactly when deleting "everything not in the file" would be catastrophic.
    """
    block_codes = _extract_boundary_codes(geography_details, "blocks")
    district_codes = _extract_boundary_codes(geography_details, "districts")

    if not block_codes and not district_codes:
        logger.warning("geography_details carries no districts or blocks - AMC rows will not be filtered by boundary")
        return (lambda boundary_code: True), False

    def is_in_scope(boundary_code: Optional[str]) -> bool:
        code = str(boundary_code or "").strip()
        if not code:
            return False
        if code in block_codes:
            return True
        return any(code == district or code.startswith(f"{district}_") for district in district_codes)

    return is_in_scope, True


def _add_months(start: datetime, months: int) -> datetime:
    """
    Calendar-month arithmetic, matching AmcConfigurationServiceUtil.generateAmcVisits on the Java side.
    A 30-days-per-month approximation drifts by over a month on a 60-month contract, which would put
    the configuration end date before the last visit the scheduler generates.
    """
    month_index = start.month - 1 + months
    year = start.year + month_index // 12
    month = month_index % 12 + 1
    day = min(start.day, [31, 29 if year % 4 == 0 and (year % 100 != 0 or year % 400 == 0) else 28,
                          31, 30, 31, 30, 31, 31, 30, 31, 30, 31][month - 1])
    return start.replace(year=year, month=month, day=day)


def _normalize_amc_assignments(assignments) -> List[dict]:
    """
    Rebuilds the nested auditDetails the AMC model expects on each assignment.

    The search query aggregates assignments with jsonb_build_object and emits the audit columns flat
    ("createdBy", "lastModifiedTime", ...), while AmcConfigurationAssignment declares an auditDetails
    object. Because the service disables FAIL_ON_UNKNOWN_PROPERTIES, echoing the flat form back on an
    update is accepted but silently drops the values, and the assignment upsert then writes NULL into
    last_modified_by/last_modified_time.
    """
    normalized: List[dict] = []
    for assignment in assignments or []:
        if not isinstance(assignment, dict):
            continue
        assignment = dict(assignment)
        audit_details = assignment.get("auditDetails")
        if not audit_details:
            audit_details = {
                key: assignment.get(key)
                for key in ("createdBy", "createdTime", "lastModifiedBy", "lastModifiedTime")
                if assignment.get(key) is not None
            }
        if audit_details:
            assignment["auditDetails"] = audit_details
        normalized.append(assignment)
    return normalized


def build_amc_configuration_update_payload(
        existing_config: dict,
        duration_months: int,
        frequency_months: int,
) -> dict:
    """
    Full update payload for an existing AMC configuration, carrying only a new duration/frequency.

    The AMC service replays the entire create validation on update, so a partial payload is rejected:
    everything it validated on create has to be sent back. Two fields are easy to get wrong -
      * status: the persister writes it straight from the payload and nothing restores it, so omitting
        it would null the column out;
      * geographyDetails: its state is read-only, so the stored value is reused rather than the one
        from the upload form.
    Search-only enrichments (vendor, facility, project, assetsAmc, totalVisits, completedVisits) are
    deliberately left out. configurationEndDate is recomputed server-side from durationMonths.
    """
    return {
        "id": existing_config.get("id"),
        "tenantId": existing_config.get("tenantId"),
        "vendorId": existing_config.get("vendorId"),
        "facilityId": existing_config.get("facilityId"),
        "projectId": existing_config.get("projectId"),
        "durationMonths": duration_months,
        "visitFrequencyMonths": frequency_months,
        "status": existing_config.get("status") or "ACTIVE",
        "configurationStartDate": existing_config.get("configurationStartDate"),
        "configurationEndDate": existing_config.get("configurationEndDate"),
        "assetTypes": existing_config.get("assetTypes"),
        "assignments": _normalize_amc_assignments(existing_config.get("assignments")),
        "geographyDetails": existing_config.get("geographyDetails"),
        "additionalDetails": existing_config.get("additionalDetails"),
        "auditDetails": existing_config.get("auditDetails"),
    }


def get_vendor_id_for_amc_field_staff(user_info_data: List[dict]) -> str:
    # Vendor id (or name fallback) for the vendor that has a user with role AMC_FIELD_STAFF.
    role_code = "AMC_FIELD_STAFF"
    candidates: Set[str] = set()

    for vendor_mapping in user_info_data:
        vendor_id = (vendor_mapping.get("vendorId") or "").strip()
        vendor_name = (vendor_mapping.get("vendor") or "").strip()
        if not vendor_id:
            if not vendor_name:
                continue
            vendor_id = vendor_name

        users = vendor_mapping.get("users", [])
        if not users and "userId" in vendor_mapping:
            users = [{"userId": vendor_mapping.get("userId"), "userName": vendor_mapping.get("userName")}]
        if not isinstance(users, list):
            raise HTTPException(status_code=400, detail=f"users must be a list for vendorId: {vendor_id}")

        for user in users:
            for role in user.get("roles") or []:
                if role.get("code") == role_code:
                    candidates.add(vendor_id)
                    break

    if not candidates:
        raise HTTPException(
            status_code=400,
            detail="No vendor with a user having the AMC_FIELD_STAFF role was found in user_info_list",
        )
    if len(candidates) > 1:
        raise HTTPException(
            status_code=400,
            detail="Multiple vendors have the AMC_FIELD_STAFF role; user_info_list must identify a single vendor",
        )
    return next(iter(candidates))


@router.post('/amcConfigurationBulkIngest',
             summary='Bulk ingest AMC configuration template data',
             response_description="Returns processed Excel file with AMC configuration creation results")
async def bulk_ingest_amc_configurations(
        background_tasks: BackgroundTasks,
        amc_file: UploadFile = File(..., description="Excel file containing AMC configuration data"),
        amc_sheet_name: str = Form(default="amc-configurations", description="Name of the sheet containing AMC data"),
        project_id: str = Form(..., description="Project ID"),
        user_info_list: str = Form(..., description="JSON array of user info objects with vendor mapping"),
        geography_details: str = Form(..., description="JSON object with state, districts and blocks for the AMC configurations being created"),
        request_info: str = Form(default="")
):
    input_temp_file = None
    output_temp_file = None
    request_info_obj = request_info_from_json(request_info)

    # Get tenant ID from request info or use default
    tenant_id = request_info_obj.user_info.tenant_id if request_info_obj.user_info and request_info_obj.user_info.tenant_id else "in"

    try:
        # Parse user info list
        try:
            user_info_data = json.loads(user_info_list)
            if not isinstance(user_info_data, list):
                raise HTTPException(status_code=400, detail="user_info_list must be a JSON array")
        except json.JSONDecodeError as e:
            raise HTTPException(status_code=400, detail=f"Invalid JSON in user_info_list: {str(e)}")

        # Parse geography details (state/districts/blocks) - same scope applied to every AMC configuration in this batch
        try:
            geography_details_data = json.loads(geography_details)
            if not isinstance(geography_details_data, dict):
                raise HTTPException(status_code=400, detail="geography_details must be a JSON object")
        except json.JSONDecodeError as e:
            raise HTTPException(status_code=400, detail=f"Invalid JSON in geography_details: {str(e)}")

        amc_vendor_id = get_vendor_id_for_amc_field_staff(user_info_data)

        assignment_users = []
        vendor_mappings = []  # One entry per payload item with valid users
        for vendor_mapping in user_info_data:
            # Primary key: vendorId (UUID)
            vendor_id = vendor_mapping.get("vendorId", "").strip()
            # Secondary: vendor name (for backward compatibility and Excel lookup)
            vendor_name = vendor_mapping.get("vendor", "").strip()

            if not vendor_id:
                # Fallback: if no vendorId, use vendor name as key
                if not vendor_name:
                    logger.warning(f"Vendor mapping missing both vendorId and vendor name: {vendor_mapping}")
                    continue
                vendor_id = vendor_name  # Use name as fallback key

            # Support both old format (single user) and new format (list of users)
            users = vendor_mapping.get("users", [])
            if not users:
                # Backward compatibility: if "users" not found, check for single user fields
                if "userId" in vendor_mapping:
                    users = [{"userId": vendor_mapping.get("userId"), "userName": vendor_mapping.get("userName")}]
                else:
                    logger.warning(f"No users found for vendorId: {vendor_id}")
                    continue

            # Validate users list
            if not isinstance(users, list):
                raise HTTPException(status_code=400, detail=f"users must be a list for vendorId: {vendor_id}")

            # Process users
            for user in users:
                # Extract user ID - prefer 'id' (from full user object), then 'userId', then 'uuid'
                user_id = user.get("uuid") or user.get("userId") or user.get("id")

                if not user_id:
                    logger.warning(f"User object missing ID field: {user}")
                    continue

                # Extract user name - prefer 'name', then 'userName'
                user_name = user.get("name") or user.get("userName", "")

                # Extract tenant ID from user object if available
                user_tenant_id = user.get("tenantId")

                assignment_users.append({
                    "id": str(user_id),  # Convert to string for consistency
                    "userId": str(user_id),  # Keep for backward compatibility
                    "userName": user_name,
                    "name": user_name,
                    "tenantId": user_tenant_id,
                    # role/pocNumber live on the user object; fall back to the vendor_mapping level
                    # for backward compatibility with older payload shapes.
                    "role": user.get("role") or vendor_mapping.get("role"),
                    "pocNumber": user.get("pocNumber") or vendor_mapping.get("pocNumber"),
                    "fullUser": user  # Store full user object for reference
                })

            if not assignment_users:
                logger.warning(f"No valid users found for vendorId: {vendor_id}")
                continue

            vendor_mappings.append({
                "vendorId": vendor_id,
                "vendorName": vendor_name,
                "users": assignment_users
            })

        # Save uploaded file
        input_temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        content = await amc_file.read()
        input_temp_file.write(content)
        input_temp_file.close()

        # Prepare output file
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"amc_configuration_ingestion_results_{timestamp}.xlsx"
        output_temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        output_temp_file.close()
        output_file_path = output_temp_file.name

        # Read Excel file
        wb = load_workbook(input_temp_file.name)
        if amc_sheet_name not in wb.sheetnames:
            raise HTTPException(status_code=400, detail=f"Sheet '{amc_sheet_name}' not found")

        df = pd.read_excel(input_temp_file.name, sheet_name=amc_sheet_name)
        df.columns = [str(c).strip() for c in df.columns]

        for _col in ("status", "error"):
            if _col not in df.columns:
                df[_col] = ""
            else:
                df[_col] = df[_col].map(lambda x: "" if pd.isna(x) else str(x))
            df[_col] = df[_col].astype("object")

        required_columns = ["Facility Id", "Health Facility Name", "Vendor", "AMC-Frequency", "AMC-Duration"]

        # Initialize clients
        facility_client = FacilityServiceClient(facility_service_url) if facility_service_url else None
        amc_client = AMCSchedulerServiceClient(amc_scheduler_service_url) if amc_scheduler_service_url else None

        if not amc_client:
            raise HTTPException(status_code=500, detail="AMC Scheduler Service is not configured")

        if not facility_client:
            raise HTTPException(status_code=500, detail="Facility Service is not configured")

        facility_ids_from_file = []
        for _, row in df.iterrows():
            if pd.isna(row.get("Facility Id")) and pd.isna(row.get("Health Facility Name")):
                continue
            facility_id = str(row.get("Facility Id", "")).strip()
            if facility_id:
                facility_ids_from_file.append(facility_id)

        facility_batch_size = int(os.getenv("AMC_INGEST_FACILITY_ID_BATCH_SIZE", "500"))

        def load_facilities_by_id(ids: List[str]) -> Dict[str, dict]:
            """
            Facility rows keyed by facility_id, in batches. Uses the with-boundary variant because
            boundary codes drive the district/block scope check below.
            """
            loaded: Dict[str, dict] = {}
            unique_ids = list(dict.fromkeys(i for i in ids if i))
            for batch_start in range(0, len(unique_ids), facility_batch_size):
                batch_ids = unique_ids[batch_start:batch_start + facility_batch_size]
                bulk_facility_result = facility_client.bulk_search_facility_with_boundary(
                    request_info=request_info_obj,
                    tenant_ids=["in"],
                    facility_ids=batch_ids,
                    limit=max(len(batch_ids), 50),
                    send_non_paginated_response=True,
                )
                for facility in (bulk_facility_result.get("facilities", []) or []):
                    f_id = facility.get("facility_id")
                    if f_id:
                        loaded[f_id] = facility
            return loaded

        facility_map = {}
        if facility_ids_from_file:
            try:
                facility_map = load_facilities_by_id(facility_ids_from_file)
            except Exception as e:
                logger.error(f"Error bulk searching facilities for AMC ingest: {e}", exc_info=True)
                raise HTTPException(status_code=502, detail=f"Facility lookup failed: {str(e)}")

        # District/block scope selected in the UI. Rows outside it are never created, and existing
        # configurations inside it that disappeared from the file are deleted (see reconciliation below).
        is_boundary_in_scope, has_geography_scope = _build_amc_geography_scope(geography_details_data)

        # One search for the whole project instead of one per row. Configurations of any status are
        # loaded, not just ACTIVE ones: ux_amc_configuration_unique_installation
        # (tenant_id, facility_id, project_id, vendor_id) means a second create for a facility that
        # already has an EXPIRED or CANCELLED configuration would be rejected by the database - and
        # since persistence is asynchronous, that rejection would never surface to the user. So an
        # existing configuration is always updated, never duplicated.
        existing_config_by_facility: Dict[str, dict] = {}
        try:
            for config in amc_client.search_all_amc_configurations(
                    request_info_obj, project_id=project_id, tenant_id=tenant_id):
                config_facility_id = config.get("facilityId")
                if not config_facility_id:
                    continue
                current = existing_config_by_facility.get(config_facility_id)
                # An ACTIVE configuration always wins over a closed one for the same facility.
                if current is None or (
                        str(current.get("status") or "").strip().upper() != "ACTIVE"
                        and str(config.get("status") or "").strip().upper() == "ACTIVE"):
                    existing_config_by_facility[config_facility_id] = config
            logger.info(
                f"Loaded {len(existing_config_by_facility)} existing AMC configuration(s) for project {project_id}"
            )
        except Exception as e:
            logger.error(f"Error loading existing AMC configurations for project {project_id}: {e}", exc_info=True)
            raise HTTPException(status_code=502, detail=f"AMC configuration lookup failed: {str(e)}")

        configs_to_create = []
        row_indexes_for_configs = []
        configs_to_update = []
        row_indexes_for_updates = []
        # Facilities the user kept in the file with a usable AMC selection. Anything already configured
        # inside the selected districts/blocks but missing from this set loses its configuration.
        facility_ids_kept_in_file: Set[str] = set()

        for index, row in df.iterrows():
            try:
                # Skip empty rows
                if pd.isna(row.get("Facility Id")) and pd.isna(row.get("Health Facility Name")):
                    df.at[index, 'status'] = 'skipped'
                    df.at[index, 'error'] = 'Empty row'
                    continue

                # Get facility by Facility ID
                facility_id = str(row.get("Facility Id", "")).strip()
                if not facility_id:
                    df.at[index, 'status'] = 'failed'
                    df.at[index, 'error'] = 'Facility Id is required'
                    continue

                if facility_id not in facility_map:
                    df.at[index, 'status'] = 'failed'
                    df.at[index, 'error'] = f'Facility not found for Facility Id: {facility_id}'
                    continue

                # The template writes a locked BoundaryCode column; fall back to the facility record
                # when the column is missing (older template) or blank.
                facility_record = facility_map.get(facility_id) or {}
                boundary_code = str(row.get("BoundaryCode", "") or "").strip()
                if not boundary_code:
                    boundary_code = str(
                        facility_record.get("boundary_code") or facility_record.get("boundaryCode") or ""
                    ).strip()

                if not is_boundary_in_scope(boundary_code):
                    df.at[index, 'status'] = 'skipped'
                    df.at[index, 'error'] = (
                        f'Facility is outside the selected districts/blocks (boundary code: {boundary_code or "unknown"})'
                    )
                    continue

                # Get AMC frequency and duration (already validated, just convert to months)
                frequency_col = "AMC-Frequency" if "AMC-Frequency" in df.columns else "amc-frequency"
                duration_col = "AMC-Duration" if "AMC-Duration" in df.columns else "amc-duration"
                amc_frequency = "" if pd.isna(row.get(frequency_col)) else str(row.get(frequency_col, "")).strip()
                amc_duration = "" if pd.isna(row.get(duration_col)) else str(row.get(duration_col, "")).strip()

                # A row left blank means the user does not want an AMC on this facility. That is a
                # deliberate choice, not an error - and it is what drives the deletion pass below.
                if not amc_frequency and not amc_duration:
                    df.at[index, 'status'] = 'skipped'
                    df.at[index, 'error'] = 'No AMC frequency/duration selected'
                    continue

                # Convert frequency to months (format already validated in validation endpoint)
                if amc_frequency == "Every 6 Months":
                    frequency_months = 6
                elif amc_frequency == "Every 1 Year":
                    frequency_months = 12
                else:
                    df.at[index, 'status'] = 'failed'
                    df.at[index, 'error'] = f'Unexpected AMC frequency value: {amc_frequency}'
                    continue

                # Convert duration to months (format already validated in validation endpoint)
                if amc_duration == "1 Year":
                    duration_months = 12
                elif amc_duration == "3 Years":
                    duration_months = 36
                elif amc_duration == "5 Years":
                    duration_months = 60
                else:
                    df.at[index, 'status'] = 'failed'
                    df.at[index, 'error'] = f'Unexpected AMC duration value: {amc_duration}'
                    continue

                facility_ids_kept_in_file.add(facility_id)

                # An existing configuration is updated in place rather than duplicated - the database
                # allows only one configuration per (tenant, facility, project, vendor).
                existing_config = existing_config_by_facility.get(facility_id)
                if existing_config:
                    if (existing_config.get("durationMonths") == duration_months
                            and existing_config.get("visitFrequencyMonths") == frequency_months
                            and str(existing_config.get("status") or "").strip().upper() == "ACTIVE"):
                        df.at[index, 'status'] = 'skipped'
                        df.at[index, 'error'] = 'No change'
                        continue

                    configs_to_update.append(build_amc_configuration_update_payload(
                        existing_config, duration_months, frequency_months
                    ))
                    row_indexes_for_updates.append(index)
                    continue

                # Create assignments array from vendor users
                assignments = []
                for user in assignment_users:
                    # Use user's id (from full user object) or userId (backward compatibility)
                    assigned_user_id = user.get("id") or user.get("userId")
                    # Prefer user's tenantId if available, otherwise use default tenant_id
                    assignment_tenant_id = user.get("tenantId") or tenant_id

                    assignment = {
                        "assignedUser": str(assigned_user_id),
                        "tenantId": assignment_tenant_id,
                        "role": user.get("role"),
                        "additionalDetails": None,
                        "pocNumber": user.get("pocNumber"),
                    }
                    assignments.append(assignment)

                # Convert asset types to API format (objects with code and name)
                asset_types_formatted = []
                asset_type_names = {
                    "INVERTER": "Inverter",
                    "PANEL": "Panel",
                    "BATTERY": "Battery"
                }
                for asset_type in DEFAULT_AMC_ASSET_TYPES:
                    asset_types_formatted.append({
                        "code": asset_type,
                        "name": asset_type_names.get(asset_type, asset_type.title())
                    })

                # Calculate configuration dates (start date = now, end date = start + duration).
                # Calendar months, not 30-day months: the AMC scheduler generates visits with calendar
                # arithmetic, so an approximation would place the last visit past the contract end.
                now = datetime.now()
                configuration_start_date = int(now.timestamp() * 1000)  # Convert to milliseconds
                configuration_end_date = int(_add_months(now, duration_months).timestamp() * 1000)

                configs_to_create.append({
                    "tenantId": tenant_id,
                    "vendorId": amc_vendor_id,
                    "facilityId": facility_id,
                    "projectId": project_id,
                    "durationMonths": duration_months,
                    "visitFrequencyMonths": frequency_months,
                    "status": "ACTIVE",
                    "configurationStartDate": configuration_start_date,
                    "configurationEndDate": configuration_end_date,
                    "assetTypes": asset_types_formatted,
                    "assignments": assignments,
                    # Same state/districts/blocks scope for every AMC configuration in this batch, same as field plan
                    "geographyDetails": geography_details_data
                })
                row_indexes_for_configs.append(index)
            except Exception as e:
                df.at[index, 'status'] = 'failed'
                df.at[index, 'error'] = f'Unexpected error: {str(e)}'
                logger.error(f"Error processing row {index}: {e}")

        chunk_size = AMC_CONFIGURATION_BULK_CHUNK_SIZE

        def _process_amc_chunk(
            action: str,
            success_status: str,
            chunk_cfgs: List[dict],
            chunk_row_indexes: List,
            http_session: requests.Session,
        ) -> None:
            # A chunk is all-or-nothing: the AMC service answers per request, not per configuration,
            # so a single rejection marks every row of the chunk. Creates and updates are therefore
            # dispatched separately, to keep one failing side from burying the other.
            send = (amc_client.create_amc_configurations_bulk if action == "create"
                    else amc_client.update_amc_configurations_bulk)
            try:
                send(request_info_obj, chunk_cfgs, session=http_session)
                df.loc[chunk_row_indexes, "status"] = success_status
                df.loc[chunk_row_indexes, "error"] = ""
            except Exception as exc:
                logger.error(
                    "Bulk AMC %s failed for %s rows: %s",
                    action,
                    len(chunk_cfgs),
                    exc,
                    exc_info=True,
                )
                err = str(exc)
                df.loc[chunk_row_indexes, "status"] = "failed"
                df.loc[chunk_row_indexes, "error"] = err

        # Facilities that already hold a configuration inside the selected districts/blocks but are no
        # longer requested in the file: the upload is the source of truth for its own scope, so those
        # configurations are deactivated (soft delete - the AMC service clears isActive, no row is
        # removed and the visit history is preserved). Facilities outside the selected
        # districts/blocks are never touched - another upload owns them.
        configs_to_delete = []
        candidate_facility_ids = [
            facility_id for facility_id, config in existing_config_by_facility.items()
            if str(config.get("status") or "").strip().upper() == "ACTIVE"
            and facility_id not in facility_ids_kept_in_file
        ] if has_geography_scope else []
        if not has_geography_scope and existing_config_by_facility:
            logger.warning(
                "Skipping AMC deletion pass for project %s: geography_details defines no districts or "
                "blocks, so the upload's scope cannot be bounded and every configuration would qualify",
                project_id,
            )
        if candidate_facility_ids:
            try:
                # Facilities absent from the file were never looked up, so resolve their boundary codes.
                missing_ids = [fid for fid in candidate_facility_ids if fid not in facility_map]
                if missing_ids:
                    facility_map.update(load_facilities_by_id(missing_ids))
            except Exception as e:
                logger.error(f"Error resolving boundaries of AMC deletion candidates: {e}", exc_info=True)
                raise HTTPException(status_code=502, detail=f"Facility lookup failed: {str(e)}")

            for facility_id in candidate_facility_ids:
                facility_record = facility_map.get(facility_id) or {}
                boundary_code = facility_record.get("boundary_code") or facility_record.get("boundaryCode")
                if not is_boundary_in_scope(boundary_code):
                    continue
                config = existing_config_by_facility[facility_id]
                configs_to_delete.append({
                    "id": config.get("id"),
                    "tenantId": config.get("tenantId") or tenant_id,
                })

        deleted_count = 0
        with requests.Session() as http_session:
            for start in range(0, len(configs_to_create), chunk_size):
                _process_amc_chunk(
                    "create",
                    "success",
                    configs_to_create[start:start + chunk_size],
                    row_indexes_for_configs[start:start + chunk_size],
                    http_session,
                )

            for start in range(0, len(configs_to_update), chunk_size):
                _process_amc_chunk(
                    "update",
                    "updated",
                    configs_to_update[start:start + chunk_size],
                    row_indexes_for_updates[start:start + chunk_size],
                    http_session,
                )

            for start in range(0, len(configs_to_delete), chunk_size):
                chunk = configs_to_delete[start:start + chunk_size]
                try:
                    amc_client.delete_amc_configurations(request_info_obj, chunk, session=http_session)
                    deleted_count += len(chunk)
                except Exception as exc:
                    logger.error(
                        "Bulk AMC delete failed for %s configuration(s): %s", len(chunk), exc, exc_info=True
                    )

        logger.info(
            f"AMC bulk ingest for project {project_id}: {len(configs_to_create)} created, "
            f"{len(configs_to_update)} updated, {deleted_count}/{len(configs_to_delete)} deleted"
        )

        with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name=amc_sheet_name)

        autofit_columns(output_file_path, amc_sheet_name, auto_fit=True)

        background_tasks.add_task(cleanup_temp_file, output_file_path)
        background_tasks.add_task(cleanup_temp_file, input_temp_file.name)

        response = FileResponse(
            path=output_file_path,
            filename=output_filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        # Deletions do not show up as rows in the returned file, so the UI needs them out of band.
        response.headers["X-Deleted-Count"] = str(deleted_count)
        return response

    except HTTPException:
        if input_temp_file and os.path.exists(input_temp_file.name):
            os.unlink(input_temp_file.name)
        if output_temp_file and os.path.exists(output_temp_file.name):
            os.unlink(output_temp_file.name)
        raise
    except Exception as e:
        logger.error(f"Unhandled error in bulk_ingest_amc_configurations: {e}")
        if input_temp_file and os.path.exists(input_temp_file.name):
            os.unlink(input_temp_file.name)
        if output_temp_file and os.path.exists(output_temp_file.name):
            os.unlink(output_temp_file.name)
        raise HTTPException(status_code=500, detail=f"An unexpected error occurred: {str(e)}")


def _read_assessment_include_sheet(file_path: str, sheet_name: str = "FacilityMapping") -> pd.DataFrame:
    wb = load_workbook(file_path, read_only=True)
    if sheet_name not in wb.sheetnames:
        for candidate in ("FacilityMapping", "AssessmentInclude"):
            if candidate in wb.sheetnames:
                sheet_name = candidate
                break
        else:
            sheet_name = wb.sheetnames[0]
    df = pd.read_excel(file_path, sheet_name=sheet_name)
    df.columns = [str(c).strip() for c in df.columns]
    return df.loc[:, ~df.columns.str.startswith("Unnamed")]


def _find_assessment_include_col(df: pd.DataFrame) -> Optional[str]:
    for col in df.columns:
        col_lower = str(col).lower()
        if "include in assessment plan" in col_lower:
            return col
    return None


def _find_assessment_facility_id_col(df: pd.DataFrame) -> Optional[str]:
    for col in df.columns:
        if "facility id" in str(col).lower():
            return col
    return None


def _row_value(row, col: Optional[str]) -> str:
    if not col:
        return ""
    val = row.get(col)
    if pd.isna(val):
        return ""
    return str(val).strip()


def _normalize_facility_id_from_excel(val) -> str:
    if pd.isna(val) or val is None:
        return ""
    if isinstance(val, bool):
        return str(val).strip()
    if isinstance(val, int):
        return str(val)
    if isinstance(val, float):
        if math.isfinite(val) and val.is_integer():
            return str(int(val))
        return str(val).strip()
    text = str(val).strip()
    if not text or text.lower() == "nan":
        return ""
    if text.endswith(".0") and text[:-2].replace("-", "", 1).isdigit():
        return text[:-2]
    return text


def _build_assessment_include_facility_payload(row, df: pd.DataFrame, facility_id_col: str) -> Dict[str, str]:
    payload = {"facilityId": _normalize_facility_id_from_excel(row.get(facility_id_col))}
    optional_fields = {
        "facilityName": next(
            (c for c in df.columns if "health centre name" in str(c).lower() or "facility name" in str(c).lower()),
            None,
        ),
        "facilityCategory": next(
            (c for c in df.columns if "category of facility" in str(c).lower() or "facility category" in str(c).lower()),
            None,
        ),
        "facilityType": next(
            (c for c in df.columns if "type of hc" in str(c).lower() or "facility type" in str(c).lower()),
            None,
        ),
        "district": next((c for c in df.columns if str(c).lower().strip() == "district"), None),
        "block": next((c for c in df.columns if str(c).lower().strip() == "block"), None),
    }
    for key, col in optional_fields.items():
        value = _row_value(row, col)
        if value:
            payload[key] = value
    return payload


def _parse_upstream_error_detail(response: Optional[requests.Response]) -> str:
    if response is None:
        return "Assessment bulk-create request failed"
    body = (response.text or "").strip()
    if not body:
        return f"Assessment bulk-create failed with HTTP {response.status_code}"
    try:
        parsed = response.json()
    except ValueError:
        return body
    errors = parsed.get("Errors") or parsed.get("errors") or []
    if isinstance(errors, list) and errors:
        messages = []
        for err in errors:
            if isinstance(err, dict):
                code = err.get("code") or err.get("errorCode") or "ERROR"
                message = err.get("message") or err.get("description") or code
                messages.append(f"{code}: {message}")
            else:
                messages.append(str(err))
        return "; ".join(messages)
    if isinstance(parsed.get("message"), str):
        return parsed["message"]
    return body


def _validate_assessment_include_rows(
        df: pd.DataFrame,
        include_col: str,
        facility_id_col: str,
        linked_ids: Set[str],
        plan_facility_ids: Set[str],
) -> List[List[str]]:
    validation_errors: List[List[str]] = []
    for _, row in df.iterrows():
        errs: List[str] = []
        include_val = str(row.get(include_col, "")).strip().lower()
        if include_val != "yes":
            validation_errors.append(errs)
            continue
        facility_id = _normalize_facility_id_from_excel(row.get(facility_id_col))
        if not facility_id:
            errs.append("Facility Id is required for Yes rows")
        elif facility_id not in linked_ids:
            errs.append("Facility is not linked to this project")
        elif facility_id in plan_facility_ids:
            errs.append("Facility already included in this assessment plan")
        validation_errors.append(errs)
    return validation_errors


def _apply_validation_results_to_dataframe(
        df: pd.DataFrame,
        validation_errors: List[List[str]],
) -> int:
    error_count = 0
    for i, errs in enumerate(validation_errors):
        if errs:
            df.at[i, "status"] = "FAILED"
            df.at[i, "error"] = "; ".join(dict.fromkeys(errs))
            error_count += 1
        else:
            df.at[i, "status"] = "PASSED"
            df.at[i, "error"] = ""
    return error_count


def _write_validated_facility_sheet(
        wb,
        df: pd.DataFrame,
        facility_sheet_name: str,
) -> None:
    ws = wb[facility_sheet_name]
    header_values = [cell.value for cell in ws[1]]

    for col_name in ["status", "error"]:
        if col_name not in header_values:
            new_col_idx = len(header_values) + 1
            cell = ws.cell(row=1, column=new_col_idx, value=col_name)
            cell.font = Font(bold=True)
            header_values.append(col_name)
            cell.protection = Protection(locked=True)
            for r_idx in range(2, ws.max_row + 1):
                ws.cell(row=r_idx, column=new_col_idx).protection = Protection(locked=True)

    grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    export_df = prepare_dataframe_for_excel_export(df)
    for r_idx, row in enumerate(dataframe_to_rows(export_df, index=False, header=False), start=2):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if ws.cell(1, c_idx).value in ["status", "error"]:
                cell.protection = Protection(locked=True)
                cell.fill = grey_fill

    ws.protection.sheet = True
    ws.protection.enable()


@router.post(
    '/assessmentPlanIncludeValidateData',
    summary='Validate assessment plan facility include Excel before apply',
    response_description='Returns validation report Excel with PASSED/FAILED rows',
)
async def validate_assessment_plan_include_data(
        background_tasks: BackgroundTasks,
        include_file: UploadFile = File(..., description="Assessment plan include Excel file"),
        plan_id: str = Form(description="Assessment plan ID"),
        project_id: str = Form(description="Project ID"),
        tenant_id: str = Form(default="in"),
        request_info: str = Form(default=""),
        facility_sheet_name: str = Form(default="FacilityMapping",
                                        description="Name of the sheet containing facility data"),
):
    if not fieldPlan_service_url:
        raise HTTPException(status_code=500, detail="FIELDPLAN_SERVICE_URL is not configured")

    request_info_obj = request_info_from_json(request_info)
    temp_input_file = None
    try:
        temp_input_file, _ = await _save_upload_to_temp_file(include_file, suffix=".xlsx")
        wb = load_workbook(temp_input_file.name)

        if facility_sheet_name not in wb.sheetnames:
            for candidate in ("FacilityMapping", "AssessmentInclude"):
                if candidate in wb.sheetnames:
                    facility_sheet_name = candidate
                    break
            else:
                raise HTTPException(
                    status_code=400,
                    detail=f"Facility sheet '{facility_sheet_name}' not found",
                )

        df = pd.read_excel(temp_input_file.name, sheet_name=facility_sheet_name)
        df.columns = [str(c).strip() for c in df.columns]
        df = df.loc[:, ~df.columns.str.startswith("Unnamed")]

        include_col = _find_assessment_include_col(df)
        if include_col is None:
            raise HTTPException(status_code=400, detail="Include in Assessment Plan column not found")

        facility_id_col = _find_assessment_facility_id_col(df)
        if facility_id_col is None:
            raise HTTPException(status_code=400, detail="Facility Id column not found")

        if "status" not in df.columns:
            df["status"] = ""
        if "error" not in df.columns:
            df["error"] = ""

        project_client = ProjectServiceClient(project_service_url)
        linked = project_client.search_project_facility(request_info_obj, project_id)
        linked_ids = {
            pf.get("facilityId") for pf in (linked.get("ProjectFacilities", []) or []) if pf.get("facilityId")
        }

        assessment_client = AssessmentServiceClient(fieldPlan_service_url)
        plan_facility_response = assessment_client.search_plan_facilities(
            request_info_obj, plan_id, export_all=True
        )
        plan_facility_ids = {
            f.get("facilityId")
            for f in (plan_facility_response.get("facilities", []) or [])
            if f.get("facilityId")
        }

        validation_errors = _validate_assessment_include_rows(
            df, include_col, facility_id_col, linked_ids, plan_facility_ids
        )
        error_count = _apply_validation_results_to_dataframe(df, validation_errors)
        _write_validated_facility_sheet(wb, df, facility_sheet_name)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_temp_file_path = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx").name
        wb.save(output_temp_file_path)
        autofit_columns(output_temp_file_path, facility_sheet_name, auto_fit=True)
        background_tasks.add_task(cleanup_temp_file, output_temp_file_path)

        response = FileResponse(
            path=output_temp_file_path,
            filename=f"assessment_plan_include_validation_results_{timestamp}.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response.headers["X-Error-Count"] = str(error_count)
        return response
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Assessment include validation failed: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Validation failed: {str(e)}")
    finally:
        if temp_input_file and os.path.exists(temp_input_file.name):
            os.unlink(temp_input_file.name)


def _write_facility_apply_results_workbook(
        wb,
        df: pd.DataFrame,
        facility_sheet_name: str,
        output_path: str,
) -> None:
    ws = wb[facility_sheet_name]
    header_values = [cell.value for cell in ws[1]]
    for col_name in df.columns:
        if col_name not in header_values:
            cell = ws.cell(row=1, column=len(header_values) + 1, value=col_name)
            cell.font = Font(bold=True)
            header_values.append(col_name)

    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    export_df = prepare_dataframe_for_excel_export(df)
    for r_idx, row in enumerate(dataframe_to_rows(export_df, index=False, header=False), start=2):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    wb.save(output_path)


@router.post(
    '/assessmentPlanIncludeApply',
    summary='Apply assessment plan facility include Excel (Yes rows only)',
    response_description='Returns processed Excel file with assessment include results',
)
async def apply_assessment_plan_include_data(
        background_tasks: BackgroundTasks,
        include_file: UploadFile = File(..., description="Assessment plan include Excel file"),
        plan_id: str = Form(description="Assessment plan ID"),
        project_id: str = Form(description="Project ID"),
        tenant_id: str = Form(default="in"),
        request_info: str = Form(default=""),
        facility_sheet_name: str = Form(default="FacilityMapping",
                                        description="Name of the sheet containing facility data"),
):
    if not fieldPlan_service_url:
        raise HTTPException(status_code=500, detail="FIELDPLAN_SERVICE_URL is not configured")

    request_info_obj = request_info_from_json(request_info)
    assessment_client = AssessmentServiceClient(fieldPlan_service_url)
    temp_input_file = None
    output_file_path = None
    try:
        temp_input_file, _ = await _save_upload_to_temp_file(include_file, suffix=".xlsx")
        wb = load_workbook(temp_input_file.name)

        if facility_sheet_name not in wb.sheetnames:
            for candidate in ("FacilityMapping", "AssessmentInclude"):
                if candidate in wb.sheetnames:
                    facility_sheet_name = candidate
                    break
            else:
                raise HTTPException(
                    status_code=400,
                    detail=f"Facility sheet '{facility_sheet_name}' not found",
                )

        df = pd.read_excel(temp_input_file.name, sheet_name=facility_sheet_name)
        df.columns = [str(c).strip() for c in df.columns]
        df = df.loc[:, ~df.columns.str.startswith("Unnamed")]

        include_col = _find_assessment_include_col(df)
        if include_col is None:
            raise HTTPException(status_code=400, detail="Include in Assessment Plan column not found")

        facility_id_col = _find_assessment_facility_id_col(df)
        if facility_id_col is None:
            raise HTTPException(status_code=400, detail="Facility Id column not found")

        if "status" not in df.columns or "error" not in df.columns:
            raise HTTPException(
                status_code=400,
                detail="Missing 'status'/'error' columns. Please upload validated file.",
            )

        failed_rows = df[df["status"].astype(str).str.upper() != "PASSED"]
        if not failed_rows.empty:
            raise HTTPException(
                status_code=400,
                detail="Validation failed: Some rows are not marked as PASSED. Please upload a fully validated file.",
            )

        result_col = "Assessment Plan Include Status"
        if result_col not in df.columns:
            df[result_col] = ""

        yes_row_indices: List[int] = []
        facilities = []
        for index, row in df.iterrows():
            include_val = str(row.get(include_col, "")).strip().lower()
            if include_val != "yes":
                df.at[index, result_col] = "Skipped (Include in Assessment Plan != Yes)"
                continue

            facility_id = _normalize_facility_id_from_excel(row.get(facility_id_col))
            if not facility_id:
                df.at[index, result_col] = "Skipped (missing Facility Id)"
                continue

            yes_row_indices.append(index)
            facilities.append(_build_assessment_include_facility_payload(row, df, facility_id_col))

        if facilities:
            try:
                response = assessment_client.bulk_create_plan_facilities(
                    request_info_obj, plan_id, tenant_id, facilities
                )
                created = response.get("created", []) or []
                api_errors = response.get("errors", []) or []

                created_ids = {
                    _normalize_facility_id_from_excel(item.get("facilityId"))
                    for item in created
                    if item.get("facilityId")
                }
                error_by_facility_id = {
                    _normalize_facility_id_from_excel(item.get("facilityId")): item
                    for item in api_errors
                    if item.get("facilityId")
                }

                for index in yes_row_indices:
                    facility_id = _normalize_facility_id_from_excel(df.at[index, facility_id_col])
                    if facility_id in created_ids:
                        df.at[index, result_col] = "Included"
                    elif facility_id in error_by_facility_id:
                        err = error_by_facility_id[facility_id]
                        message = err.get("message") or err.get("code") or "Include failed"
                        df.at[index, result_col] = f"Failed: {message}"
                    else:
                        df.at[index, result_col] = "Not Attempted"
            except requests.HTTPError as http_err:
                detail = _parse_upstream_error_detail(http_err.response)
                for index in yes_row_indices:
                    df.at[index, result_col] = f"Failed: {detail}"

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"assessment_plan_include_apply_results_{timestamp}.xlsx"
        output_temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        output_temp_file.close()
        output_file_path = output_temp_file.name

        _write_facility_apply_results_workbook(wb, df, facility_sheet_name, output_file_path)
        autofit_columns(output_file_path, facility_sheet_name, auto_fit=True)
        background_tasks.add_task(cleanup_temp_file, output_file_path)

        return FileResponse(
            path=output_file_path,
            filename=output_filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Assessment include apply failed: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Apply failed: {str(e)}")
    finally:
        if temp_input_file and os.path.exists(temp_input_file.name):
            os.unlink(temp_input_file.name)