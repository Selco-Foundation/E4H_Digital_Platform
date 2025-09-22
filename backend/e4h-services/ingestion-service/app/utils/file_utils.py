import os
import pandas as pd
import tempfile

from openpyxl import load_workbook

from app.core.logging import AppLogger
from app.ingest.excel_data_writer import ExcelDataWriter

logger = AppLogger().get_logger()


def create_temp_file(suffix: str = None) -> str:
    """Create a temporary file and return its path"""
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
    temp_file.close()
    return temp_file.name


def cleanup_temp_file(file_path: str) -> None:
    """Clean up a temporary file if it exists"""
    try:
        if file_path and os.path.exists(file_path):
            os.unlink(file_path)
    except Exception as e:
        logger.error(f"Error cleaning up temporary file {file_path}: {e}")


def create_excel_data_writer(file_path: str, output_sheet: str) -> ExcelDataWriter:
    """Factory method to create an ExcelDataWriter instance"""
    return ExcelDataWriter(file_path, output_sheet)


def create_empty_excel_file(file_path: str) -> None:
    """Create an empty Excel file"""
    try:
        pd.DataFrame().to_excel(file_path, index=False)
        logger.info(f"Created empty Excel file at {file_path}")
    except Exception as e:
        logger.error(f"Error creating empty Excel file: {e}")
        raise

def remove_default_empty_sheet(file_path: str) -> None:
    """
    Removes the default empty sheet (Sheet1) if it exists and there are other sheets present.
    This should be called at the end of template generation.
    """
    try:
        wb = load_workbook(file_path)
        if "Sheet1" in wb.sheetnames and len(wb.sheetnames) > 1:
            std = wb["Sheet1"]
            wb.remove(std)
            wb.save(file_path)
            logger.info("Removed default empty sheet from workbook")
        else:
            logger.info("No default empty sheet found or only one sheet exists")
    except Exception as e:
        logger.error(f"Error removing default empty sheet: {e}")
        raise

