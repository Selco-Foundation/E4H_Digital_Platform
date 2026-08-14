"""
Helpers for assessment eligible → installation field plan handoff (LLD §2.2.7, §2.2.9, API §8).
"""
from __future__ import annotations

from typing import Any, Dict, List, Optional

import pandas as pd

from app.core.logging import AppLogger
from app.schemas.request_info import RequestInfo
from app.utils.assessment_service_client import AssessmentServiceClient

logger = AppLogger().get_logger()

PLAN_FACILITY_ID_COLUMN = "Plan Facility Id"
ASSESSMENT_PLAN_ID_COLUMN = "Assessment Plan Id"
ASSESSMENT_PLAN_NAME_COLUMN = "Assessment Plan Name"


def should_use_assessment_handoff_validation(
    assessment_client: Optional[AssessmentServiceClient],
    request_info: RequestInfo,
    project_id: Optional[str],
    tenant_id: str,
) -> bool:
    """True when there are eligible, unassigned facilities ready for field-plan handoff."""
    if not project_id or assessment_client is None:
        return False
    try:
        response = assessment_client.search_eligible_facilities(
            request_info=request_info,
            project_id=project_id,
            tenant_id=tenant_id,
        )
        facilities = response.get("facilities") or []
        if facilities:
            logger.info(
                "Project %s has %s eligible assessment facility(ies) for handoff validation",
                project_id,
                len(facilities),
            )
            return True
        return False
    except Exception as exc:
        logger.warning(
            "Could not load eligible assessment facilities for project %s: %s",
            project_id,
            exc,
        )
        return False


def fetch_eligible_assessment_facilities(
    assessment_client: AssessmentServiceClient,
    request_info: RequestInfo,
    project_id: str,
    tenant_id: str,
) -> List[Dict[str, Any]]:
    response = assessment_client.search_eligible_facilities(
        request_info=request_info,
        project_id=project_id,
        tenant_id=tenant_id,
    )
    return response.get("facilities") or []


def append_assessment_handoff_columns(
    file_path: str,
    sheet_name: str,
    eligible_facilities: List[Dict[str, Any]],
) -> None:
    """Prepend assessment handoff columns; populate rows that match eligible facility ids."""
    from openpyxl import load_workbook

    meta_by_facility_id = {
        str(entry.get("facilityId")): entry
        for entry in eligible_facilities
        if entry.get("facilityId")
    }
    if not meta_by_facility_id:
        return

    workbook = load_workbook(file_path)
    worksheet = workbook[sheet_name]
    worksheet.insert_cols(1, 3)
    worksheet.cell(row=1, column=1, value=PLAN_FACILITY_ID_COLUMN)
    worksheet.cell(row=1, column=2, value=ASSESSMENT_PLAN_ID_COLUMN)
    worksheet.cell(row=1, column=3, value=ASSESSMENT_PLAN_NAME_COLUMN)

    facility_id_col = None
    for col_idx in range(1, worksheet.max_column + 1):
        header = worksheet.cell(row=1, column=col_idx).value
        if header and "Facility Id" in str(header):
            facility_id_col = col_idx
            break
    if facility_id_col is None:
        workbook.save(file_path)
        return

    for row_idx in range(2, worksheet.max_row + 1):
        raw_facility_id = worksheet.cell(row=row_idx, column=facility_id_col).value
        facility_id = str(raw_facility_id).strip() if raw_facility_id is not None else ""
        meta = meta_by_facility_id.get(facility_id, {})
        worksheet.cell(row=row_idx, column=1, value=meta.get("planFacilityId", ""))
        worksheet.cell(row=row_idx, column=2, value=meta.get("assessmentPlanId", ""))
        worksheet.cell(row=row_idx, column=3, value=meta.get("assessmentPlanName", ""))
    workbook.save(file_path)


def merge_eligible_facilities_into_list(
    all_facilities: List[Dict[str, Any]],
    eligible_facilities: List[Dict[str, Any]],
    facilities_by_id: Dict[str, Dict[str, Any]],
) -> List[Dict[str, Any]]:
    """Ensure eligible assessment facilities appear even if geography filtering skipped them."""
    existing_ids = {
        str(f.get("facility_id") or f.get("facilityId"))
        for f in all_facilities
        if f.get("facility_id") or f.get("facilityId")
    }
    merged = list(all_facilities)
    for eligible in eligible_facilities:
        facility_id = eligible.get("facilityId")
        if not facility_id or str(facility_id) in existing_ids:
            continue
        facility = dict(facilities_by_id.get(facility_id) or {})
        if not facility.get("facility_id"):
            facility["facility_id"] = facility_id
        facility["include_in_fieldplan"] = "No"
        merged.append(facility)
        existing_ids.add(str(facility_id))
    return merged


def load_eligible_facility_map(
    assessment_client: AssessmentServiceClient,
    request_info: RequestInfo,
    project_id: str,
    tenant_id: str,
) -> Dict[str, Dict[str, Any]]:
    """Map planFacilityId → eligible facility record (all closed plans in project)."""
    response = assessment_client.search_eligible_facilities(
        request_info=request_info,
        project_id=project_id,
        tenant_id=tenant_id,
    )
    facilities = response.get("facilities") or []
    return {
        str(f.get("planFacilityId")): f
        for f in facilities
        if f.get("planFacilityId")
    }


def find_column(df: pd.DataFrame, partial: str) -> Optional[str]:
    partial_lower = partial.lower()
    for col in df.columns:
        if partial_lower in str(col).lower():
            return col
    return None


def validate_assessment_handoff_rows(
    df: pd.DataFrame,
    eligible_by_plan_facility_id: Dict[str, Dict[str, Any]],
) -> List[List[str]]:
    """Return per-row validation error messages (empty list = passed)."""
    df = df.reset_index(drop=True)
    errors: List[List[str]] = [[] for _ in range(len(df))]

    plan_facility_col = find_column(df, "plan facility id")
    include_col = find_column(df, "included in field plan")

    for i, row in df.iterrows():
        include_val = ""
        if include_col:
            include_val = str(row.get(include_col, "")).strip().lower()
        elif "Included in Field Plan (Mandatory)" in df.columns:
            include_val = str(row.get("Included in Field Plan (Mandatory)", "")).strip().lower()
        if include_val != "yes":
            continue

        plan_facility_id = ""
        if plan_facility_col:
            plan_facility_id = str(row.get(plan_facility_col, "")).strip()
        if not plan_facility_id or plan_facility_id.lower() in ("nan", "none"):
            # Legacy field-plan include without assessment handoff.
            continue

        if plan_facility_id not in eligible_by_plan_facility_id:
            errors[i].append(
                f"ASSESSMENT_FACILITY_NOT_ELIGIBLE: planFacilityId '{plan_facility_id}' "
                "is not ELIGIBLE or already handed off"
            )

    return errors


def merge_assessment_validation_errors(
    base_errors: List[List[str]],
    assessment_errors: List[List[str]],
) -> List[List[str]]:
    merged: List[List[str]] = []
    for i in range(max(len(base_errors), len(assessment_errors))):
        row_errors = []
        if i < len(base_errors):
            row_errors.extend(base_errors[i])
        if i < len(assessment_errors):
            row_errors.extend(assessment_errors[i])
        merged.append(row_errors)
    return merged


def build_assessment_fieldplan_template_rows(
    eligible_facilities: List[Dict[str, Any]],
    facilities_by_id: Dict[str, Dict[str, Any]],
) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for eligible in eligible_facilities:
        facility_id = eligible.get("facilityId")
        facility = facilities_by_id.get(facility_id, {})
        rows.append({
            PLAN_FACILITY_ID_COLUMN: eligible.get("planFacilityId", ""),
            ASSESSMENT_PLAN_ID_COLUMN: eligible.get("assessmentPlanId", ""),
            ASSESSMENT_PLAN_NAME_COLUMN: eligible.get("assessmentPlanName", ""),
            "Facility Id": facility_id or "",
            "Health Centre Name (Mandatory)": (
                eligible.get("facilityName")
                or facility.get("facility_name")
                or facility.get("facilityName")
                or ""
            ),
            "Category of Facility (Mandatory)": facility.get("facility_category") or facility.get("facilityCategory") or "",
            "Type of HC (Mandatory)": facility.get("facility_type") or facility.get("facilityType") or "",
            "Boundary Code (Mandatory)": facility.get("boundary_code") or facility.get("boundaryCode") or "",
            "Included in Field Plan (Mandatory)": "",
        })
    return rows


def extract_assessment_link_meta(row: pd.Series, df: pd.DataFrame) -> tuple[Optional[str], Optional[str]]:
    plan_facility_col = find_column(df, "plan facility id")
    assessment_plan_col = find_column(df, "assessment plan id")
    plan_facility_id = None
    assessment_plan_id = None
    if plan_facility_col:
        val = row.get(plan_facility_col)
        if pd.notna(val) and str(val).strip():
            plan_facility_id = str(val).strip()
    if assessment_plan_col:
        val = row.get(assessment_plan_col)
        if pd.notna(val) and str(val).strip():
            assessment_plan_id = str(val).strip()
    return plan_facility_id, assessment_plan_id
