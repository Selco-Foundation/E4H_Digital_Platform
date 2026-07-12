"""
Validate and convert an ICC (Installation Completion Certificate) report Excel file into a
flat JSON keyed by the {{placeholder}} field names used in the pdf-service BOM format-config
templates bundled under app/config/icc_templates/.

This mirrors the standalone icc_report_to_json.py script (built and verified against real
sample files for the "dc", "ac_off", "hybrid", and "ac_on_grid" system types), reshaped into
importable functions for the /icc-reports endpoint: structural validation happens before
conversion, and conversion never guesses a System Type the caller didn't ask for - see
validate_icc_report().

Ground truth for "Field Label -> JSON key" comes from the templates directly (they place
{{field_name}} placeholders right next to the same item labels used in the Excel form), not
from fuzzy-matching against an MDMS schema. See detect_system_type()'s docstring for how the
4 supported formats are told apart.
"""
import json
import math
import os
import re
from collections import OrderedDict

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

_BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_DIR = os.path.abspath(os.path.join(_BASE_DIR, "..", "config", "icc_templates"))

DATA_SHEET = "ICC Report"
MAP_SHEET = "Data_Ingestion_Map"
REQUIRED_SHEETS = (DATA_SHEET, MAP_SHEET)
REQUIRED_MAP_COLUMNS = ("Section", "Field Label", "Cell Address", "Row No.", "Column No.", "Expected Input Type")

# Public systemType values the /icc-reports endpoint accepts -> internal format key used by
# TEMPLATE_BY_TYPE/SFP_FIELD_MAPS/detect_system_type().
#
# facility.SystemType now issues a separate MDMS code per phase (e.g. AC_ON_GRID_SINGLE_PHASE /
# AC_ON_GRID_THREE_PHASE), but phase never changes which Excel cell maps to which template
# key (see detect_system_type()'s docstring) - so every phase variant of a given system type
# maps to the same internal key here. The pre-phase-split codes (DC_OFF_GRID, AC_HYBRID,
# AC_ON_GRID) are kept as aliases for backward compatibility with any caller still sending them.
SYSTEM_TYPE_TO_INTERNAL = {
    # Legacy codes (pre phase-split), kept as aliases.
    "DC_OFF_GRID": "dc",
    "AC_HYBRID": "hybrid",
    "AC_ON_GRID": "ac_on_grid",
    # Current facility.SystemType MDMS codes.
    "DC": "dc",
    "AC_OFF_GRID": "ac_off",
    "AC_OFF_GRID_THREE_PHASE": "ac_off",
    "HYBRID_SINGLE_PHASE": "hybrid",
    "HYBRID_THREE_PHASE": "hybrid",
    "AC_ON_GRID_SINGLE_PHASE": "ac_on_grid",
    "AC_ON_GRID_THREE_PHASE": "ac_on_grid",
}

TEMPLATE_BY_TYPE = {
    "dc": "bom_dc_system.json",
    "hybrid": "bom_hybrid_three.json",  # identical to bom_hybrid_single.json
    "ac_off": "bom_ac_off_three.json",  # cosmetically identical to bom_ac_off_single.json
    "ac_on_grid": "bom_ac_on_grid_single.json",
}

SECTION_TEMPLATE_HEADERS = {
    "Bill Of Material (For Solar System)": "Bill Of Material(For Solar System)",
    "Bill of material (For Luminaries & Fans)": "Bill of material (For Luminaries & Fans)",
    "Bill of materials (For RMS)": "Bill of material (For RMS)",
    "Bill Of Material (For Load Wiring)": "Bill of material (For Load Wiring)",
    # ac_on_grid's Data_Ingestion_Map names its (only) BOM section just "Bill Of Material",
    # without the "(For Solar System)" qualifier every other type uses.
    "Bill Of Material": "Bill Of Material(For Solar System)",
}

SERIAL_RE = re.compile(r"^\d+[a-z]?$")
PLACEHOLDER_RE = re.compile(r"\{\{\s*([a-zA-Z0-9_]+)\s*\}\}")

LABEL_ROLE_SUFFIX_RE = re.compile(
    r"\s*-\s*(make|capacity|quantity|qty\.?|remarks|description(?:\s*&\s*make)?)\s*$",
    re.IGNORECASE,
)

NO_TEMPLATE_COVERAGE_PREFIXES = ("header", "image", "annexure")


class ICCValidationError(Exception):
    """Raised for any Validation 1 / Validation 2 failure - message is user-facing."""


def normalize_label(label):
    return re.sub(r"\s+", " ", str(label)).strip().lower().rstrip(":.")


TEMPLATE_LABEL_MERGE = {
    normalize_label("Battery rack with the following: 1. Acid absorbent mat"): "battery rack (merged)",
    normalize_label("Battery rack with the following: 2. Electrical Insulation mat (Minimum (0.4 kV)"): "battery rack (merged)",
}

SFP_FIELD_MAP_HYBRID = {
    (141, 3): "array_size_kwp",
    (141, 8): "array_no_modules",
    (141, 13): "array_no_strings",
    (141, 18): "array_no_modules_per_string",
    (141, 23): "array_each_module_wp",
    (148, 12): "ajb_off_voc_string_1",
    (149, 12): "ajb_off_voc_string_2",
    (150, 12): "ajb_off_voc_string_3",
    (151, 12): "ajb_off_voc_string_4",
    (152, 12): "ajb_off_voc_string_5",
    (153, 12): "ajb_off_voc_mv_anchor",
    (154, 12): "ajb_on_vmp_string_1", (154, 21): "ajb_on_imp_string_1",
    (155, 12): "ajb_on_vmp_string_2", (155, 21): "ajb_on_imp_string_2",
    (156, 12): "ajb_on_vmp_string_3", (156, 21): "ajb_on_imp_string_3",
    (157, 12): "ajb_on_vmp_string_4", (157, 21): "ajb_on_imp_string_4",
    (158, 12): "ajb_on_vmp_string_5", (158, 21): "ajb_on_imp_string_5",
    (159, 12): "ajb_on_vmp_mv_anchor", (159, 21): "ajb_on_imp_mv_anchor",
    (160, 12): "ajb_on_pmp_string_1",
    (161, 12): "ajb_on_pmp_string_2",
    (162, 12): "ajb_on_pmp_string_3",
    (163, 12): "ajb_on_pmp_string_4",
    (164, 12): "ajb_on_pmp_string_5",
    (165, 12): "ajb_on_pmp_array",
    (167, 4): "bbank_size_kwh",
    (167, 10): "bbank_capacity_ah",
    (167, 16): "bbank_no_batteries",
    (167, 22): "bbank_no_strings",
    (168, 4): "bbank_batteries_per_string",
    (168, 10): "bbank_each_capacity_ah",
    (168, 16): "bbank_each_voltage_v",
    (168, 22): "bbank_each_c_rating",
    (171, 17): "bb_voltage_mv",
    (172, 17): "bb_current_mv",
    (174, 7): "pcu_inverter_count",
    (174, 20): "charge_controller_count",
    (176, 20): "charge_controller_ratings",
    (181, 9): "ph1_load_voltage_mv", (181, 20): "ph1_grid_input_voltage_mv",
    (182, 9): "ph1_output_freq_mv", (182, 20): "ph1_grid_frequency_mv",
    (183, 9): "ph1_output_current_full_load_mv", (183, 20): "ph1_grid_current_mv",
    (187, 12): "ph3_off_load_inverter_voltage_rn_mv", (187, 20): "ph3_on_grid_voltage_rn_mv",
    (188, 12): "ph3_off_inverter_output_frequency_rn_mv", (188, 20): "ph3_on_grid_frequency_rn_mv",
    (189, 12): "ph3_off_inverter_full_load_current_rn_mv", (189, 20): "ph3_on_grid_current_rn_mv",
    (190, 12): "ph3_off_load_inverter_voltage_yn_mv", (190, 20): "ph3_on_grid_voltage_yn_mv",
    (191, 12): "ph3_off_inverter_output_frequency_yn_mv", (191, 20): "ph3_on_grid_frequency_yn_mv",
    (192, 12): "ph3_off_inverter_full_load_current_yn_mv", (192, 20): "ph3_on_grid_current_yn_mv",
    (193, 12): "ph3_off_load_inverter_voltage_bn_mv", (193, 20): "ph3_on_grid_voltage_bn_mv",
    (194, 12): "ph3_off_inverter_output_frequency_bn_mv", (194, 20): "ph3_on_grid_frequency_bn_mv",
    (195, 12): "ph3_off_inverter_full_load_current_bn_mv", (195, 20): "ph3_on_grid_current_bn_mv",
    (196, 12): "ph3_off_inverter_output_voltage_ry_mv", (196, 20): "ph3_on_grid_voltage_ry_mv",
    (197, 12): "ph3_off_inverter_output_frequency_ry_mv", (197, 20): "ph3_on_grid_frequency_ry_mv",
    (198, 12): "ph3_off_inverter_output_voltage_rb_mv", (198, 20): "ph3_on_grid_voltage_rb_mv",
    (199, 12): "ph3_off_inverter_output_frequency_rb_mv", (199, 20): "ph3_on_grid_frequency_rb_mv",
    (200, 12): "ph3_off_inverter_output_voltage_yb_mv", (200, 20): "ph3_on_grid_voltage_yb_mv",
    (201, 12): "ph3_off_inverter_output_frequency_yb_mv", (201, 20): "ph3_on_grid_frequency_yb_mv",
    (213, 13): "array_pcu_isolator_on_v", (213, 22): "array_pcu_isolator_off_v",
    (214, 13): "battery_pcu_isolator_on_v", (214, 22): "battery_pcu_isolator_off_v",
    (215, 13): "grid_pcu_isolator_on_v", (215, 22): "grid_pcu_isolator_off_v",
    (216, 13): "pcu_cos1_isolator_on_v", (216, 22): "pcu_cos1_isolator_off_v",
    (217, 13): "rccb_load_acdb_r_on_v", (217, 22): "rccb_load_acdb_r_off_v",
    (218, 13): "rccb_load_acdb_y_on_v", (218, 22): "rccb_load_acdb_y_off_v",
    (219, 13): "rccb_load_acdb_b_on_v", (219, 22): "rccb_load_acdb_b_off_v",
    (220, 13): "isolator_load_acdb_r_on_v", (220, 22): "isolator_load_acdb_r_off_v",
    (223, 13): "main_mcb_load_acdb_on_v", (223, 22): "main_mcb_load_acdb_off_v",
    (253, 18): "sock_v_pn_mv",
    (254, 18): "sock_v_pe_mv",
    (255, 18): "sock_v_en_mv",
    (259, 1): "ept_dc_earthing_mv",
    (259, 6): "ept_ac_earthing_mv",
    (259, 13): "ept_lightning_arrester1_mv",
    (259, 18): "ept_lightning_arrester2_mv",
    (262, 19): "dc_panel_panel_res",
    (263, 19): "dc_panel_mms_res",
    (264, 19): "dc_mms_ajb_res",
    (265, 19): "dc_ajb_busbar_res",
    (266, 19): "dc_battery_rack_busbar_res",
    (267, 19): "dc_isolator_box_busbar_res",
    (268, 19): "dc_busbar_earth_pit_res",
    (269, 19): "ac_pcu_busbar_res",
    (270, 19): "ac_changeover1_busbar_res",
    (271, 19): "ac_changeover2_busbar_res",
    (272, 19): "ac_gipb_busbar_res",
    (273, 19): "ac_load_acdb_busbar_res",
    (274, 19): "ac_busbar_earth_pit_res",
    (275, 19): "la1_la_aluminium_cable_res",
    (276, 19): "la1_aluminium_gi_strip_res",
    (277, 19): "la1_gi_strip_earth_pit_res",
    (278, 19): "la2_la_aluminium_cable_res",
    (279, 19): "la2_aluminium_gi_strip_res",
    (280, 19): "la2_gi_strip_earth_pit_res",
    (282, 19): "la1_la2_res",
    (285, 15): "orientation_operational_instructions_status",
    (286, 15): "orientation_routine_maintenance_status",
    (287, 15): "stickers_pasted_status",
    (288, 15): "posters_pasted_status",
    (289, 15): "orientation_warranty_solar_panel_status",
    (290, 15): "orientation_warranty_batteries_status",
    (291, 15): "orientation_warranty_pcu_status",
    (292, 15): "orientation_warranty_bidirectional_meter_status",
    (293, 15): "orientation_warranty_check_meter_status",
}

SFP_FIELD_MAP_AC_OFF = {
    (138, 1): "sfp_date",
    (139, 6): "sfp_weather",
    (141, 4): "array_size_kwp",
    (141, 10): "array_no_modules",
    (141, 14): "array_no_strings",
    (141, 19): "array_no_modules_per_string",
    (141, 23): "array_each_module_wp",
    (144, 8): "ajb_off_voc_string_1",
    (145, 8): "ajb_off_voc_string_2",
    (146, 8): "ajb_off_voc_string_3",
    (147, 8): "ajb_off_voc_string_4",
    (148, 8): "ajb_off_voc_string_5",
    (149, 8): "ajb_off_voc_mv_anchor",
    (150, 8): "ajb_on_vmp_string_1", (150, 18): "ajb_on_imp_string_1",
    (151, 8): "ajb_on_vmp_string_2", (151, 18): "ajb_on_imp_string_2",
    (152, 8): "ajb_on_vmp_string_3", (152, 18): "ajb_on_imp_string_3",
    (153, 8): "ajb_on_vmp_string_4", (153, 18): "ajb_on_imp_string_4",
    (154, 8): "ajb_on_vmp_string_5", (154, 18): "ajb_on_imp_string_5",
    (155, 8): "ajb_on_vmp_mv_anchor", (155, 18): "ajb_on_imp_mv_anchor",
    (156, 8): "ajb_on_pmp_string_1",
    (157, 8): "ajb_on_pmp_string_2",
    (158, 8): "ajb_on_pmp_string_3",
    (159, 8): "ajb_on_pmp_string_4",
    (160, 8): "ajb_on_pmp_string_5",
    (161, 8): "ajb_on_pmp_array",
    (164, 15): "bb_voltage_mv",
    (165, 15): "bb_current_mv",
    (169, 9): "ph1_load_voltage_mv", (169, 20): "ph1_grid_input_voltage_mv",
    (170, 9): "ph1_output_freq_mv", (170, 20): "ph1_inverter_output_frequency_mv",
    (171, 9): "ph1_output_current_full_load_mv", (171, 20): "ph1_grid_input_current_mv",
    (175, 12): "ph3_rpn_voltage_mv_off", (175, 20): "ph3_rpn_voltage_mv_on",
    (176, 12): "ph3_rpn_freq_mv_off", (176, 20): "ph3_rpn_freq_mv_on",
    (177, 12): "inv_full_i_rn", (177, 20): "ph3_ypn_voltage_mv_on",
    (178, 12): "ph3_ypn_voltage_mv_off", (178, 20): "ph3_ypn_freq_mv_on",
    (179, 12): "ph3_ypn_freq_mv_off", (179, 20): "ph3_bpn_voltage_mv_on",
    (180, 12): "inv_full_i_yn", (180, 20): "ph3_bpn_freq_mv_on",
    (181, 12): "ph3_bpn_voltage_mv_off", (181, 20): "grid_on_v_ry",
    (182, 12): "ph3_bpn_freq_mv_off", (182, 20): "grid_on_f_ry",
    (183, 12): "inv_full_i_bn", (183, 20): "grid_on_v_rb",
    (184, 12): "ph3_r_y_voltage_mv", (184, 20): "grid_on_f_rb",
    (185, 12): "ph3_r_b_voltage_mv", (185, 20): "grid_on_v_by",
    (186, 12): "ph3_y_b_voltage_mv", (186, 20): "grid_on_f_by",
    (189, 6): "disp_inverter_priority",
    (190, 6): "disp_grid_export",
    (191, 6): "disp_battery_settings_ah",
    (192, 6): "disp_load_running_on",
    (194, 12): "chg_switch_orientation",
    (195, 12): "chg_switch_functional",
    (198, 17): "sock_v_pn_mv",
    (199, 17): "sock_v_pe_mv",
    (200, 17): "sock_v_en_mv",
    (204, 4): "ept_dc_earthing_mv",
    (204, 9): "ept_ac_earthing_mv",
    (204, 14): "ept_lightning_arrester1_mv",
    (204, 19): "ept_lightning_arrester2_mv",
    (207, 13): "orientation_operational_instructions_status",
    (208, 13): "orientation_routine_maintenance_status",
    (209, 13): "stickers_pasted_status",
    (210, 13): "posters_pasted_status",
}

SFP_FIELD_MAP_DC = {
    (76, 1): "dc_date_of_recording",
    (76, 13): "dc_time_of_recording",
    (77, 6): "dc_weather_condition",
    (79, 4): "dc_array_size_kwp",
    (79, 10): "dc_array_no_modules",
    (79, 14): "dc_array_no_strings",
    (79, 19): "dc_array_modules_per_string",
    (79, 23): "dc_array_each_module_capacity_wp",
    (82, 8): "dc_on_vmp_rn_string1_mv",
    (83, 8): "dc_on_vmp_rn_string2_mv",
    (85, 8): "ajb_on_vmp_string_1", (85, 18): "dc_on_imp_rn_string1_mv",
    (86, 8): "ajb_on_vmp_string_2", (86, 18): "dc_on_imp_rn_string2_mv",
    (87, 18): "dc_on_imp_rn_array_mv",
    (93, 15): "bb_voltage_mv",
    (94, 15): "bb_current_mv",
    (97, 15): "dc_ccu_output_voltage_mv",
    (98, 15): "dc_ccu_output_current_full_load_mv",
    (101, 9): "dc_ccu_ldv_set_value",
    (101, 20): "dc_ccu_lrv_set_value",
    (102, 6): "dc_battery_type",
    (105, 13): "orientation_operational_instructions_status",
    (106, 13): "orientation_routine_maintenance_status",
    (107, 13): "stickers_pasted_status",
    (108, 13): "posters_pasted_status",
}

# Verified against ICC_Report_Ongrid.xlsx + bom_ac_on_grid_single.json. Unlike the other 3
# sample files, this one's Data_Ingestion_Map labels are already fully descriptive and its row
# order lines up with the template's 1:1 everywhere, so no header-text reconstruction or
# "mislabeled by a nearby static cell" workarounds were needed here.
SFP_FIELD_MAP_AC_ON_GRID = {
    (42, 1): "sfp_date",
    (42, 13): "sfp_time",
    (43, 6): "sfp_weather",
    # Array Details
    (45, 4): "array_size_kwp",
    (45, 8): "array_no_modules",
    (45, 13): "array_no_strings",
    (45, 18): "array_no_modules_per_string",
    (45, 23): "array_each_module_wp",
    # (46-49, 18) "No. of modules String 2-5" have no template key (singular
    # array_no_modules_per_string only) -> fallback
    # Array Parameters (At the AJB/CCU side) - Voc (AJB MCB OFF)
    (52, 8): "ajb_off_voc_string_1",
    (53, 8): "ajb_off_voc_string_2",
    (54, 8): "ajb_off_voc_string_3",
    (55, 8): "ajb_off_voc_string_4",
    (56, 8): "ajb_off_voc_string_5",
    (57, 8): "ajb_off_voc_mv_anchor",
    # Array Parameters (At the AJB/CCU side) - Vmp + Imp (AJB MCB ON)
    (58, 8): "ajb_on_vmp_string_1", (58, 18): "ajb_on_imp_string_1",
    (59, 8): "ajb_on_vmp_string_2", (59, 18): "ajb_on_imp_string_2",
    (60, 8): "ajb_on_vmp_string_3", (60, 18): "ajb_on_imp_string_3",
    (61, 8): "ajb_on_vmp_string_4", (61, 18): "ajb_on_imp_string_4",
    (62, 8): "ajb_on_vmp_string_5", (62, 18): "ajb_on_imp_string_5",
    (63, 8): "ajb_on_vmp_mv_anchor", (63, 18): "ajb_on_imp_mv_anchor",
    # Array Parameters (At the AJB/CCU side) - Pmp (AJB MCB ON)
    (64, 8): "ajb_on_pmp_string_1",
    (65, 8): "ajb_on_pmp_string_2",
    (66, 8): "ajb_on_pmp_string_3",
    (67, 8): "ajb_on_pmp_string_4",
    (68, 8): "ajb_on_pmp_string_5",
    (69, 8): "ajb_on_pmp_array",
    # PCU/Inverter & Charge Controller Details
    (71, 7): "pcu_inverter_count",
    (71, 20): "charge_controller_count",
    (72, 7): "pcu_inverter_rating_kva",
    (72, 10): "pcu_inverter_rating_phase",
    (72, 20): "charge_controller_ratings",
    (73, 7): "pcu_inverter_total_capacity_kva",
    (74, 7): "pcu_inverter_mppt_count",
    # 1-Phase System - PCU/Inverter parameters (no Grid ON section for 1-phase in this template)
    (77, 15): "ph1_load_voltage_mv",
    (78, 15): "ph1_output_freq_mv",
    (79, 15): "ph1_output_power_mv",
    # 3-Phase System - PCU/Inverter parameters (R-N, Y-N, B-N, R-Y, R-B, Y-B, in that order)
    (82, 15): "ph3_rpn_voltage_mv",
    (83, 15): "ph3_rpn_freq_mv",
    (84, 15): "ph3_rpn_current_mv",
    (85, 15): "ph3_ypn_voltage_mv",
    (86, 15): "ph3_ypn_freq_mv",
    (87, 15): "ph3_ypn_current_mv",
    (88, 15): "ph3_bpn_voltage_mv",
    (89, 15): "ph3_bpn_freq_mv",
    (90, 15): "ph3_bpn_current_mv",
    (91, 15): "ph3_r_y_voltage_mv",
    (92, 15): "ph3_r_y_freq_mv",
    (93, 15): "ph3_r_b_voltage_mv",
    (94, 15): "ph3_r_b_freq_mv",
    (95, 15): "ph3_y_b_voltage_mv",
    (96, 15): "ph3_y_b_freq_mv",
    # PCU/Inverter Display
    (99, 6): "pcu_display_status",
    (100, 6): "pcu_display_parameters",
    # Isolator & Breaker functionality check (col13 = ON, col23 = OFF)
    (152, 13): "pv_line1_pcu_mcb_isolator_on_v", (152, 23): "pv_line1_pcu_mcb_isolator_off_v",
    (153, 13): "pv_line2_pcu_mcb_isolator_on_v", (153, 23): "pv_line2_pcu_mcb_isolator_off_v",
    (154, 13): "pv_line3_pcu_mcb_isolator_on_v", (154, 23): "pv_line3_pcu_mcb_isolator_off_v",
    (155, 13): "pv_line4_pcu_mcb_isolator_on_v", (155, 23): "pv_line4_pcu_mcb_isolator_off_v",
    (156, 13): "pv_line5_pcu_mcb_isolator_on_v", (156, 23): "pv_line5_pcu_mcb_isolator_off_v",
    (157, 13): "acdb_contactor_ry_on_v", (157, 23): "acdb_contactor_ry_off_v",
    (158, 13): "acdb_contactor_yb_on_v", (158, 23): "acdb_contactor_yb_off_v",
    (159, 13): "acdb_contactor_rb_on_v", (159, 23): "acdb_contactor_rb_off_v",
    (160, 13): "acdb_mcb_ry_on_v", (160, 23): "acdb_mcb_ry_off_v",
    (161, 13): "acdb_mcb_yb_on_v", (161, 23): "acdb_mcb_yb_off_v",
    (162, 13): "acdb_mcb_rb_on_v", (162, 23): "acdb_mcb_rb_off_v",
    (163, 13): "lt_panel_main_mcb_ry_on_v", (163, 23): "lt_panel_main_mcb_ry_off_v",
    (164, 13): "lt_panel_main_mcb_yb_on_v", (164, 23): "lt_panel_main_mcb_yb_off_v",
    (165, 13): "lt_panel_main_mcb_rb_on_v", (165, 23): "lt_panel_main_mcb_rb_off_v",
    (166, 13): "lt_panel_phase_mcb1_ry_on_v", (166, 23): "lt_panel_phase_mcb1_ry_off_v",
    (167, 13): "lt_panel_phase_mcb1_yb_on_v", (167, 23): "lt_panel_phase_mcb1_yb_off_v",
    (168, 13): "lt_panel_phase_mcb1_rb_on_v", (168, 23): "lt_panel_phase_mcb1_rb_off_v",
    (169, 13): "lt_panel_phase_mcb2_ry_on_v", (169, 23): "lt_panel_phase_mcb2_ry_off_v",
    (170, 13): "lt_panel_phase_mcb2_yb_on_v", (170, 23): "lt_panel_phase_mcb2_yb_off_v",
    (171, 13): "lt_panel_phase_mcb2_rb_on_v", (171, 23): "lt_panel_phase_mcb2_rb_off_v",
    (172, 13): "lt_panel_phase_mcb3_ry_on_v", (172, 23): "lt_panel_phase_mcb3_ry_off_v",
    (173, 13): "lt_panel_phase_mcb3_yb_on_v", (173, 23): "lt_panel_phase_mcb3_yb_off_v",
    (174, 13): "lt_panel_phase_mcb3_rb_on_v", (174, 23): "lt_panel_phase_mcb3_rb_off_v",
    # Earth Pit Tests
    (129, 1): "ept_dc_earthing_mv",
    (129, 6): "ept_ac_earthing_mv",
    (129, 13): "ept_lightning_arrester1_mv",
    (129, 18): "ept_lightning_arrester2_mv",
    # Earth-Down Conductor Continuity Test
    (132, 19): "dc_panel_panel_res",
    (133, 19): "dc_panel_mms_res",
    (134, 19): "dc_mms_ajb_res",
    (135, 19): "dc_ajb_busbar_res",
    (136, 19): "dc_busbar_earth_pit_res",
    (137, 19): "ac_inverter_busbar_res",
    (138, 19): "ac_acdb_busbar_res",
    (139, 19): "ac_ltpanel_busbar_res",
    (140, 19): "ac_loads_busbar_res",
    (141, 19): "ac_coswitch_busbar_res",
    (142, 19): "ac_busbar_earth_pit_res",
    (143, 19): "la1_la_aluminium_cable_res",
    (144, 19): "la1_aluminium_gi_strip_res",
    (145, 19): "la1_gi_strip_earth_pit_res",
    (146, 19): "la2_la_aluminium_cable_res",
    (147, 19): "la2_aluminium_gi_strip_res",
    (148, 19): "la2_gi_strip_earth_pit_res",
    (149, 19): "earthpit_ac_dc_interconnection_res",
    (150, 19): "earthpit_la1_la2_interconnection_res",
    # Orientation Checklist
    (177, 13): "orientation_operational_instructions_status",
    (178, 13): "orientation_routine_maintenance_status",
    (179, 13): "stickers_pasted_status",
    (180, 13): "posters_pasted_status",
    (181, 13): "orientation_warranty_solar_panel_status",
    (182, 13): "orientation_warranty_pcu_inverter_ccu_status",
    (183, 13): "orientation_warranty_bidirectional_meter_status",
    (184, 13): "orientation_warranty_check_meter_status",
}

SFP_FIELD_MAPS = {
    "hybrid": SFP_FIELD_MAP_HYBRID,
    "ac_off": SFP_FIELD_MAP_AC_OFF,
    "dc": SFP_FIELD_MAP_DC,
    "ac_on_grid": SFP_FIELD_MAP_AC_ON_GRID,
}


def slugify(label):
    label = re.sub(r"\(([^)]*)\)", r" \1 ", label)
    label = label.replace("'", "").replace("’", "")
    label = label.lower()
    label = re.sub(r"[^a-z0-9]+", "_", label)
    label = re.sub(r"(?<=[a-z])(?=[0-9])", "_", label)
    label = re.sub(r"(?<=[0-9])(?=[a-z])", "_", label)
    return re.sub(r"_+", "_", label).strip("_")


def make_fallback_key(section, label, idx, cell_count):
    key = f"{slugify(section)}_{slugify(label)}"
    return key if cell_count == 1 else f"{key}_{idx + 1}"


def strip_label_role_suffix(label):
    return LABEL_ROLE_SUFFIX_RE.sub("", label).strip()


def cell_text(cell):
    if isinstance(cell, dict):
        return cell.get("text", "") or ""
    if isinstance(cell, list):
        return " ".join(cell_text(c) for c in cell)
    return "" if cell is None else str(cell)


def validate_icc_report(wb, requested_system_type):
    """Validation 1 (ICC Format Verification) + Validation 2 (System Type Matching).

    Raises ICCValidationError with a caller-facing message on any failure; returns the
    detected internal system type ("dc" / "ac_off" / "hybrid") on success.
    """
    internal_type = SYSTEM_TYPE_TO_INTERNAL.get(requested_system_type)
    if requested_system_type not in SYSTEM_TYPE_TO_INTERNAL:
        raise ICCValidationError(
            f"Unknown System Type '{requested_system_type}'. "
            f"Expected one of {list(SYSTEM_TYPE_TO_INTERNAL)}."
        )
    if internal_type is None:
        raise ICCValidationError(
            f"System Type '{requested_system_type}' is not yet supported - no approved ICC "
            f"template is available for it."
        )

    missing_sheets = [s for s in REQUIRED_SHEETS if s not in wb.sheetnames]
    if missing_sheets:
        raise ICCValidationError(
            f"Uploaded file structure does not match the expected ICC format: "
            f"missing sheet(s) {missing_sheets}."
        )

    ws_map = wb[MAP_SHEET]
    header_row = [str(c.value).strip() if c.value is not None else "" for c in next(ws_map.iter_rows(min_row=1, max_row=1))]
    missing_columns = [c for c in REQUIRED_MAP_COLUMNS if c not in header_row]
    if missing_columns:
        raise ICCValidationError(
            f"Uploaded file structure does not match the expected ICC format: "
            f"'{MAP_SHEET}' is missing required column(s) {missing_columns}."
        )

    instances = load_data_ingestion_map(wb)
    if not instances:
        raise ICCValidationError(
            f"Uploaded file structure does not match the expected ICC format: "
            f"'{MAP_SHEET}' has no data rows."
        )

    detected_type = detect_system_type(instances)
    if detected_type != internal_type:
        raise ICCValidationError(
            f"Uploaded template does not correspond to the selected System Type. "
            f"Selected: {requested_system_type}; uploaded template matches: "
            f"{_INTERNAL_TO_SYSTEM_TYPE.get(detected_type, detected_type)}."
        )

    return detected_type


_INTERNAL_TO_SYSTEM_TYPE = {v: k for k, v in SYSTEM_TYPE_TO_INTERNAL.items() if v}


def load_data_ingestion_map(wb):
    ws = wb[MAP_SHEET]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    instances = []
    for section, label, _addr, row_no, col_no, _etype in rows:
        if not section or row_no is None or col_no is None:
            continue
        base_label = strip_label_role_suffix(label)
        if (
            instances
            and instances[-1]["section"] == section
            and instances[-1]["label"] == base_label
            and instances[-1]["cells"][-1][0] == row_no
        ):
            instances[-1]["cells"].append((row_no, col_no))
        else:
            instances.append({"section": section, "label": base_label, "cells": [(row_no, col_no)]})
    return instances


def detect_system_type(instances):
    """See icc_report_to_json.py's detect_system_type() for the full rationale: phase
    (single/three) is not detected because it never changes which Excel cell maps to which
    key for any of the supported formats.

    4 formats now, told apart purely from Data_Ingestion_Map's own Section/Field Label text:
      - no "Bill of materials (For RMS)" section at all, AND its solar BOM section is named
        exactly "Bill Of Material" (no "(For Solar System)" qualifier)  -> ac_on_grid
      - no RMS section otherwise                                        -> dc
      - "Solar Hybrid PCU" or a "Net Meter" BOM row                     -> hybrid
      - otherwise (has RMS)                                             -> ac_off
    """
    labels_lower = [inst["label"].lower() for inst in instances]

    has_rms = any(inst["section"] == "Bill of materials (For RMS)" for inst in instances)
    if not has_rms:
        is_on_grid = any(inst["section"] == "Bill Of Material" for inst in instances)
        return "ac_on_grid" if is_on_grid else "dc"

    is_hybrid = any("solar hybrid pcu" in l or l.strip() == "net meter" for l in labels_lower)
    return "hybrid" if is_hybrid else "ac_off"


def extract_section_rows(template, header_text):
    content = template["config"]["content"]
    rows = []
    collecting = False
    for node in content:
        if isinstance(node, dict) and node.get("style") == "header" and "text" in node:
            collecting = node["text"].strip() == header_text
        elif collecting and isinstance(node, dict) and "table" in node:
            rows.extend(node["table"].get("body", []))
    return rows


def extract_bom_entries(rows):
    entries = []
    for row in rows:
        texts = [cell_text(c).strip() for c in row]
        if len(texts) < 3 or not SERIAL_RE.match(texts[0]):
            continue
        label = texts[1]
        value_keys = [PLACEHOLDER_RE.findall(t) for t in texts[2:]]
        entries.append((label, value_keys))
    return entries


def run_length_group(items, key_fn):
    runs = []
    for item in items:
        k = key_fn(item)
        if runs and runs[-1][0] == k:
            runs[-1][1].append(item)
        else:
            runs.append((k, [item]))
    return runs


def align_bom_rows(template_entries, excel_instances, section, unmatched):
    """
    Shared row-alignment between a BOM template's rows and the uploaded Excel's matching
    instances - yields (label, value_keys, excel_cells) per aligned row, where value_keys is
    the per-column list of placeholder keys (see extract_bom_entries) and excel_cells is the
    per-column list of (row_no, col_no) cells that row landed on in the uploaded workbook.

    Used by both zip_section() (build the {{field_name}}: value JSON) and
    validate_bom_editable_fields_filled() (Validation 3: Make/Capacity/Quantity aren't blank).
    """
    template_key_fn = lambda e: TEMPLATE_LABEL_MERGE.get(normalize_label(e[0]), normalize_label(e[0]))
    template_runs = run_length_group(template_entries, key_fn=template_key_fn)
    excel_runs = run_length_group(excel_instances, key_fn=lambda e: normalize_label(e["label"]))

    for (_t_label, t_items), (_e_label, e_items) in zip(template_runs, excel_runs):
        for i, (label, value_keys) in enumerate(t_items):
            excel_inst = e_items[i] if i < len(e_items) else e_items[-1]
            yield label, value_keys, excel_inst["cells"]
        for extra in e_items[len(t_items):]:
            unmatched.append((section, extra["label"]))


def zip_section(template_entries, excel_instances, section, unmatched):
    output = {}
    for _label, value_keys, excel_cells in align_bom_rows(template_entries, excel_instances, section, unmatched):
        for col_idx, keys in enumerate(value_keys):
            if col_idx >= len(excel_cells) or not keys:
                continue
            row_no, col_no = excel_cells[col_idx]
            for key in keys:
                output.setdefault(key, []).append((row_no, col_no))
    return output


# Placeholder key suffixes that identify an editable BOM table cell's role - every template
# uses these exact suffixes consistently (confirmed against all 4 bundled templates).
BOM_ROLE_SUFFIXES = OrderedDict([
    ("_make", "Make"),
    ("_capacity", "Capacity"),
    ("_qty", "Quantity"),
])


def _bom_role_for_keys(keys):
    """Returns the display role name ('Make'/'Capacity'/'Quantity') for a template column's
    placeholder key(s), or None if none of them are a required BOM role (e.g. "_remarks",
    "_description" alone). A column can carry more than one key (e.g. Load Wiring's combined
    "Description & Make" column has both a "_description" and a "_make" key) - the column
    counts as a "Make" column if any of its keys end with "_make"."""
    for key in keys:
        for suffix, role in BOM_ROLE_SUFFIXES.items():
            if key.endswith(suffix):
                return role
    return None


def _is_number_not_text(value):
    """True only if openpyxl handed us an actual numeric cell value - not a string, even one
    that merely looks numeric (e.g. "5"). bool is excluded since Python's bool is technically
    an int subclass but isn't a meaningful Quantity value."""
    return isinstance(value, (int, float)) and not isinstance(value, bool)


_BOM_ROLE_ORDER = list(BOM_ROLE_SUFFIXES.values())  # ["Make", "Capacity", "Quantity"]


NUMERIC_BOM_ROLES = ("Capacity", "Quantity")


def validate_bom_editable_fields_filled(wb, detected_type):
    """
    Validation 3 (BOM completeness): every row of the editable BOM tables (the "Bill Of
    Material..." sections listed in SECTION_TEMPLATE_HEADERS - Solar System / Luminaries &
    Fans / Load Wiring) must have its Make, Capacity and Quantity cells filled in, and Capacity
    and Quantity must additionally be an actual number (not text, even numeric-looking text).
    SYSTEM FUNCTIONALITY PARAMETERS (and any other non-BOM section, e.g. RMS/Header/Image/
    Annexure) is out of scope for this check, matching the same "out of scope" sections already
    excluded from template-key coverage elsewhere in this module.

    Blank fields are reported as one summary line per section with a per-role count (e.g.
    "5 Make, 5 Capacity, 5 Quantity missing") rather than one line per blank cell - an entirely
    empty upload would otherwise produce a message with hundreds of near-identical lines, which
    renders poorly on the frontend and isn't any more actionable than the count. Invalid (non-
    blank but non-numeric) Capacity/Quantity values are rare and specific, so those stay listed
    in full per-row detail, same as before.

    Raises ICCValidationError on any failure (blank and/or invalid-number); returns nothing on
    success.
    """
    ws = wb[DATA_SHEET]
    instances = load_data_ingestion_map(wb)

    template_path = os.path.join(TEMPLATE_DIR, TEMPLATE_BY_TYPE[detected_type])
    with open(template_path) as f:
        template = json.load(f)

    by_section = OrderedDict()
    for inst in instances:
        by_section.setdefault(inst["section"], []).append(inst)

    blank_counts_by_section = OrderedDict()  # section -> {role: count}
    invalid_numeric_errors = []
    unmatched = []
    for section, section_instances in by_section.items():
        header_text = SECTION_TEMPLATE_HEADERS.get(section)
        if not header_text:
            # Not an editable BOM table (SYSTEM FUNCTIONALITY PARAMETERS / RMS / Header /
            # Image / Annexure) - out of scope for this check.
            continue

        rows = extract_section_rows(template, header_text)
        template_entries = extract_bom_entries(rows)

        for label, value_keys, excel_cells in align_bom_rows(template_entries, section_instances, section, unmatched):
            for col_idx, keys in enumerate(value_keys):
                if col_idx >= len(excel_cells) or not keys:
                    continue
                role = _bom_role_for_keys(keys)
                if role is None:
                    continue
                row_no, col_no = excel_cells[col_idx]
                value = ws.cell(row=row_no, column=col_no).value
                if value is None or str(value).strip() == "":
                    section_counts = blank_counts_by_section.setdefault(section, {})
                    section_counts[role] = section_counts.get(role, 0) + 1
                elif role in NUMERIC_BOM_ROLES and not _is_number_not_text(value):
                    invalid_numeric_errors.append(
                        f"{section} > '{label}' - {role} '{value}' must be a number, not text"
                    )

    if not blank_counts_by_section and not invalid_numeric_errors:
        return

    error_lines = []
    for section, section_counts in blank_counts_by_section.items():
        parts = [f"{section_counts[role]} {role}" for role in _BOM_ROLE_ORDER if role in section_counts]
        error_lines.append(f"{section}: {', '.join(parts)} missing")
    error_lines.extend(invalid_numeric_errors)

    raise ICCValidationError(
        "The uploaded ICC report has invalid required fields in the editable BOM tables "
        "(Make/Capacity/Quantity must be filled in, and Capacity/Quantity must be a number):\n- "
        + "\n- ".join(error_lines)
    )


def convert_icc_report(wb, detected_type):
    """Convert an already-validated workbook to the flat {{field_name}}: value JSON.

    Returns (data, fallback_fields, unmatched_fields) - fallback_fields/unmatched_fields are
    diagnostics (not errors): they list which cells used a slugified key instead of a real
    template field name, and which Excel fields had no corresponding template row.
    """
    ws = wb[DATA_SHEET]
    instances = load_data_ingestion_map(wb)

    template_path = os.path.join(TEMPLATE_DIR, TEMPLATE_BY_TYPE[detected_type])
    with open(template_path) as f:
        template = json.load(f)

    sfp_map = SFP_FIELD_MAPS.get(detected_type, {})

    by_section = OrderedDict()
    for inst in instances:
        by_section.setdefault(inst["section"], []).append(inst)

    output = {}
    unmatched_excel_fields = []
    fallback_fields = []

    for section, section_instances in by_section.items():
        header_text = SECTION_TEMPLATE_HEADERS.get(section)

        if not header_text and section.strip().lower().startswith(NO_TEMPLATE_COVERAGE_PREFIXES):
            for inst in section_instances:
                for idx, (row_no, col_no) in enumerate(inst["cells"]):
                    key = make_fallback_key(section, inst["label"], idx, len(inst["cells"]))
                    output[key] = ws.cell(row=row_no, column=col_no).value
                    fallback_fields.append((section, inst["label"], key))
            continue

        if not header_text:
            # SYSTEM FUNCTIONALITY PARAMETERS (and its subsections) are out of scope for
            # conversion - always emit null rather than whatever the Data_Ingestion_Map cell
            # holds (which can be a static label, not an actual recorded value). Key resolution
            # (sfp_map vs. fallback) is left unchanged so diagnostics keep working the same way.
            for inst in section_instances:
                for idx, (row_no, col_no) in enumerate(inst["cells"]):
                    key = sfp_map.get((row_no, col_no))
                    if key is None:
                        key = make_fallback_key(section, inst["label"], idx, len(inst["cells"]))
                        fallback_fields.append((section, inst["label"], key))
                    output[key] = None
            continue

        rows = extract_section_rows(template, header_text)
        template_entries = extract_bom_entries(rows)
        key_cells = zip_section(template_entries, section_instances, section, unmatched_excel_fields)
        for key, cell_list in key_cells.items():
            row_no, col_no = cell_list[-1]
            output[key] = ws.cell(row=row_no, column=col_no).value

    return output, fallback_fields, unmatched_excel_fields


# MDMS schema listing every registered brand, grouped by asset type (PANEL/BATTERY/INVERTER).
BRAND_SCHEMA_CODE = "asset-registry.BrandSchema"

# Which converted-JSON key holds the Make value for panel/battery/inverter, per detected system
# type, and which asset_type_code (from BRAND_SCHEMA_CODE) it must be validated against. Mirrors
# FieldPlannerService.INVERTER_KEY_CANDIDATES on the Java side: DC has no dedicated inverter (the
# CCU serves that role), HYBRID's is the "hybrid PCU", AC_OFF_GRID/AC_ON_GRID call it "inverter".
# AC_ON_GRID has no battery bank at all, so it's omitted from that type's mapping.
BOM_BRAND_FIELDS_BY_TYPE = {
    "dc": {
        "solar_module_make": "PANEL",
        "solar_battery_make": "BATTERY",
        "solar_charge_controller_make": "INVERTER",
    },
    "ac_off": {
        "solar_module_make": "PANEL",
        "solar_battery_make": "BATTERY",
        "inverter_make": "INVERTER",
    },
    "hybrid": {
        "solar_module_make": "PANEL",
        "solar_battery_make": "BATTERY",
        "solar_hybrid_pcu_make": "INVERTER",
    },
    "ac_on_grid": {
        "solar_module_make": "PANEL",
        "inverter_make": "INVERTER",
    },
}


def _brand_names_by_asset_type(mdms_client, request_info):
    """Returns {asset_type_code: {registered brand names}}, e.g. {"PANEL": {"Gautam", "ReNew", ...}}.

    Returns an empty dict (never raises) if the schema can't be fetched - brand validation is then
    silently skipped, the same graceful-degradation choice already made for
    facility.FacilitySolarConfigurationRule elsewhere in this codebase, rather than blocking every
    upload on an MDMS outage.
    """
    try:
        records = mdms_client.fetch_mdms_records(request_info, BRAND_SCHEMA_CODE)
    except Exception:
        return {}

    brands = records[0].get("Brand") or [] if records else []
    names_by_asset_type = {}
    for brand in brands:
        if brand.get("active") is False:
            continue
        asset_type = brand.get("asset_type_code")
        name = brand.get("name")
        if not asset_type or not name:
            continue
        names_by_asset_type.setdefault(asset_type, set()).add(name)
    return names_by_asset_type


def validate_bom_brand_names(data, detected_type, mdms_client, request_info):
    """
    Validation 4 (brand names): the Make value recorded for the Solar Module (panel), Solar
    Battery, and the system's inverter-equivalent component (Solar Charge Controller for DC, the
    hybrid PCU for HYBRID, Inverter for AC_OFF_GRID/AC_ON_GRID) must be one of the active brand
    names registered in MDMS asset-registry.BrandSchema for that component's asset type - e.g.
    Solar Battery's Make must be "NED" or "Coslight", not an arbitrary string.

    Raises ICCValidationError listing every invalid Make value found; returns nothing on success
    (including when the brand schema itself can't be fetched, or has no entries for a given
    asset type - this check simply doesn't block in either case).
    """
    fields_by_role = BOM_BRAND_FIELDS_BY_TYPE.get(detected_type, {})
    if not fields_by_role:
        return

    names_by_asset_type = _brand_names_by_asset_type(mdms_client, request_info)
    if not names_by_asset_type:
        return

    errors = []
    for key, asset_type in fields_by_role.items():
        value = data.get(key)
        if value is None or str(value).strip() == "":
            continue  # blank Make is already reported by Validation 3

        allowed_names = names_by_asset_type.get(asset_type)
        if not allowed_names:
            continue  # no MDMS brands registered for this asset type - nothing to check against

        value_str = str(value).strip()
        if value_str not in allowed_names:
            errors.append(
                f"'{value_str}' is not a recognized {asset_type.title()} brand - expected one of: "
                + ", ".join(sorted(allowed_names))
            )

    if errors:
        raise ICCValidationError(
            "The uploaded ICC report has invalid brand names in the editable BOM tables:\n- "
            + "\n- ".join(errors)
        )


# MDMS schema listing every registered capacity option, grouped by asset type (BATTERY/PANEL/
# INVERTER) AND by systemType (the "system" field there uses the same phase-specific codes as
# facility.SystemType, e.g. "AC_OFF_GRID_THREE_PHASE" - not the internal dc/ac_off/hybrid/
# ac_on_grid keys used elsewhere in this module).
ASSET_TYPE_SCHEMA_CODE = "asset-registry.AssetTypeSchema"

_CAPACITY_MATCH_TOLERANCE = 0.001


def _capacity_options_by_asset_type(mdms_client, request_info, requested_system_type):
    """Returns {asset_type_code: {allowed capacity option strings}} for one systemType, read from
    MDMS asset-registry.AssetTypeSchema's per-asset-type form_fields where key == "capacity" and
    system == requested_system_type. Never raises - returns {} on any fetch/shape problem, same
    graceful-degradation choice as _brand_names_by_asset_type."""
    try:
        records = mdms_client.fetch_mdms_records(request_info, ASSET_TYPE_SCHEMA_CODE)
    except Exception:
        return {}

    asset_types = records[0].get("AssetType") or [] if records else []
    options_by_asset_type = {}
    for asset_type in asset_types:
        if asset_type.get("active") is False:
            continue
        code = asset_type.get("code")
        if not code:
            continue
        for field in asset_type.get("form_fields") or []:
            if field.get("key") != "capacity" or field.get("system") != requested_system_type:
                continue
            options = field.get("options") or []
            options_by_asset_type.setdefault(code, set()).update(str(o).strip() for o in options)
    return options_by_asset_type


def _capacity_matches_any_option(value, allowed_options):
    """Numeric-tolerant match: "550" and "550.0" (or a real float 550.0 from openpyxl) are the
    same capacity, so compare as numbers rather than exact text - same reasoning as the "45" vs
    "45.0" totalCapacity fix on the field-planner side."""
    value_str = str(value).strip()
    if value_str in allowed_options:
        return True
    try:
        value_num = float(value_str)
    except (TypeError, ValueError):
        return False
    for option in allowed_options:
        try:
            if math.isclose(value_num, float(option), rel_tol=0, abs_tol=_CAPACITY_MATCH_TOLERANCE):
                return True
        except (TypeError, ValueError):
            continue
    return False


def validate_bom_capacity_options(data, detected_type, requested_system_type, mdms_client, request_info):
    """
    Validation 5 (capacity options): the Capacity value recorded for the Solar Module (panel),
    Solar Battery, and the system's inverter-equivalent component must be one of the allowed
    values registered in MDMS asset-registry.AssetTypeSchema for that component's asset type and
    the requested systemType - e.g. Solar Battery's Capacity for AC_OFF_GRID must be one of
    "125", "150", "180", "200", "220".

    Reuses BOM_BRAND_FIELDS_BY_TYPE's make-field -> asset_type mapping (same 3 components), just
    deriving each field's paired "_capacity" key from its "_make" key.

    Raises ICCValidationError listing every invalid Capacity value found; returns nothing on
    success (including when the schema can't be fetched, or has no options registered for this
    systemType/asset type combination - this check simply doesn't block in either case).
    """
    fields_by_role = BOM_BRAND_FIELDS_BY_TYPE.get(detected_type, {})
    if not fields_by_role:
        return

    options_by_asset_type = _capacity_options_by_asset_type(mdms_client, request_info, requested_system_type)
    if not options_by_asset_type:
        return

    errors = []
    for make_key, asset_type in fields_by_role.items():
        capacity_key = make_key.replace("_make", "_capacity")
        value = data.get(capacity_key)
        if value is None or str(value).strip() == "":
            continue  # blank Capacity is already reported by Validation 3

        allowed_options = options_by_asset_type.get(asset_type)
        if not allowed_options:
            continue  # no MDMS options registered for this asset type/systemType - nothing to check against

        if not _capacity_matches_any_option(value, allowed_options):
            errors.append(
                f"'{value}' is not an allowed {asset_type.title()} capacity for systemType "
                f"'{requested_system_type}' - expected one of: "
                + ", ".join(sorted(allowed_options, key=lambda o: (len(o), o)))
            )

    if errors:
        raise ICCValidationError(
            "The uploaded ICC report has invalid capacity values in the editable BOM tables:\n- "
            + "\n- ".join(errors)
        )


def validate_and_convert(xlsx_path, requested_system_type, mdms_client=None, request_info=None):
    """Single entry point for the /icc-reports endpoint: raises ICCValidationError on any
    Validation 1/2/3/4/5 failure (nothing is saved in that case); otherwise returns
    (detected_type, data, fallback_fields, unmatched_fields).

    mdms_client/request_info are optional: pass both to also run Validations 4/5 (brand names,
    capacity options); omit either to skip them (e.g. for callers/tests without MDMS access).
    """
    try:
        wb = load_workbook(xlsx_path, data_only=True)
    except InvalidFileException as exc:
        raise ICCValidationError(f"Uploaded file is not a valid Excel (.xlsx) file: {exc}") from exc

    detected_type = validate_icc_report(wb, requested_system_type)
    validate_bom_editable_fields_filled(wb, detected_type)
    data, fallback_fields, unmatched_fields = convert_icc_report(wb, detected_type)
    if mdms_client is not None and request_info is not None:
        validate_bom_brand_names(data, detected_type, mdms_client, request_info)
        validate_bom_capacity_options(data, detected_type, requested_system_type, mdms_client, request_info)
    return detected_type, data, fallback_fields, unmatched_fields
