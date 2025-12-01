from typing import Dict, List, Union, Any, Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Protection, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

def add_dropdowns_to_excel(
        file_path: str,
        sheet_name: str,
        dropdowns: Dict[str, List[str]],
        allow_blank_map: Optional[Dict[str, bool]],
        max_extra_rows: int = 1000
):
    wb = load_workbook(file_path)
    ws = wb[sheet_name]
    header_row = 1
    max_row = ws.max_row + max_extra_rows  # extend range

    for column_header, options in dropdowns.items():
        if not options:
            continue
        options_str = ",".join(options)
        allow_blank = (allow_blank_map or {}).get(column_header, True)
        dv = DataValidation(type="list",showErrorMessage=True, formula1=f'"{options_str}"', allow_blank=allow_blank)
        dv.error = 'Please select from the list'
        dv.errorTitle = 'Invalid Entry'
        for cell in ws[header_row]:
            if cell.value == column_header:
                col_letter = cell.column_letter
                dv.add(f"{col_letter}2:{col_letter}{max_row}")
                break
        ws.add_data_validation(dv)

    wb.save(file_path)


def lock_excel_columns(
        file_path: str,
        sheet_name: str,
        column_headers_to_unlock: List[Union[str, int]]
) -> None:
    """
    Locks the entire Excel sheet except the specified columns by header name or index.

    Args:
        file_path: Path to the Excel file.
        sheet_name: Name of the sheet to modify.
        column_headers_to_unlock: List of column headers (e.g., 'Selection?') or 1-based column indices to keep editable.
    """
    wb = load_workbook(file_path)
    ws = wb[sheet_name]

    # Find column indices to unlock based on header names or 1-based indices
    column_indices_to_unlock = set()
    for identifier in column_headers_to_unlock:
        if isinstance(identifier, str):
            for col_index, cell in enumerate(ws[1], 1):  # Header is assumed to be in the first row
                if cell.value and str(cell.value).strip() == identifier.strip():
                    column_indices_to_unlock.add(col_index)
                    break
        elif isinstance(identifier, int):
            column_indices_to_unlock.add(identifier)

    max_rows = ws.max_row

    # Lock all cells by default
    for row in range(1, max_rows + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).protection = Protection(locked=True)

    # Unlock only the specified columns
    for col_index in column_indices_to_unlock:
        for row in range(1, max_rows + 1):
            ws.cell(row=row, column=col_index).protection = Protection(locked=False)

    # Enable worksheet protection and allow selection of unlocked cells
    ws.protection.sheet = True
    ws.protection.formatColumns = True
    wb.save(file_path)


def lock_prefilled_rows_in_excel(
    file_path: str,
    sheet_name: str,
    editable_columns: list,
    total_rows: int,
    total_columns: int,
    always_locked_columns: list = None,  # new parameter
    extra_append_rows: int = 1000
):
    always_locked_columns = always_locked_columns or []

    wb = load_workbook(file_path)
    ws = wb[sheet_name]

    grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    no_fill = PatternFill()  # reset

    # Unlock all cells first
    for row in ws.iter_rows():
        for cell in row:
            cell.protection = Protection(locked=False)
            cell.fill = no_fill

    # Get header row values
    header_row = [cell.value for cell in ws[1]]

    # Editable column indices
    editable_indices = [
        i + 1 for i, col in enumerate(header_row)
        if col and any(c in col for c in editable_columns)
    ]

    # Always locked column indices
    always_locked_indices = [
        i + 1 for i, col in enumerate(header_row)
        if col and any(c in col for c in always_locked_columns)
    ]

    # Lock prefilled rows completely (grey out non-editable cells)
    for row_idx in range(2, total_rows + 2):
        for col_idx in range(1, total_columns + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            # If column is editable -> unlock
            if col_idx in editable_indices:
                cell.protection = Protection(locked=False)
                cell.fill = no_fill
            # If column is always locked -> lock + grey
            elif col_idx in always_locked_indices:
                cell.protection = Protection(locked=True)
                cell.fill = grey_fill
            else:
                # other prefilled cells -> lock + grey
                cell.protection = Protection(locked=True)
                cell.fill = grey_fill

    # Leave appendable rows fully unlocked, but respect always_locked columns
    # Only apply gray fill to cells that have values, not empty cells
    for row_idx in range(total_rows + 2, total_rows + extra_append_rows + 2):
        for col_idx in range(1, total_columns + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell_value = cell.value
            is_empty = cell_value is None or (isinstance(cell_value, str) and cell_value.strip() == "")

            if col_idx in always_locked_indices:
                cell.protection = Protection(locked=True)
                # Only apply gray fill if cell has a value, otherwise leave it white
                if not is_empty:
                    cell.fill = grey_fill
                else:
                    cell.fill = no_fill
            else:
                cell.protection = Protection(locked=False)
                cell.fill = no_fill

    # Enable protection
    ws.protection.select_unlocked_cells = True
    ws.protection.formatColumns = True
    ws.protection.insertRows = True
    ws.protection.sheet = True
    ws.protection.enable()

    wb.save(file_path)


def add_validations_to_excel(file_path: str,
                             sheet_name: str,
                             validations: Dict[str, Dict[str, str]],
                             allow_blank_map: Dict[str, bool],
                             max_extra_rows: int = 1000
                             ):
    """
    Adds Excel data validation (pattern + uniqueness) to specified columns.
    :param file_path: Excel file path
    :param sheet_name: Sheet where validation needs to be added
    :param validations: Dict[column_name] = {"type": "regex"/"unique", "pattern"/"message": str}
    :param allow_blank_map: Dict[str,bool]
    :param max_extra_rows: Maximum extra number of rows to allow
    """
    wb = load_workbook(file_path)
    ws = wb[sheet_name]

    header_row = 1
    max_row = ws.max_row + max_extra_rows  # allow future rows for data entry
    header_cells = {cell.value.strip(): cell for cell in ws[header_row] if cell.value}

    for col_name, config in validations.items():
        header_cell = header_cells.get(col_name)
        if not header_cell:
            continue

        col_idx = header_cell.column
        col_letter = get_column_letter(col_idx)
        data_range = f"{col_letter}2:{col_letter}{max_row}"
        allow_blank = allow_blank_map.get(col_name, True)

        if config["type"] == "regex":
            pattern = config["pattern"]
            # Handle known simple patterns
            if pattern == "^\\d{10}$":  # 10-digit number validation
                formula = f'AND(ISNUMBER({col_letter}2),LEN({col_letter}2)=10)'
                dv = DataValidation(type="custom", formula1=formula,
                                    showErrorMessage=True, error=config.get("message", "Invalid value"),allow_blank=allow_blank)
                ws.add_data_validation(dv)
                dv.add(data_range)
            else:
                # For unknown/complex regex, just add a comment instead of hard validation
                for cell in ws[data_range]:
                    for c in cell:
                        c.comment = Comment(f"Validation: {config['message']}", "System")

        elif config["type"] == "unique":
            # Enforce uniqueness using COUNTIF formula
            if allow_blank:
                formula = f'OR(LEN({col_letter}2)=0, COUNTIF(${col_letter}:${col_letter},{col_letter}2)=1)'
            else:
                formula = f'COUNTIF(${col_letter}:${col_letter},{col_letter}2)=1'
            dv_unique = DataValidation(type="custom", formula1=formula,
                                       showErrorMessage=True, error=config.get("message", "Must be unique"),allow_blank=allow_blank)
            ws.add_data_validation(dv_unique)
            dv_unique.add(data_range)

    wb.save(file_path)


def autofit_columns(
    file_path: str,
    sheet_name: str,
    auto_fit: bool = True,
    default_width: int = 20,
    max_width: int = 40,
    enable_wrap_text: bool = True
) -> None:
    """
    Adjust column widths in a given Excel sheet and optionally apply wrap text.

    Args:
        file_path: Path to the Excel file
        sheet_name: Name of the sheet to adjust
        auto_fit: If True, width is based on longest text length in each column
        default_width: Default width if auto_fit is False
        max_width: Maximum allowed column width
        enable_wrap_text: If True, applies wrap text to all cells
    """
    wb = load_workbook(file_path)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in {file_path}")

    ws = wb[sheet_name]

    # Load data with pandas for convenience
    try:
        df = pd.read_excel(file_path, sheet_name=sheet_name)
    except Exception:
        df = None  # fallback: will just use worksheet cells

    for i, col in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row), start=1):
        col_letter = get_column_letter(i)

        if auto_fit:
            if df is not None and df.shape[1] >= i:
                texts = [str(df.columns[i - 1])] + df.iloc[:, i - 1].astype(str).tolist()
            else:
                texts = [str(cell.value) for cell in col if cell.value is not None]

            max_length = max((len(str(t)) for t in texts if t), default=default_width)
            ws.column_dimensions[col_letter].width = min(max_length + 2, max_width)  # padding
        else:
            ws.column_dimensions[col_letter].width = default_width

        # Apply wrap text to all cells in this column
        if enable_wrap_text:
            for cell in col:
                cell.alignment = Alignment(wrap_text=True)

    wb.save(file_path)


def add_non_blank_validations_to_file(
    file_path: str,
    sheet_name: str,
    facility_schema: List[Dict[str, Any]],
    allow_blank_map: Dict[str, bool]
) -> None:
    """
    Add non-blank validations to an Excel sheet given file path and sheet name.

    :param file_path: Path to the Excel file
    :param sheet_name: Name of the worksheet where validations should be applied
    :param facility_schema: List of dicts with column definitions
    :param allow_blank_map: Dict {header_name: bool} to override allow_blank
    """
    try:
        wb = load_workbook(file_path)
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in {file_path}")

        ws = wb[sheet_name]

        for col_idx, col in enumerate(facility_schema, start=1):
            # Match how headers are generated
            mandatory_indicator = "(Mandatory)" if col.get("required") else ""
            header_name = f"{col.get('name')} {mandatory_indicator}".strip()

            required = col.get("required", False)

            # Skip if regex/dropdown/unique present
            has_rule = any(k in col for k in ("regex", "unique", "dropdown"))

            if required and not has_rule:
                # Default allow_blank = False for required, unless overridden
                allow_blank = allow_blank_map.get(header_name, not required)
                col_letter = get_column_letter(col_idx)

                dv = DataValidation(
                    type="custom",
                    formula1=f'LEN(TRIM({col_letter}2))>0',
                    allow_blank=allow_blank,
                    showErrorMessage=True,
                    error="This field cannot be left blank"
                )
                ws.add_data_validation(dv)
                dv.add(f"{col_letter}2:{col_letter}1048576")

        wb.save(file_path)

    except Exception as e:
        raise